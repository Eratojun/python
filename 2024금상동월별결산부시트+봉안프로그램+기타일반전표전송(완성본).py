import pandas as pd			
from pathlib import Path
import openpyxl	
import pandas as pd

from openpyxl import load_workbook

import pandas as pd			
from pathlib import Path
import openpyxl	
from datetime import datetime
import pandas as pd			
from pathlib import Path
import openpyxl	
from openpyxl.styles import NamedStyle

from openpyxl.styles import PatternFill #엑셀셀을 빨간색으로 표시



# 여러 개의 데이터프레임을 엑셀 파일로 내보내기 함수
def export_multiple_dfs_to_excel(dfs, sheet_names, file_name):
    """
    여러 개의 데이터프레임을 각각 다른 시트에 엑셀 파일로 내보내는 함수.
    : param dfs: 데이터프레임 리스트 
    : param sheet_names: 시트 이름 리스트 
    : param file_name: 저장할 엑셀 파일 이름
   """ 
    
    if len(dfs) != len(sheet_names) :

      raise ValueError("데이터프레임 리스트와 시트 이름 리스트의 길이가 같아야 합니다. ")


    with pd. ExcelWriter (file_name) as writer:
     for df, sheet_name in zip(dfs, sheet_names) :
        df. to_excel(writer, sheet_name=sheet_name,index=False)
    #Print(f"{file_name} 파일이 생성되었습니다.")






input_folder='C:\dev\myproj01\금상동2024'
raw_data_dir =Path(input_folder)
excel_files=raw_data_dir.glob('24년*')





total_df=pd.DataFrame() #빈 데이터프래임 생성
total_df2=pd.DataFrame() #빈 데이터프래임 생성
total_df3=pd.DataFrame() #빈 데이터프래임 생성
total_df4=pd.DataFrame() #빈 데이터프래임 생성
total_df5=pd.DataFrame() #빈 데이터프래임 생성











i=0
for excel_file in excel_files:
    i=i+1
   
    
    
    
    df =pd.read_excel(excel_file,sheet_name="사용선수금",skiprows=3,usecols="a:g")
    df.columns = ['날짜', '성명', '봉안함', '적요1', '차변', '대변', '잔액']
    #df['날짜'] = pd.to_datetime(df['날짜']).dt.strftime('%Y-%m-%d')
    
    df2 =pd.read_excel(excel_file,sheet_name="관리선수금",skiprows=3,usecols="a:g")
    df2.columns = ['날짜', '성명', '봉안함', '적요1', '차변', '대변', '잔액']
    #f2['날짜'] = pd.to_datetime(df2['날짜']).dt.strftime('%Y-%m-%d')
    
    df3 =pd.read_excel(excel_file,sheet_name="사용료수입",skiprows=3,usecols="a:g")
    df3.columns = ['날짜', '성명', '봉안함', '적요1', '차변', '대변', '잔액']
    #df3['날짜'] = pd.to_datetime(df3['날짜']).dt.strftime('%Y-%m-%d')
    
    try:
     # 엑셀 파일의 시트 목록을 가져옴
        xls = pd.ExcelFile(excel_file)
        print(f"'{excel_file}' 파일의 시트 목록: {xls.sheet_names}")
    
        # '회전관선' 시트가 있는지 확인
        if "회전관선" in xls.sheet_names:
          print(f"'{excel_file}' 파일에서 '회전관선' 시트를 읽습니다.")
          # 시트를 읽어 데이터프레임으로 저장
          df4 = pd.read_excel(excel_file, sheet_name="회전관선", skiprows=3, usecols="A:G")
          df4.columns = ['날짜', '성명', '봉안함', '적요1', '차변', '대변', '잔액']
          total_df4=total_df4._append(df4,ignore_index=True)
          print(df4.head())  # 데이터프레임 출력
        else:
          print(f"'{excel_file}' 파일에 '회전관선' 시트가 없습니다. 다른 작업을 진행합니다.")
    
    except ValueError as e:
      print(f"오류 발생: {e}")
    except Exception as e:
      print(f"파일 '{excel_file}' 처리 중 오류 발생: {e}")

    print("프로그램이 종료되었습니다.")
    
       
    
    
    
    
    
    df4 = pd.read_excel(excel_file, sheet_name="회전관선-최준철세무사", skiprows=3, usecols="A:G")
    df4.columns = ['날짜', '성명', '봉안함', '적요1', '차변', '대변', '잔액']

    



    df5= pd.read_excel(excel_file,sheet_name=f"24-{i}",skiprows=2,usecols="a:i")
     
    
    
    
    
    
    
    total_df=total_df._append(df,ignore_index=True)	

    total_df2=total_df2._append(df2,ignore_index=True)

    total_df3=total_df3._append(df3,ignore_index=True)

    total_df4=total_df4._append(df4,ignore_index=True)

    total_df5=total_df5._append(df5,ignore_index=True)
    
   
    




total_df2['날짜'] = pd.to_datetime(total_df2['날짜'], errors='coerce')
total_df2['날짜'] = total_df2['날짜'].dt.strftime('%Y-%m-%d')

total_df3['날짜'] = pd.to_datetime(total_df3['날짜'], errors='coerce')
total_df3['날짜'] = total_df3['날짜'].dt.strftime('%Y-%m-%d')

total_df4['날짜'] = pd.to_datetime(total_df4['날짜'], errors='coerce')
total_df4['날짜'] = total_df4['날짜'].dt.strftime('%Y-%m-%d')


total_df5= total_df5.dropna(subset=["회계계정"]) #아래 누적액,월계액을 없애기위해 회계계정없는부분을 삭제함

total_df5['월일']=total_df5['월일'].dt.strftime('%Y-%m-%d') #구매자결제일이 날짜함수이므로 문자함수로 변환

total_df5['월일']=total_df5['월일'].str.replace('-','')

total_df5.info()
#total_df5['월일']=int(total_df5['월일'])




total_df5.columns = total_df5.columns.str.strip()
total_df5['월일']=total_df5['월일'].fillna(method='ffill') # 최근일 월일없는행은 전일자로 처리

#total_df5=total_df5['적요'].fillna(method='ffill') # 최근일 월일없는행은 전일자로 처리
total_df5['적       요']=total_df5['적       요'].fillna(method='ffill') # 최근일 월일없는행은 전일자로 처리















total_df=total_df.sort_values(["봉안함"]) # 봉안함번호로정렬
total_df2=total_df2.sort_values(["봉안함"]) # 봉안함번호로정렬
total_df3=total_df3.sort_values(["봉안함"]) # 봉안함번호로정렬
total_df4=total_df4.sort_values(["봉안함"]) # 봉안함번호로정렬







total_df= total_df.dropna(subset=["봉안함"]) #아래 봉안함번로 없는 행을 없애기위해 봉안함없는 부분을 삭제함
total_df2= total_df2.dropna(subset=["봉안함"]) #아래 봉안함번로 없는 행을 없애기위해 봉안함없는 부분을 삭제함
total_df3= total_df3.dropna(subset=["봉안함"]) #아래 봉안함번로 없는 행을 없애기위해 봉안함없는 부분을 삭제함
total_df4= total_df4.dropna(subset=["봉안함"]) #아래 봉안함번로 없는 행을 없애기위해 봉안함없는 부분을 삭제함
















total_df.columns = ['날짜','성명','봉안함','적요1','차변','대변','잔액']
total_df2.columns = ['날짜','성명','봉안함','적요1','차변','대변','잔액']
total_df3.columns = ['날짜','성명','봉안함','적요1','차변','대변','잔액']
total_df4.columns = ['날짜', '성명','봉안함','적요1','차변','대변','잔액']
    


# 날짜 열을 먼저 datetime 형식으로 변환
total_df['날짜'] = pd.to_datetime(total_df['날짜'], errors='coerce')
total_df['날짜'] = total_df['날짜'].dt.strftime('%Y-%m-%d')

total_df2['날짜'] = pd.to_datetime(total_df2['날짜'], errors='coerce')
total_df2['날짜'] = total_df2['날짜'].dt.strftime('%Y-%m-%d')

total_df3['날짜'] = pd.to_datetime(total_df3['날짜'], errors='coerce')
total_df3['날짜'] = total_df3['날짜'].dt.strftime('%Y-%m-%d')

total_df4['날짜'] = pd.to_datetime(total_df4['날짜'], errors='coerce')
total_df4['날짜'] = total_df4['날짜'].dt.strftime('%Y-%m-%d')


total_df['날짜']=total_df['날짜'].fillna(method='ffill') # 최근일 월일없는행은 전일자로 처리
total_df=total_df.sort_values(["봉안함"]) # 일자와 환율을 일자와 환율로 정렬
total_df6=total_df.sort_values(["날짜"]) # 일자와 환율을 일자와 환율로 정렬

    
total_df2['날짜']=total_df2['날짜'].fillna(method='ffill') # 최근일 월일없는행은 전일자로 처리
total_df2=total_df2.sort_values(["봉안함"]) # 일자와 환율을 일자와 환율로 정렬
total_df7=total_df2.sort_values(["날짜"]) # 일자와 환율을 일자와 환율로 정렬


total_df3['날짜']=total_df3['날짜'].fillna(method='ffill') # 최근일 월일없는행은 전일자로 처리
total_df3=total_df3.sort_values(["봉안함"]) # 일자와 환율을 일자와 환율로 정렬
total_df8=total_df3.sort_values(["날짜"]) # 일자와 환율을 일자와 환율로 정렬



total_df4['날짜']=total_df4['날짜'].fillna(method='ffill') # 최근일 월일없는행은 전일자로 처리
total_df4=total_df4.sort_values(["봉안함"]) # 일자와 환율을 일자와 환율로 정렬
total_df9=total_df4.sort_values(["날짜"]) # 일자와 환율을 일자와 환율로 정렬


#total_df = pd.concat([total_df, df1])
total_df.to_excel("C:\dev\myproj01\금상동\사용선수금2024.xlsx")
total_df2.to_excel("C:\dev\myproj01\금상동\관리선수금2024.xlsx")
total_df3.to_excel("C:\dev\myproj01\금상동\사용료수입2024.xlsx")
total_df4.to_excel("C:\dev\myproj01\금상동\회전관리선수금2024.xlsx")
total_df5.to_excel("C:\dev\myproj01\금상동\결과폴더\월별결산부2024.xlsx") 




#######################################################################################
# 엑셀 파일 읽어오기
#mon_df = pd.read_excel('C:\dev\myproj01\금상동\결과폴더\월별결산부2024.xlsx')  # 메인 엑셀 파일
mon_df = total_df5
mon_df['수입'] = mon_df['수입'].fillna(0)
mon_df['지출'] = mon_df['지출'].fillna(0)

# '구분1', '수입', '지출' 열에 NaN 값이 있는 행 제거
#mon_df = mon_df.dropna(subset=['구분1', '수입', '지출'])
mon_df['금액2']=mon_df['금액2']=0


# 금액 계산을 위한 함수 정의
mon_df['지출']=mon_df['지출'].astype(int)
mon_df['수입']=mon_df['수입'].astype(int)









mon_df['코드1'] = None
mon_df['구분1']=None
mon_df['구분2']=None
mon_df.info()



#mon_df.info()

# 두 문자열 칼럼을 합쳐서 새로운 칼럼에 저장
#mon_df['적요명'] = mon_df['적       요'].astype(str) + mon_df['봉안실'].astype(str)

import pandas as pd

# 1. 엑셀 파일에서 회계 데이터를 불러오기
#mon_df = pd.read_excel('accounting_data.xlsx')





account_code_df = pd.read_excel('C:\dev\myproj01\금상동\결과폴더\금상동계정코드(2024-11-21).xlsx')  # 계정코드 엑셀 파일

account_code_df.info()

# '코드1'이라는 새로운 열 추가

# 조건에 따른 값을 계산하는 함수 정의
def calculate_gubun(row):
    E = str(row['납부방법'])  # E3608
    D = str(row['계정과목']) # D3608
    G =str(row['회계계정'])  # G3608
    H = row['수입']  # H3608
    I = row['지출']  # I3608

    if  ("이체" in E and "관리수입" in G and H < 100000) : #관리수입중 5만이상인 경우만 인식하기
        return 0 
    elif "현금" in E and H > 0:
        return 2
    elif ("대체" in E or "이체" in E) and "해약" in D and I > 0:
        return 4
    elif (
        ("이체" in E and H > 0) or
        ("이체" in E and "일반관리선수" in G and I > 0) or
        ("이체" in E and "회전관리" in G and I > 0) or
        ("이체" in E and "사용수입" in G and I > 0) or
       
        ("대체" in E and "사용수입" in G and H > 0) or
        ("대체" in E and "관리선수" in G and H > 0) or
        ("대체" in E and "관리선수" in G and I > 0) or
        ("대체" in E and "사용수입" in G and H > 0) or
        ("신용" in E and H > 0)
    ):
        return 4
    else:
        return 0

# 각 행에 대해 함수 적용
mon_df['구분1'] = mon_df.apply(calculate_gubun, axis=1)

mon_df['구분1']=mon_df['구분1'].astype(float)



mon_df['구분1'] = mon_df['구분1'].fillna(0)







# 인덱스를 재설정
mon_df = mon_df.reset_index(drop=True)





# 조건에 따른 값 설정
for index, row in mon_df.iterrows():
    # 마지막 행 처리 방지
    if index + 1 >= len(mon_df):
        continue
    
    # 첫번째 조건
    if "해약" in str(row['계정과목']) and row['납부방법'] == "이체" and (row['회계계정'] in ["관리선수", "사용선수"]) and row['지출'] > 0:
        mon_df.at[index, '코드1'] = 103
    # 두번째 조건
    elif "해약" in str(row['계정과목']) and row['회계계정'] == "사용수입" and row['지출'] > 0:
        mon_df.at[index, '코드1'] = 420
    # 세번째 조건
    elif "관리비" in str(row['계정과목']) and "관리수입" in str(row['회계계정']) and "일반관리선수" in str(mon_df.at[index+1, '회계계정']) and row['수입'] > 0:
        mon_df.at[index, '코드1'] = 274
    # 네번째 조건
    elif "관리비" in str(row['계정과목']) and "회전관리선수" in str(mon_df.iloc[index+1]['회계계정']) and row['수입'] > 0:
        mon_df.at[index, '코드1'] = 275
    # 다섯번째 조건
    elif "대체" in str(row['납부방법']) and "관리선수" in str(row['회계계정']) and row['지출'] > 0:
        mon_df.at[index, '코드1'] = 421
    # 여섯번째 조건
    elif "이체" in str(row['납부방법']) and row['회계계정'] == "관리수입":
        mon_df.at[index, '코드1'] = 274
    # 일곱번째 조건
    elif row['수입'] == 0:
        mon_df.at[index, '코드1'] = 0
    # VLOOKUP 유사 작업
    else:
        lookup_value = str(row['회계계정'])[:6]  # G열의 첫 6글자 추출
        code_row = account_code_df[account_code_df['계정코드'].fillna('').str.startswith(lookup_value)]
        
        
        
        #code_row = account_code_df[account_code_df['계정'].str.startswith(lookup_value)]
        if not code_row.empty:
            mon_df.at[index, '코드1'] = code_row.iloc[0]['코드1']
        else:
            mon_df.at[index, '코드1'] = 0  # 일치하는 값이 없으면 0 처리
    
     
    
    
    
    
    
# 결과를 엑셀 파일로 저장
# 3. 병합된 결과를 엑셀 파일로 저장
mon_df.to_excel('C:\dev\myproj01\금상동\결과폴더\월별병합엑셀파일9-20.xlsx', index=False)



# 2. VLOOKUP 유사 기능: '코드1' 열을 기준으로 lookup_df의 4번째 열과 병합
# lookup_df의 4번째 열의 이름을 임시로 설정
lookup_df_column_name = account_code_df.columns[5]  # 4번째 열의 이름 가져오기

# 병합 수행
merged_df = pd.merge(mon_df, account_code_df[['코드2','계정코드2']], how='left',left_on='코드1',right_on='코드2')








# 3. 불필요한 '코드2' 열 제거 (선택 사항)
merged_df.drop(columns=['코드2'], inplace=True)

#mon_df=mon_df['적       요'].fillna(method='ffill') # 최근일 월일없는행은 전일자로 처리




merged_df.iloc[:, 3] = merged_df.iloc[:, 3].fillna(method='ffill')


# 두 문자열 칼럼을 합쳐서 새로운 칼럼에 저장
merged_df['적요명'] = merged_df.iloc[:, 2].astype(str) + merged_df.iloc[:, 5].astype(str)


# 3. 병합된 결과를 엑셀 파일로 저장
merged_df.to_excel('C:\dev\myproj01\금상동\결과폴더\월별병합코드엑셀파일9-20.xlsx', index=False)




merged_df.info()

def calculate_amount(row):
    수입 = row['수입']
    지출 = row['지출']
    구분1 = float(row['구분1'])
    계정과목 = str(row['계정과목'])
    납부방법 = str(row['납부방법'])
    회계계정 = str(row['회계계정'])

    if 구분1 == 0:
        return (수입 + 지출) * 구분1

    # '해약' 관련 조건
    if '해약' in 계정과목 and 지출 > 0:
        if '이체' in 납부방법 and ('관리선수' in 회계계정 or '사용선수' in 회계계정):
            return 수입 + 지출
        return 수입 + 지출 * -1

    # '사용료' 관련 조건
    if '사용료' in 계정과목 and '현금' in 납부방법 and ('사용수입' in 회계계정 or '관리수입' in 회계계정) and 수입 > 0:
        return 수입 + 지출

    # '관리비' 관련 조건
    if '관리비' in 계정과목 and '이체' in 납부방법 and ('관리수입' in 회계계정 or '사용선수' in 회계계정) and 수입 > 0:
        return 수입 + 지출

    # 기본 계산
    return 수입 + 지출

# '금액' 열을 계산하여 추가
merged_df['금액'] = merged_df.apply(calculate_amount, axis=1)


def calculate_amount2(row):
    수입 = row['수입']
    지출 = row['지출']
    구분1 = float(row['구분1'])
    계정과목 = str(row['계정과목'])
    납부방법 = str(row['납부방법'])
    회계계정 = str(row['회계계정'])

    if 구분1 == 0:
        return (수입 + 지출) * 구분1

    # '해약' 관련 조건
    if '해약' in 계정과목 and 지출 > 0:
        if '이체' in 납부방법 and ('관리선수' in 회계계정 or '사용선수' in 회계계정):
            return 수입 + 지출
        return 수입 + 지출 * 1

    # '사용료' 관련 조건
    if '사용료' in 계정과목 and '현금' in 납부방법 and ('사용수입' in 회계계정 or '관리수입' in 회계계정) and 수입 > 0:
        return 수입 + 지출

    # '관리비' 관련 조건
    if '관리비' in 계정과목 and '이체' in 납부방법 and ('관리수입' in 회계계정 or '사용선수' in 회계계정) and 수입 > 0:
        return 수입 + 지출

    # 기본 계산
    return 수입 + 지출

# '금액' 열을 계산하여 추가
merged_df['금액2'] = merged_df.apply(calculate_amount2, axis=1)




# 데이터프레임에 '금액' 열 추가
#mon_df['금액'] = mon_df.apply(calculate_amount, axis=1)

merged_df['지출'] =merged_df['지출'].fillna(0)

merged_df['수입'] =merged_df['수입'].fillna(0)

merged_df['금액'] = merged_df['금액'].fillna(0)

merged_df['금액']=merged_df['금액'].astype(int)






# 코드2 값을 계산하는 함수 정의

def calculate_code2(row):
    if row['구분1'] == 0:
        return 0

    # '납부방법' 및 '회계계정' 조건을 변수로 처리
    납부방법 = str(row['납부방법'])
    회계계정 = str(row['회계계정'])
    계정과목 = str(row.get('계정과목', ''))  # '계정과목'이 없을 경우 빈 문자열

    # 지출이 발생한 경우
    if row['지출'] > 0:
        if '이체' in 납부방법:
            if '일반관리' in 회계계정:
                return 274
            elif '회전관리선수' in 회계계정 or '관리선수' in 회계계정:
                return 274
            elif '회전' in 회계계정:
                return 275
            elif '사용수입' in 회계계정 or '사용선수' in 회계계정 :
                return 273
        elif '대체' in 납부방법:
            if '관리선수' in 회계계정 or '관리수입' in 회계계정:
                return 274
            elif '사용수입' in 회계계정:
                return 273

    # 수입이 발생한 경우
    if row['수입'] > 0:
        if '이체' in 납부방법 and '관리수입' in 회계계정 and row['수입'] < 100000:
            return 0  # C50 외상매출금 참조
        elif '이체' in 납부방법:
            return 108  # C20 미수금 참
        
        elif '신용' in 납부방법:
            return 120  # C20 미수금 참조
        elif '현금' in 납부방법:
            return 103  # C73 현금 참조

    # '대체' 납부방법과 '해약' 또는 '사용선수' 관련 조건
    if '대체' in 납부방법:
        if '해약' in 계정과목 or '사용선수' in 회계계정:
            return row['코드1']
    elif '이체' in 납부방법 and '해약' in 계정과목:
        return row['코드1']

    # 기본값 반환
    return 0

# '코드2' 열을 계산하여 추가
merged_df['코드2'] = merged_df.apply(calculate_code2, axis=1)





# 엑셀 파일 불러오기

code_df = pd.read_excel('C:\dev\myproj01\금상동\결과폴더\금상동계정코드(2024-11-21).xlsx')

# 금상동 계정코드 파일을 딕셔너리로 변환
code_dict = code_df.set_index(code_df.columns[0]).to_dict()[code_df.columns[1]]



def calculate_gubun2(row):
    # 납부방법과 회계계정, 계정과목 변수 정의
    납부방법 = str(row['납부방법'])
    회계계정 = str(row['회계계정'])
    계정과목 = str(row.get('계정과목', ''))  # '계정과목'이 없을 경우 빈 문자열

    # '대체' 관련 조건
    if '대체' in 납부방법:
        if '해약' in 계정과목 or '사용선수' in 회계계정:
            return 3
        if '사용수입' in 회계계정 and row['지출'] > 0:
            return 3
        if '관리선수' in 회계계정 and row['지출'] > 0:
            return 3

    # '이체' 관련 조건
    if '이체' in 납부방법 and '해약' in 계정과목:
        return 3

    # 코드1이 0일 때
    if row['코드1'] == 0:
        return 0

    # 구분1이 2일 때
    if row['구분1'] == 2:
        return 0

    # 그 외의 경우
    return row['구분1'] - 1

# '구분2' 열을 계산하여 추가
merged_df['구분2'] = merged_df.apply(calculate_gubun2, axis=1)




code_df1=code_df[['코드2','계정코드2']]

# VLOOKUP 함수 구현
def vlookup(lookup_value, code_df1, col_index, range_lookup=False):
    # 'range_lookup'에 따라 정확히 일치하는 값 또는 근사치를 찾을 수 있지만,
    # 현재는 정확한 일치를 기준으로 구현
    result = code_df1[code_df1.iloc[:, 0] == lookup_value]
    if not result.empty:
        return result.iloc[0, col_index - 1]  # col_index는 1부터 시작
    return 0 if not range_lookup else "Approximate Not Found"



# VLOOKUP을 적용할 데이터
lookup_col = '코드2'  # VLOOKUP에 사용할 열 이름
value_col = '계정코드2'      # 반환할 열 이름
result_col = '계정코드3'  # 결과를 저장할 열 이름

# VLOOKUP 함수 적용
merged_df[result_col] = merged_df[lookup_col].apply(lambda x: vlookup(x, code_df1, 2))

#merged_df['금액2']=None
#merged_df['금액2']=merged_df['금액'].astype(int)


# 금액2 열을 맨 끝으로 이동시킵니다
# 현재 열 순서를 얻습니다
columns = list(merged_df.columns)

# 금액2가 맨 끝으로 오도록 열 순서를 재배열합니다
new_columns_order = [col for col in columns if col != '금액2'] + ['금액2']

# DataFrame을 새로운 열 순서로 재정렬합니다
merged_df=merged_df[new_columns_order]





merged_df['번호']=merged_df['Unnamed: 1']


print(merged_df)
#merged_df.info()

merged_df.to_excel('C:\dev\myproj01\금상동\결과폴더\금상동월별결산부2024업로드20.xlsx', index=False)
upload_df = pd.DataFrame()
dae_df=merged_df[['번호','월일','구분1','코드1','계정코드2','적요명','금액']]
cha_df=merged_df[['번호','월일','구분2','코드2','계정코드3','적요명','금액2']]

cha_df['구분1']=cha_df['구분2']
cha_df['코드1']=cha_df['코드2']
cha_df['계정코드2']=cha_df['계정코드3']
cha_df['금액']=cha_df['금액2']


#merged_df = pd.merge(mon_df, account_code_df[['코드2','계정코드2']], how='left',left_on='코드1',right_on='코드2')

#upload_df = pd.merge(cha_df, dae_df, how='left',left_on='구분1',right_on='구분1')



cha_df2=cha_df[['번호','월일','구분1','코드1','계정코드2','적요명','금액']]







#dae_df.info()
cha_df2.to_excel('C:\dev\myproj01\금상동\결과폴더\차변업로드.xlsx', index=False)
dae_df.to_excel('C:\dev\myproj01\금상동\결과폴더\대변업로드.xlsx', index=False)
#upload_df = pd.merge(cha_df, dae_df[['월일','구분2','코드2','계정코드3']], how='left',left_on='월일',right_on='코드2')

#merged_df = pd.merge(mon_df, account_code_df[['코드2','계정코드2']], how='left',left_on='코드1',right_on='코드2')




# 데이터프레임 1과 2의 행 수가 같은지 확인합니다
#assert len(cha_df) == len(dae_df), "두 데이터프레임의 행 수가 다릅니다."

# 빈 데이터프레임을 만듭니다
upload_df = pd.DataFrame()

# 교차로 행을 추가합니다
for i in range(len(cha_df)):
    upload_df= pd.concat([upload_df, cha_df.iloc[[i]], dae_df.iloc[[i]]], ignore_index=True)

#upload_df['차대']=cha_df['구분1']+dae_df['구분2']
#upload_df=upload_df.drop(columns=['구분1'], inplace=True)
#upload_df.info()



#mon_df['구분1'] > 1) & (mon_df['코드1'] > 0)


#mon_df = mon_df[mon_df['구분1'] >= 1]
#upload_df['구분1']=upload_df['구분1'].astype(int)

upload_df=upload_df[['월일','구분1','코드1','계정코드2','적요명','금액']]


#upload_df2.info()

upload_df2= upload_df[(upload_df['구분1'] >= 1) & (upload_df['코드1'] > 0)] #구분1 값이 1보다 큰값은 남기고 코드1값이 0 이면 삭제하기위함


#upload_df2= upload_df[upload_df['구분1'] >= 1]


upload_df2.info()

upload_df2.insert(4, '코드4', '')      # 6th column, index 5
upload_df2.insert(5, '거래처명', '')   # 7th column, index 6
upload_df2.insert(6, '코드5', '')      # 8th column, index 7








#upload_df2.to_excel('C:\dev\myproj01\금상동\결과폴더\금상동세무사랑2024업로드920.xlsx', index=False)


# 1. 엑셀 파일 불러오기
#input_file = 'C:\dev\myproj01\금상동\결과폴더\금상동세무사랑2024업로드920.xlsx'  # 여기에 불러올 엑셀 파일의 경로를 입력하세요.
#df7 = pd.read_excel(input_file)
df7 = upload_df2

df7.rename(columns={'월일' : '년도월일'}, inplace=False)
df7.rename(columns={'계정코드2' : '계정과목'}, inplace=False)
df7.rename(columns={'코드4' : '코드'}, inplace=False)







# 2. 데이터프레임을 엑셀 파일로 저장 (임시 저장)
output_file = 'C:\dev\myproj01\금상동\결과폴더\금상동세무사랑기타업로드수정11-23.xlsx'  # 여기에 저장할 엑셀 파일의 경로를 입력하세요.
df7.to_excel(output_file, index=False)

# 3. openpyxl을 사용하여 엑셀 파일 열기
wb = load_workbook(output_file)
ws = wb.active

# 4. 열 너비 설정
ws.column_dimensions['D'].width = 20  # '구분1' 열 (가정: A열에 위치)
ws.column_dimensions['H'].width = 25  # '적요명' 열 (가정: B열에 위치)

# 5. 엑셀 파일 저장
wb.save(output_file)

#

















#####################################################################################
#total_df4=total_df4.drop(columns=["잔액"])

#total_df4=total_df4.drop(columns=["차변"], inplace=True)


 # 엑셀 파일 불러오기
file_path3 = r"C:\dev\myproj01\금상동세무사랑회전관리선수금2022-2023.xls"   # 엑셀 파일 경로를 지정하세요
df17= pd.read_excel(file_path3,sheet_name="회전관리선수금2022-2023봉안함순",skiprows=0,usecols="a:e") 

df17=df17.rename(columns={"일자":"날짜"})

df17=df17.rename(columns={"계약자":"성명"})
df17=df17.rename(columns={"대    변":"대변"})
df17["잔액"]=0
df17.info()



df17.to_excel(r'C:\dev\myproj01\금상동\결과폴더\회전관리선수금2022-2023봉안함순2.xlsx') 



df20=pd.concat([total_df4,df17])
df20['날짜'] =pd.to_datetime(df20['날짜'])
df20['날짜']=df20['날짜'].dt.strftime('%y-%m-%d') #날짜를 날짜함수이므로 문자함수로 변환
df20=df20.sort_values("봉안함")
df20.to_excel(r'C:\dev\myproj01\금상동\결과폴더\회전관리선수금2022-2023봉안함순.xlsx')







# 데이터프레임과 시트 이름 리스트
dfs = [total_df, total_df2,total_df3,total_df4,df20,total_df5,total_df6,total_df7,total_df8,total_df9]











sheet_names = ['사용선수금', '관리선수금','사용료수입','회전관리선수금','회전관리선수금통합-봉안함순','월별결산부','사용선수금-일자순','관리선수금-일자순','사용료수입-일자순','회전관리선수금-일자순']
# 함수 호출
#export_multiple_dfs_to_excel(dfs, sheet_names,'C:\dev\myproj01\금상동\결과폴더\월별결산부사용선수금-관리선수금-사용료수입-회전관리선수금2024.xlsx')

export_multiple_dfs_to_excel(dfs, sheet_names,'C:\dev\myproj01\금상동\결과폴더\월별결산부+회전관리선수금통합.xlsx')

# 엑셀 파일 열기
#file_path = r'C:\dev\myproj01\금상동\결과폴더\월별결산부사용선수금-관리선수금-사용료수입-회전관리선수금2024.xlsx'  # 엑셀 파일 경로를 입력하세요
file_path = r'C:\dev\myproj01\금상동\결과폴더\월별결산부+회전관리선수금통합.xlsx'  # 엑셀 파일 경로를 입력하세요
wb = openpyxl.load_workbook(file_path)

# 모든 시트에 대해 A열부터 D열까지의 너비를 25로 설정
columns_to_resize = ['A', 'B', 'C', 'D','E','F','G']  # A열부터 D열까지 지정

for sheet in wb.worksheets:
    for column_letter in columns_to_resize:
        # 지정된 열(column_letter)의 너비를 25로 설정
        sheet.column_dimensions[column_letter].width = 15



# 통화 기호 없는 회계 스타일 정의
accounting_style = NamedStyle(name="accounting_style", number_format='#,##0;[Red](#,##0)')

# 모든 시트에 대해 E열부터 G열까지 회계 스타일 적용
for sheet in wb.worksheets:
    for row in range(1, sheet.max_row + 1):
        for col in ['E', 'F', 'G']:  # E열부터 G열까지
            cell = sheet[f'{col}{row}']
            cell.style = accounting_style








# 수정된 엑셀 파일 저장

# 수정된 엑셀 파일 저장 (경로와 파일명 수정)
#save_path = r'C:\dev\myproj01\금상동\결과폴더\월별결산부사용선수금-관리선수금-사용료수입-회전관리선수금(열조정2024).xlsx'
save_path = r'C:\dev\myproj01\금상동\결과폴더\월별결산부+회전관리선수금통합(열조정2024).xlsx'
wb.save(save_path)












total_df.info()
			
import pandas as pd
from openpyxl import load_workbook

# 파일 경로 설정
#file_path = r'C:\dev\myproj01\금상동\결과폴더\월별결산부사용선수금-관리선수금-사용료수입-회전관리선수금(열조정2024).xlsx'  # 여기에 엑셀 파일 경로 입력
file_path = r'C:\dev\myproj01\금상동\결과폴더\월별결산부+회전관리선수금통합(열조정2024).xlsx'  # 여기에 엑셀 파일 경로 입력



# 엑셀 파일 불러오기 (openpyxl 사용)
wb = load_workbook(file_path)
#sheet_names = ['사용선수금', '관리선수금', '사용료수입', '회전관리선수금']  # 작업할 시트 이름들
sheet_names = ['사용선수금', '관리선수금','사용료수입','회전관리선수금','회전관리선수금통합-봉안함순','월별결산부','사용선수금-일자순','관리선수금-일자순','사용료수입-일자순','회전관리선수금-일자순']  # 작업할 시트 이름들

#'사용선수금', '관리선수금','사용료수입','회전관리선수금','월별결산부','사용선수금-일자순','관리선수금-일자순','사용료수입-일자순','회전관리선수금-일자순','회전관리선수금누적(봉안함순)'
# 각 시트에 대해 작업 수행
for sheet_name in sheet_names:
    # pandas로 시트를 읽어오기
    df21 = pd.read_excel(file_path, sheet_name=sheet_name)
    
    # "계정과목" 열을 추가하고 시트 이름으로 채우기
    df21['계정과목'] = sheet_name
    
    # pandas로 수정된 데이터를 다시 엑셀 파일로 저장
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df21.to_excel(writer, sheet_name=sheet_name, index=False)

print("모든 시트에 '계정과목' 열 추가 완료.")

import pandas as pd

# 파일 경로 설정
a_file = r'C:\dev\myproj01\금상동\결과폴더\금상동_2024.xlsx'  # a파일 경로
#b_file = r'C:\dev\myproj01\금상동\결과폴더\월별결산부사용선수금-관리선수금-사용료수입-회전관리선수금(열조정2024).xlsx'  # b파일 경로
b_file = r'C:\dev\myproj01\금상동\결과폴더\월별결산부+회전관리선수금통합(열조정2024).xlsx'   # b파일 경로

# 1. a 파일 불러와서 데이터프레임 a 생성
df_a = pd.read_excel(a_file)
#df_b = pd.read_excel(b_file)
# 2. b 파일의 첫 번째 시트부터 네 번째 시트까지 데이터프레임 생성
b_sheets = [pd.read_excel(b_file, sheet_name=i) for i in range(5)]

# 3. 첫 번째 시트를 a 데이터프레임에 병합하여 a-1 데이터프레임 생성
df_a1 =df_a





# 4. 두 번째, 세 번째, 네 번째 시트를 차례로 병합
for i in range(1, 5):
    #b_sheets[i] = b_sheets[i].drop_duplicates(subset=['봉안함'])  # 봉안함 중복 제거 
    df_a1 = pd.merge(df_a1, 
                     b_sheets[i][['봉안함', '차변', '대변', '계정과목','날짜']], 
                     on='봉안함', 
                     how='left',
                     suffixes=('', f'_b{i}'))


# 4. 병합된 결과를 엑셀 파일로 저장
# 날짜를 문자열로 포맷 (예: '2024-10-03')
today = datetime.today().strftime('%Y-%m-%d')

output_file = f'C:\dev\myproj01\금상동\결과폴더\금상동봉안+월별결산부병합결과-{today}.xlsx'


df_a1=df_a1.rename(columns={"대변_b4":"누적-회전관리선수금"})
df_a1=df_a1.rename(columns={"대변_b3":"2024-회전관리선수금"})
df_a1=df_a1.rename(columns={"차변_b2":"사용료수입"})



df_a1.to_excel(output_file, index=False)

print(f"병합된 데이터가 {output_file}로 저장되었습니다.")


# 엑셀 파일 열기
file_path = f'C:\dev\myproj01\금상동\결과폴더\금상동봉안+월별결산부병합결과-{today}.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb.active  # 활성화된 시트를 선택
# 모든 시트에 대해 A열부터 D열까지의 너비를 25로 설정
columns_to_resize = ['A', 'B', 'C', 'D','E','F','G','Y', 'Z','AA', 'AB','AC','AD','AH', 'AI','AJ', 'AK','AL','AM','AN']  # A열부터 D열까지 지정

for sheet in wb.worksheets:
    for column_letter in columns_to_resize:
        # 지정된 열(column_letter)의 너비를 25로 설정
        sheet.column_dimensions[column_letter].width = 15



       # 통화 기호 없는 회계 스타일 정의
        accounting_style = NamedStyle(name="accounting_style", number_format='#,##0;[Red](#,##0)')

# 모든 시트에 대해 E열부터 G열까지 회계 스타일 적용
    for sheet in wb.worksheets:
      for row in range(1, sheet.max_row + 1):
        for col in ['Y', 'Z','AA', 'AB','AC','AD','AH', 'AI','AJ', 'AK','AL','AM','AN']:  # E열부터 G열까지
            cell = sheet[f'{col}{row}']
            cell.style = accounting_style

    
    # '전체월수' 열의 값이 180을 초과하는 경우 색상을 빨간색으로 변경
    fill_red = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

    # 전체월수 열의 값을 차례대로 불러와서 처리
    for idx, value in enumerate(df_a1['전체월수'], start=2):  # start=2는 헤더를 건너뛰기 위해
        if value > 180:
            ws[f'AE{idx}'].fill = fill_red  # C열에 '전체월수'가 있다고 가정
    
    
    
    
    # 수정된 엑셀 파일 저장
    # A2 셀 기준으로 틀 고정 (freeze panes)
    ws.freeze_panes = 'A2'

    # 첫 번째 행에 필터 추가
    ws.auto_filter.ref = ws.dimensions  # ws.dimensions는 사용 중인 셀의 범위를 자동으로 가져옴

    # 수정된 엑셀 파일 저장 (경로와 파일명 수정)
    
    today = datetime.today().strftime('%Y-%m-%d')

    
    
    save_path = f'C:\dev\myproj01\금상동\결과폴더\금상동봉안+월별결산부병합결과-{today}.xlsx' 
    wb.save(save_path)
    print(f"새로운 데이터가 {save_path} 파일로 저장되었습니다.")        

    # 첫 번째 파일 (기존 파일)
file1 = r'C:\dev\myproj01\금상동\결과폴더\월별결산부+회전관리선수금통합(열조정2024).xlsx'

# 두 번째 파일 (추가할 파일)
file2 = r'C:\dev\myproj01\금상동\결과폴더\금상동세무사랑기타업로드수정11-23.xlsx'

# 파일을 읽어들임
dfs_file2 = pd.read_excel(file2, sheet_name=None)  # 두 번째 파일의 모든 시트를 딕셔너리로 읽기

# 엑셀 파일의 기존 워크북 불러오기
with pd.ExcelWriter(file1, engine="openpyxl", mode="a") as writer:
    for sheet_name, df in dfs_file2.items():  # 두 번째 파일의 시트들 순회
        df.to_excel(writer, sheet_name="엑셀기타업로드", index=False)  # 새로운 시트로 추가

print(f"'{file2}' 파일의 모든 시트가 '{file1}' 파일에 추가되었습니다.")
