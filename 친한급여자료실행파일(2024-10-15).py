from PIL import Image
import matplotlib.pyplot as plt
# -*- coding: utf-8 -*-
import pandas as pd

import pandas as pd
from openpyxl import load_workbook

import os




   # 이미지 파일 경로
image_path = r'C:\dev\myproj01\17블럭토지도면.jpg'  # 여기에 JPG 파일 경로를 입력하세요.

   # 이미지 열기
img = Image.open(image_path)

   # 이미지 표시
plt.imshow(img)
plt.axis('off')  # 축을 숨김
plt.show()





# 1. 병합된 셀 해제 및 엑셀 파일 불러오기
file_path = 'C:\dev\myproj01\복사본 2024년9월친한.xlsx'  # 불러올 엑셀 파일 경로

# 엑셀 파일 로드
wb = load_workbook(file_path,data_only=True)
#ws = wb.active


# 엑셀 파일 로드
#wb = load_workbook(file_path)
ws = wb['2024급여(9)']  # 특정 시트 선택





# openpyxl을 사용하여 엑셀 파일 열기 (data_only=True로 수식 계산된 값만 불러옴)
#wb = openpyxl.load_workbook(input_file, data_only=True)

# 2. "2024급여(8)" 시트 불러오기
source_sheet = wb['2024급여(9)']

# 3. 새로운 시트 "2024급여8" 만들기
new_sheet_name = '2024급여9'
if new_sheet_name in wb.sheetnames:
    wb.remove(wb[new_sheet_name])  # 이미 있으면 삭제하고 새로 만들기
new_sheet = wb.create_sheet(new_sheet_name)

# 4. 원본 시트의 데이터를 새로운 시트에 복사 (값만 복사)
for row in source_sheet.iter_rows(values_only=True):
    new_sheet.append(row)  # 계산된 값만 복사

# 5. 엑셀 파일 저장
#wb.save(output_file)




# 병합된 셀 해제
for merge in list(ws.merged_cells):
    ws.unmerge_cells(str(merge))


#data = []
#for row in ws.iter_rows(values_only=True):  # 수식이 아닌 현재 값을 읽음
#    data.append(row)

#df = pd.DataFrame(data)







# 병합 해제된 파일을 임시로 저장하고, pandas로 다시 읽기
wb.save('c:/dev/myproj01/temp_unmerged0912.xlsx')

new_column_names = ['순번', '성명', 'vlookup', '월기본', '월소정', '기본시급','기본', '연장','(휴일)연장', '기본급', '연장수당','휴일연장', '식대','교통비','기타수당','지급합계','비과세제외합계','갑근세','국민연금','건강보험','착오정산','공제합계','차인지급액','비고']
#new_column_names = ['순번', '성명', 'vlookup', '월기본', '월소정', '기본시급','기본', '연장','(휴일)연장', '기본급', '연장수당','휴일연장', '식대']





df =pd.read_excel("c:/dev/myproj01/temp_unmerged0912.xlsx",sheet_name="2024급여9",skiprows=4,usecols="a:y")








df.columns = new_column_names  # 칼럼명 변경


output_file_path = 'C:\dev\myproj01\친한9월급여결과0912.xlsx'  # 저장할 엑셀 파일 경로
#df.info()
#df2=df
df.to_excel(output_file_path, index=False)


# 1. 엑셀 파일 불러오기
input_file = 'C:\dev\myproj01\친한9월급여결과0912.xlsx'  # 엑셀 파일 경로
output_file = 'C:\dev\myproj01\친한9월급여결과0914.xlsx'  # 저장할 엑셀 파일 경로

# 2. 데이터프레임1으로 엑셀 파일 읽기
df1 = pd.read_excel(input_file)

# 3. 짝수행의 값들을 가져오기 위한 열 리스트 설정
source_columns = ['vlookup','기본급', '연장수당', '휴일연장', '갑근세', '국민연금', '건강보험']
target_columns = ['성명','휴게수당', '휴일수당', '야간수당', '주민세', '고용보험', '요양보험']

# 짝수 행의 값 가져오기 (pandas는 0부터 시작하므로 짝수 행은 index 1, 3, 5,...)
even_rows_df1 = df1.loc[1::2, source_columns].reset_index(drop=True)

#df3= df1.drop_duplicates(subset=['성명']) #결제일과 환율이 중복되는  행삭제
df3=df1.dropna(subset=['성명'])


df3.to_excel('C:\dev\myproj01\친한9월급여결과df3.xlsx',index=False)
# 4. 새로운 데이터프레임2 생성 (짝수행의 값을 새로운 열에 넣기)
df2 = pd.DataFrame()

# source_columns의 짝수행 값을 target_columns에 복사
for source_col, target_col in zip(source_columns, target_columns):
    df2[target_col] = even_rows_df1[source_col]

# 5. 데이터프레임2를 엑셀로 저장
df2.to_excel(output_file, index=False) 


df5=df3.merge(df2,how="outer",on=["성명"])

#df5= df5.drop_duplicates(subset=['성명']) #결제일과 환율이 중복되는  행삭제
#df5=df5.dropna(subset=['성명'])
#df5=pd.merge(df3,df2,'성명')
df5.to_excel('C:\dev\myproj01\친한9월급여결과0915.xlsx',index=False)

# 엑셀 파일 불러오기
file_path = 'C:\dev\myproj01\친한사번성명.xlsx'
df7 = pd.read_excel(file_path)

# '사번' 열을 복사하여 '사번2' 열 생성
df7['사번2'] = df7['사번']
df8=df5.merge(df7,how="outer",on=["성명"])

df8['주민번호'] = df8['주민(외국인)번호.1']
df8['사원코드'] = df8['사번']
df8['사원명'] = df8['성명']
df8['기본급-1001'] = df8['기본급']
df8['상여-1002'] = 0
df8['직책수당-1003'] = 0
df8['월차수당-1004'] = 0
df8['식대-1005'] = df8['식대']
df8['자가운전보조금-1006'] = df8['교통비']
df8['연장수당-2001'] = df8['연장수당']
df8['연차수당-2002'] = 0
df8['교통비-2003'] = 0
df8['공제(결근 등)-2004'] = 0
df8['명절상여-2005'] = 0
df8['휴일수당-2006'] = 0
df8['기타수당-2007'] = df8['기타수당']
df8['고정수당-2008'] = 0
df8['야근수당-2009'] = 0
df8['기타지급-2010'] = 0
df8['기타금품-2011'] = 0
df8['휴게수당-2012'] = df8['휴게수당']
df8['휴일연장수당-2013'] = 0

df8['국민연금-5001'] = df8['국민연금']
df8['건강보험-5002'] = df8['건강보험']

df8['장기요양보험-5003'] = df8['요양보험']
df8['고용보험-5004'] = df8['고용보험']
df8['학자금상환-5005'] =0
df8['기타공제-6001'] =0
df8['보험료과오납-6002'] =0
df8['기지급분공제-6003'] =0

df8['건강보험험정산-6004'] =0
df8['세금정산사회보험정산-6005'] =0
df8['가불-6006'] =0
df8['소득세-9994'] = df8['갑근세']
df8['지방소득세-9995'] = df8['주민세']
df8['농특세-9996'] =0
df8.info()




df8.to_excel('C:\dev\myproj01\친한9월급여결과1010.xlsx',index=False)

#df9=df8[['사원코드','사원명','주민번호','기본급-1001','상여-1002','직책수당-1003','월차수당-1004','식대-1005','자가운전보조금-1006','연장수당-2001','연차수당-2002','교통비-2003','공제(결근 등)-2004','명절상여-2005','휴일수당-2006','기타수당-2007','고정수당-2008','야근수당-2009','기타지급-2010','기타금품-2011','휴게수당-2012','휴일연장수당-2013','국민연금-5001','건강보험-5002','장기요양보험-5003','고용보험-5004','학자금상환-5005','기타공제-6001','보험료과오납-6002','기지급분공제-6003','건강보험험정산-6004','세금정산사회보험정산-6005','가불-6006','소득세-9994','지방소득세-9995','농특세-9996']]
df9=df8[['사원코드','사원명','주민번호','기본급-1001','상여-1002','직책수당-1003','월차수당-1004','식대-1005','자가운전보조금-1006','연장수당-2001','연차수당-2002','교통비-2003','공제(결근 등)-2004','명절상여-2005','휴일수당-2006','기타수당-2007','고정수당-2008','야근수당-2009','기타지급-2010','기타금품-2011','휴게수당-2012','휴일연장수당-2013','국민연금-5001','건강보험-5002','장기요양보험-5003','고용보험-5004','학자금상환-5005','기타공제-6001','보험료과오납-6002','기지급분공제-6003','건강보험험정산-6004','세금정산사회보험정산-6005','가불-6006','소득세-9994','지방소득세-9995','농특세-9996']]

# 조건: 사원명이 결측치(NaN)이면서 기본급-1001 열의 값이 0인 행 제거
df10 = df9[df9['기본급-1001'] >= 10000]


df10.to_excel('C:\dev\myproj01\친한9월급여결과1017.xlsx',index=False)

# 엑셀 파일 불러오기
file_a = r'C:\dev\myproj01\친한급여자료업로드2024년9월(1015).xlsx'  # 엑셀 파일 A 경로
file_b = 'C:\dev\myproj01\친한9월급여결과1017.xlsx'  # 엑셀 파일 B 경로

output_file = r'C:\dev\myproj01\친한급여자료업로드2024년9월(1015)-2.xlsx'

# 첫 번째 파일의 전체 데이터를 불러옵니다 (a8 셀부터 데이터 대치 예정)
df_a = pd.read_excel(file_a, header=None)

# 두 번째 파일의 모든 데이터를 불러옵니다 (a2부터 값 대치)
df_b = pd.read_excel(file_b, header=None)

# 두 번째 파일의 셀 값 (A2부터 시작)만큼을 첫 번째 파일의 A8부터 대치합니다.
# df_b는 전체 셀을 A2부터 입력한다고 가정
df_a.iloc[7:7 + len(df_b), :len(df_b.columns)] = df_b.values

# 변경된 첫 번째 파일을 새로운 파일로 저장
df_a.to_excel(output_file, index=False, header=False)

print(f"새로운 파일이 {output_file}로 저장되었습니다.")






