import pandas as pd			
from pathlib import Path
import openpyxl	
import pandas as pd

from openpyxl import load_workbook

import pandas as pd			
from pathlib import Path
import openpyxl	
from datetime import datetime
import pandas as pd			
from pathlib import Path
import openpyxl	
from openpyxl.styles import NamedStyle


# 여러 개의 데이터프레임을 엑셀 파일로 내보내기 함수
def export_multiple_dfs_to_excel(dfs, sheet_names, file_name):
    """
    여러 개의 데이터프레임을 각각 다른 시트에 엑셀 파일로 내보내는 함수.
    : param dfs: 데이터프레임 리스트 
    : param sheet_names: 시트 이름 리스트 
    : param file_name: 저장할 엑셀 파일 이름
   """ 
    
    if len(dfs) != len(sheet_names) :

      raise ValueError("데이터프레임 리스트와 시트 이름 리스트의 길이가 같아야 합니다. ")


    with pd. ExcelWriter (file_name) as writer:
     for df, sheet_name in zip(dfs, sheet_names) :
        df. to_excel(writer, sheet_name=sheet_name,index=False)
    #Print(f"{file_name} 파일이 생성되었습니다.")






input_folder='C:\dev\myproj01\금상동2024'
raw_data_dir =Path(input_folder)
excel_files=raw_data_dir.glob('24년*')





total_df=pd.DataFrame() #빈 데이터프래임 생성
total_df2=pd.DataFrame() #빈 데이터프래임 생성
total_df3=pd.DataFrame() #빈 데이터프래임 생성
total_df4=pd.DataFrame() #빈 데이터프래임 생성
total_df5=pd.DataFrame() #빈 데이터프래임 생성











i=0
for excel_file in excel_files:
    i=i+1
   
    
    
    
    df =pd.read_excel(excel_file,sheet_name="사용선수금",skiprows=3,usecols="a:g")
    df.columns = ['날짜', '성명', '봉안함', '적요1', '차변', '대변', '잔액']
    #df['날짜'] = pd.to_datetime(df['날짜']).dt.strftime('%Y-%m-%d')
    
    df2 =pd.read_excel(excel_file,sheet_name="관리선수금",skiprows=3,usecols="a:g")
    df2.columns = ['날짜', '성명', '봉안함', '적요1', '차변', '대변', '잔액']
    #f2['날짜'] = pd.to_datetime(df2['날짜']).dt.strftime('%Y-%m-%d')
    
    df3 =pd.read_excel(excel_file,sheet_name="사용료수입",skiprows=3,usecols="a:g")
    df3.columns = ['날짜', '성명', '봉안함', '적요1', '차변', '대변', '잔액']
    #df3['날짜'] = pd.to_datetime(df3['날짜']).dt.strftime('%Y-%m-%d')
    
    try:
     # 엑셀 파일의 시트 목록을 가져옴
        xls = pd.ExcelFile(excel_file)
        print(f"'{excel_file}' 파일의 시트 목록: {xls.sheet_names}")
    
        # '회전관선' 시트가 있는지 확인
        if "회전관선" in xls.sheet_names:
          print(f"'{excel_file}' 파일에서 '회전관선' 시트를 읽습니다.")
          # 시트를 읽어 데이터프레임으로 저장
          df4 = pd.read_excel(excel_file, sheet_name="회전관선", skiprows=3, usecols="A:G")
          df4.columns = ['날짜', '성명', '봉안함', '적요1', '차변', '대변', '잔액']
          total_df4=total_df4._append(df4,ignore_index=True)
          print(df4.head())  # 데이터프레임 출력
        else:
          print(f"'{excel_file}' 파일에 '회전관선' 시트가 없습니다. 다른 작업을 진행합니다.")
    
    except ValueError as e:
      print(f"오류 발생: {e}")
    except Exception as e:
      print(f"파일 '{excel_file}' 처리 중 오류 발생: {e}")

    print("프로그램이 종료되었습니다.")
    
       
    
    
    
    
    
    df4 = pd.read_excel(excel_file, sheet_name="회전관선-최준철세무사", skiprows=3, usecols="A:G")
    df4.columns = ['날짜', '성명', '봉안함', '적요1', '차변', '대변', '잔액']

    
    #df4 =pd.read_excel(excel_file,sheet_name="회전관선",skiprows=3,usecols="a:g")
    
    




    df5= pd.read_excel(excel_file,sheet_name=f"24-{i}",skiprows=2,usecols="a:i")
     
    
    
    
    
    
    
    total_df=total_df._append(df,ignore_index=True)	

    total_df2=total_df2._append(df2,ignore_index=True)

    total_df3=total_df3._append(df3,ignore_index=True)

    total_df4=total_df4._append(df4,ignore_index=True)

    total_df5=total_df5._append(df5,ignore_index=True)
    
   
    




total_df2['날짜'] = pd.to_datetime(total_df2['날짜'], errors='coerce')
total_df2['날짜'] = total_df2['날짜'].dt.strftime('%Y-%m-%d')

total_df3['날짜'] = pd.to_datetime(total_df3['날짜'], errors='coerce')
total_df3['날짜'] = total_df3['날짜'].dt.strftime('%Y-%m-%d')

total_df4['날짜'] = pd.to_datetime(total_df4['날짜'], errors='coerce')
total_df4['날짜'] = total_df4['날짜'].dt.strftime('%Y-%m-%d')


total_df5= total_df5.dropna(subset=["회계계정"]) #아래 누적액,월계액을 없애기위해 회계계정없는부분을 삭제함

total_df5['월일']=total_df5['월일'].dt.strftime('%Y-%m-%d') #구매자결제일이 날짜함수이므로 문자함수로 변환

total_df5['월일']=total_df5['월일'].str.replace('-','')

total_df5.info()
#total_df5['월일']=int(total_df5['월일'])




total_df5.columns = total_df5.columns.str.strip()
total_df5['월일']=total_df5['월일'].fillna(method='ffill') # 최근일 월일없는행은 전일자로 처리

#total_df5=total_df5['적요'].fillna(method='ffill') # 최근일 월일없는행은 전일자로 처리
total_df5['적       요']=total_df5['적       요'].fillna(method='ffill') # 최근일 월일없는행은 전일자로 처리















total_df=total_df.sort_values(["봉안함"]) # 일자로 정렬
total_df2=total_df2.sort_values(["봉안함"]) # 일자로 정렬
total_df3=total_df3.sort_values(["봉안함"]) # 일자로 정렬
total_df4=total_df4.sort_values(["봉안함"]) # 일자로 정렬







total_df= total_df.dropna(subset=["봉안함"]) #아래 봉안함번로 없는 행을 없애기위해 봉안함없는 부분을 삭제함
total_df2= total_df2.dropna(subset=["봉안함"]) #아래 봉안함번로 없는 행을 없애기위해 봉안함없는 부분을 삭제함
total_df3= total_df3.dropna(subset=["봉안함"]) #아래 봉안함번로 없는 행을 없애기위해 봉안함없는 부분을 삭제함
total_df4= total_df4.dropna(subset=["봉안함"]) #아래 봉안함번로 없는 행을 없애기위해 봉안함없는 부분을 삭제함















#total_df = pd.concat([total_df, df1])
total_df.to_excel("C:\dev\myproj01\금상동\사용선수금2024.xlsx")
total_df2.to_excel("C:\dev\myproj01\금상동\관리선수금2024.xlsx")
total_df3.to_excel("C:\dev\myproj01\금상동\사용료수입2024.xlsx")
total_df4.to_excel("C:\dev\myproj01\금상동\회전관리선수금2024.xlsx")
total_df5.to_excel("C:\dev\myproj01\금상동\결과폴더\월별결산부2024.xlsx") 
# 데이터프레임과 시트 이름 리스트
dfs = [total_df, total_df2,total_df3,total_df4,total_df5]

total_df.columns = ['날짜','성명','봉안함','적요1','차변','대변','잔액']
total_df2.columns = ['날짜','성명','봉안함','적요1','차변','대변','잔액']
total_df3.columns = ['날짜','성명','봉안함','적요1','차변','대변','잔액']
total_df4.columns = ['날짜', '성명','봉안함','적요1','차변','대변','잔액']
    


# 날짜 열을 먼저 datetime 형식으로 변환
total_df['날짜'] = pd.to_datetime(total_df['날짜'], errors='coerce')
total_df['날짜'] = total_df['날짜'].dt.strftime('%Y-%m-%d')

total_df2['날짜'] = pd.to_datetime(total_df2['날짜'], errors='coerce')
total_df2['날짜'] = total_df2['날짜'].dt.strftime('%Y-%m-%d')

total_df3['날짜'] = pd.to_datetime(total_df3['날짜'], errors='coerce')
total_df3['날짜'] = total_df3['날짜'].dt.strftime('%Y-%m-%d')

total_df4['날짜'] = pd.to_datetime(total_df4['날짜'], errors='coerce')
total_df4['날짜'] = total_df4['날짜'].dt.strftime('%Y-%m-%d')


total_df['날짜']=total_df['날짜'].fillna(method='ffill') # 최근일 월일없는행은 전일자로 처리
total_df=total_df.sort_values(["봉안함"]) # 일자와 환율을 일자와 환율로 정렬
    
total_df2['날짜']=total_df2['날짜'].fillna(method='ffill') # 최근일 월일없는행은 전일자로 처리
total_df2=total_df2.sort_values(["봉안함"]) # 일자와 환율을 일자와 환율로 정렬

total_df3['날짜']=total_df3['날짜'].fillna(method='ffill') # 최근일 월일없는행은 전일자로 처리
total_df3=total_df3.sort_values(["봉안함"]) # 일자와 환율을 일자와 환율로 정렬

total_df4['날짜']=total_df4['날짜'].fillna(method='ffill') # 최근일 월일없는행은 전일자로 처리
total_df4=total_df4.sort_values(["봉안함"]) # 일자와 환율을 일자와 환율로 정렬




sheet_names = ['사용선수금', '관리선수금','사용료수입','회전관리선수금','월별결산부']
# 함수 호출
export_multiple_dfs_to_excel(dfs, sheet_names,'C:\dev\myproj01\금상동\결과폴더\월별결산부사용선수금-관리선수금-사용료수입-회전관리선수금2024.xlsx')



# 엑셀 파일 열기
file_path = r'C:\dev\myproj01\금상동\결과폴더\월별결산부사용선수금-관리선수금-사용료수입-회전관리선수금2024.xlsx'  # 엑셀 파일 경로를 입력하세요
wb = openpyxl.load_workbook(file_path)

# 모든 시트에 대해 A열부터 D열까지의 너비를 25로 설정
columns_to_resize = ['A', 'B', 'C', 'D','E','F','G']  # A열부터 D열까지 지정

for sheet in wb.worksheets:
    for column_letter in columns_to_resize:
        # 지정된 열(column_letter)의 너비를 25로 설정
        sheet.column_dimensions[column_letter].width = 15



# 통화 기호 없는 회계 스타일 정의
accounting_style = NamedStyle(name="accounting_style", number_format='#,##0;[Red](#,##0)')

# 모든 시트에 대해 E열부터 G열까지 회계 스타일 적용
for sheet in wb.worksheets:
    for row in range(1, sheet.max_row + 1):
        for col in ['E', 'F', 'G']:  # E열부터 G열까지
            cell = sheet[f'{col}{row}']
            cell.style = accounting_style








# 수정된 엑셀 파일 저장

# 수정된 엑셀 파일 저장 (경로와 파일명 수정)
save_path = r'C:\dev\myproj01\금상동\결과폴더\월별결산부사용선수금-관리선수금-사용료수입-회전관리선수금(열조정2024).xlsx'
wb.save(save_path)












total_df.info()
			
import pandas as pd
from openpyxl import load_workbook

# 파일 경로 설정
file_path = r'C:\dev\myproj01\금상동\결과폴더\월별결산부사용선수금-관리선수금-사용료수입-회전관리선수금(열조정2024).xlsx'  # 여기에 엑셀 파일 경로 입력

# 엑셀 파일 불러오기 (openpyxl 사용)
wb = load_workbook(file_path)
sheet_names = ['사용선수금', '관리선수금', '사용료수입', '회전관리선수금']  # 작업할 시트 이름들

# 각 시트에 대해 작업 수행
for sheet_name in sheet_names:
    # pandas로 시트를 읽어오기
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    
    # "계정과목" 열을 추가하고 시트 이름으로 채우기
    df['계정과목'] = sheet_name
    
    # pandas로 수정된 데이터를 다시 엑셀 파일로 저장
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

print("모든 시트에 '계정과목' 열 추가 완료.")

import pandas as pd

# 파일 경로 설정
a_file = r'C:\dev\myproj01\금상동\결과폴더\금상동_2024.xlsx'  # a파일 경로
b_file = r'C:\dev\myproj01\금상동\결과폴더\월별결산부사용선수금-관리선수금-사용료수입-회전관리선수금(열조정2024).xlsx'  # b파일 경로

# 1. a 파일 불러와서 데이터프레임 a 생성
df_a = pd.read_excel(a_file)
#df_b = pd.read_excel(b_file)
# 2. b 파일의 첫 번째 시트부터 네 번째 시트까지 데이터프레임 생성
b_sheets = [pd.read_excel(b_file, sheet_name=i) for i in range(4)]

# 3. 첫 번째 시트를 a 데이터프레임에 병합하여 a-1 데이터프레임 생성
df_a1 =df_a

# 4. 두 번째, 세 번째, 네 번째 시트를 차례로 병합
for i in range(1, 4):
    b_sheets[i] = b_sheets[i].drop_duplicates(subset=['봉안함'])  # 봉안함 중복 제거 
    df_a1 = pd.merge(df_a1, 
                     b_sheets[i][['봉안함', '차변', '대변', '계정과목']], 
                     on='봉안함', 
                     how='left',
                     suffixes=('', f'_b{i}'))


# 4. 병합된 결과를 엑셀 파일로 저장
output_file = r'C:\dev\myproj01\금상동\결과폴더\병합결과1025.xlsx'
df_a1.to_excel(output_file, index=False)

print(f"병합된 데이터가 {output_file}로 저장되었습니다.")
