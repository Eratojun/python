import pandas as pd			
from pathlib import Path
import openpyxl	
import pandas as pd

from openpyxl import load_workbook

import pandas as pd			
from pathlib import Path
import openpyxl	
from datetime import datetime
from dateutil.relativedelta import relativedelta
import pandas as pd			
from pathlib import Path
import openpyxl	
from openpyxl.styles import NamedStyle
from openpyxl import Workbook
from openpyxl.styles import PatternFill #엑셀셀을 빨간색으로 표시
import socket
from openpyxl.utils.dataframe import dataframe_to_rows



def get_local_ip():
    """현재 컴퓨터의 IP 주소를 반환합니다."""
    hostname = socket.gethostname()  # 컴퓨터의 호스트 이름 가져오기
    local_ip = socket.gethostbyname(hostname)  # 호스트 이름 기반 IP 가져오기
    return local_ip

# IP 주소에 따른 경로 설정
local_ip = get_local_ip()

if local_ip == "192.168.0.29":  # 사무실에서 사용하는 IP
    file_path30 = r'C:\Users\erato2022\OneDrive\문서\2.금상동천호성지\2024기초데이터\통합조회_20241203금상동.xlsx'  # 엑셀 파일 경로를 지정하세요
elif local_ip != "192.168.0.29":  # 사무실이 아닌경우 IP
    file_path30 =r'C:\Users\erato\OneDrive\문서\2.금상동천호성지\2024기초데이터\통합조회_20241203금상동.xlsx'
else:  # 기본 경로 (기타 환경)
    file_path30 = "/mnt/default_path/"

#print(f"Current Path: {file_path30}")






#########################################################################################
#####    봉안프로그램 

# 엑셀 파일 불러오기
#file_path30 = r'C:\dev\myproj01\통합조회_20241123금상동.xlsx'  # 엑셀 파일 경로를 지정하세요
#file_path30 = r'C:\Users\erato2022\OneDrive\문서\2.금상동천호성지\2024기초데이터\통합조회_20241128금상동.xlsx'  # 엑셀 파일 경로를 지정하세요
#file_path30 =r'C:\Users\erato\OneDrive\문서\2.금상동천호성지\2024기초데이터\통합조회_20241128금상동.xlsx'
             

if local_ip == "192.168.0.29":  # 사무실에서 사용하는 IP
    #file_path30 = r'C:\Users\erato2022\OneDrive\문서\2.금상동천호성지\2024기초데이터\통합조회_20241128금상동.xlsx'  # 엑셀 파일 경로를 지정하세요
    output_file = r'C:\Users\erato2022\OneDrive\문서\2.금상동천호성지\2024기초데이터\통합조회_20241128금상동(관리선수금월수).xlsx'
elif local_ip != "192.168.0.29":  # 사무실이 아닌경우 IP
    #file_path30 =r'C:\Users\erato\OneDrive\문서\2.금상동천호성지\2024기초데이터\통합조회_20241128금상동.xlsx'
     output_file = r'C:\Users\erato\OneDrive\문서\2.금상동천호성지\2024기초데이터\통합조회_20241128금상동(관리선수금월수).xlsx'
else:  # 기본 경로 (기타 환경)
    file_path30 = "/mnt/default_path/"

df30 = pd.read_excel(file_path30)
df30.info()

#output_file = r'C:\Users\erato2022\OneDrive\문서\2.금상동천호성지\2024기초데이터\통합조회_20241128금상동(관리선수금월수).xlsx'

# 날짜가 유효하지 않다면 2월 28일로 수정하는 함수
def fix_invalid_date(date_str):
    try:
        # 날짜가 유효한지 확인
        date_obj = pd.to_datetime(date_str)
    except ValueError:
        # 유효하지 않은 날짜인 경우 처리 (윤년이 아닌 해의 2월 29일 등)
        if "02-29" in date_str:
            # 날짜가 2월 29일인 경우 2월 28일로 수정
            date_str = date_str.replace("02-29", "02-28")
            date_obj = pd.to_datetime(date_str)
        else:
            date_obj = pd.NaT
    return date_obj

# '관리시작열'과 '관리종료열'의 날짜를 변환하며 오류를 수정
df30['관리시작'] = df30['관리시작'].apply(fix_invalid_date)
df30['관리종료'] = df30['관리종료'].apply(fix_invalid_date)
df30['봉안일1'] = df30['봉안일1'].apply(fix_invalid_date)
df30['봉안일2'] = df30['봉안일2'].apply(fix_invalid_date)
df30['만기일1'] = df30['만기일1'].apply(fix_invalid_date)
df30['만기일2'] = df30['만기일2'].apply(fix_invalid_date)



# 연도를 입력받음
input_year = int(input("연도를 입력하세요 (예: 2023): "))
start_of_input_year = datetime(input_year, 1, 1)
end_of_input_year = datetime(input_year, 12, 31)



# 날짜 형식 변환
df30['봉안일1'] = pd.to_datetime(df30['봉안일1'],errors='coerce')
df30['봉안일2'] = pd.to_datetime(df30['봉안일2'],errors='coerce')
df30['만기일1'] = pd.to_datetime(df30['만기일1'],errors='coerce')
df30['만기일2'] = pd.to_datetime(df30['만기일2'],errors='coerce')

df30['관리시작'] = pd.to_datetime(df30['관리시작'],errors='coerce')
df30['관리종료'] = pd.to_datetime(df30['관리종료'],errors='coerce')

# 관리선수금총월수를 계산하는 함수
def calculate_total_months_with_extra_month(row):
    #def calculate_total_months(row):
    try:
        # 날짜 값이 모두 유효한 경우에만 계산
        if pd.notna(row['관리종료']) and pd.notna(row['관리시작']):
    
          if pd.notna(row['봉안일1']) or pd.notna(row['봉안일2']):
             #조건 A
             #if row['봉안일1'] == row['봉안일2'] == row['관리시작'] or row['봉안일1'] == row['봉안일2']:
             # 관리종료 - 관리시작의 월수 계산
              delta = relativedelta(row['관리종료'], row['관리시작'])
              total_months =delta.years * 12 + delta.months+1
              return total_months               
             
              
         
    except Exception as e:
        print(f"Error processing row: {row}, Error: {e}")
        return 0       
    
      # 조건을 만족하지 않는 경우 0 반환



# 회전관리선수금총월수를 계산하는 함수
def calculate_total_months_with_extra_month2(row):
    #def calculate_total_months(row):
    try:
        # 날짜 값이 모두 유효한 경우에만 계산
        if pd.notna(row['만기일1']) or  pd.notna(row['만기일2']):
    
          if pd.notna(row['봉안일1']) :
             #조건 A
             #if row['봉안일1'] == row['봉안일2'] == row['관리시작'] or row['봉안일1'] == row['봉안일2']:
             # 관리종료 - 관리시작의 월수 계산
              delta = relativedelta(row['만기일1'],row['관리종료'], )
              total_months =delta.years * 12 + delta.months
              return total_months               
             
          if pd.notna(row['봉안일2']) :
             #조건 A
             #if row['봉안일1'] == row['봉안일2'] == row['관리시작'] or row['봉안일1'] == row['봉안일2']:
             # 관리종료 - 관리시작의 월수 계산
              delta = relativedelta(row['만기일2'],row['관리종료'], )
              total_months =delta.years * 12 + delta.months
              return total_months     
         
    except Exception as e:
        print(f"Error processing row: {row}, Error: {e}")
        return 0       
    
      # 조건을 만족하지 않는 경우 0 반환


# 관리선수금 년중월수 계산 함수
def calculate_year_months(row):
    start_date = pd.to_datetime(row['관리시작'])
    end_date = pd.to_datetime(row['관리종료'])

    # 관리 시작 날짜가 입력된 연도보다 이전인 경우 처리
    if start_date < start_of_input_year:
        # 입력된 연도의 시작일부터 계산
        actual_start = start_of_input_year
    else:
        actual_start = start_date

    # 관리 종료 날짜와 입력된 연도의 종료일 중 빠른 날짜 선택
    actual_end = min(end_date, end_of_input_year)

    # 기간이 겹치지 않는 경우 0 반환
    if actual_end < actual_start:
        return 0

    # 월 수 계산
    months_in_year = (actual_end.year - actual_start.year) * 12 + actual_end.month - actual_start.month
    return months_in_year+1  # 시작과 종료 포함


# 연중회전관리선수금월수를 계산하는 함수
def calculate_total_months_with_extra_month3(row):
    #def calculate_total_months(row):
    try:
        # 날짜 값이 모두 유효한 경우에만 계산
        if pd.notna(row['만기일1']) or  pd.notna(row['만기일2']):
    
          if pd.notna(row['봉안일1']) :
            start_date = pd.to_datetime(row['관리종료'])
            end_date = pd.to_datetime(row['만기일1'])
          
            # 관리 시작 날짜가 입력된 연도보다 이전인 경우 처리
            if start_date < start_of_input_year:
             # 입력된 연도의 시작일부터 계산
               actual_start = start_of_input_year
            else:
               actual_start = start_date

               # 관리 종료 날짜와 입력된 연도의 종료일 중 빠른 날짜 선택
            actual_end = min(end_date, end_of_input_year)

            # 기간이 겹치지 않는 경우 0 반환
            if actual_end < actual_start:
              return 0
            months_in_year = (actual_end.year - actual_start.year) * 12 + actual_end.month - actual_start.month
            return months_in_year+1
            # 시작과 종료 포함





          if pd.notna(row['봉안일2']) :
            start_date = pd.to_datetime(row['관리종료'])
            end_date = pd.to_datetime(row['만기일2'])
          
            # 관리 시작 날짜가 입력된 연도보다 이전인 경우 처리
            if start_date < start_of_input_year:
             # 입력된 연도의 시작일부터 계산
               actual_start = start_of_input_year
            else:
               actual_start = start_date

               # 관리 종료 날짜와 입력된 연도의 종료일 중 빠른 날짜 선택
            actual_end = min(end_date, end_of_input_year)

            # 기간이 겹치지 않는 경우 0 반환
            if actual_end < actual_start:
              return 0# 월 수 계산
            months_in_year = (actual_end.year - actual_start.year) * 12 + actual_end.month - actual_start.month
            return months_in_year+1  # 시작과 종료 포함 #조건 A
            
    except Exception as e:
        print(f"Error processing row: {row}, Error: {e}")
        return 0       















# 연말잔여월수 계산 함수
# 입력된 연도의 종료일을 계산
end_of_input_year = datetime(input_year, 12, 31)

# 잔여월수 계산 함수
def calculate_remaining_months(row):
    end_date = pd.to_datetime(row['관리종료'])
    
    # 관리종료열이 입력된 연도의 종료일보다 빠른 경우 0 반환
    if end_date <= end_of_input_year:
        return 0
    
    # 관리종료열이 입력된 연도의 종료일보다 늦은 경우 잔여 월수 계산
    else:
        remaining_months = (end_date.year - end_of_input_year.year) * 12 + (end_date.month - end_of_input_year.month)
        return remaining_months + 1  # 종료 월을 포함한 월수 반환




# 3. 관리시작열에서 입력된 연도를 포함하는지 확인 후, 사용료수입 열에 값을 추가하거나 0으로 설정
def check_year(row):
    # 관리시작열에서 입력된 연도를 포함하면 사용료수입 값을 유지, 포함하지 않으면 0을 반환
    if str(input_year) in str(row['관리시작']):
        return row['사용수입']
    else:
        return 0

def check_year2(row):
    start_date = pd.to_datetime(row['관리시작'])
    end_date = pd.to_datetime(row['관리종료'])
    
    if start_date < start_of_input_year:
        # 입력된 연도의 시작일부터 계산
        return row['사용수입']
    else:
        return 0



# 각 행에 대해 함수 적용
df30['관리선수금전체월수'] = df30.apply(calculate_total_months_with_extra_month, axis=1)
df30['회전관리선수금전체월수'] = df30.apply(calculate_total_months_with_extra_month2, axis=1)
df30['년중회전관리선수금월수'] = df30.apply(calculate_total_months_with_extra_month3, axis=1)

# 년중월수열 추가
df30['년중월수'] = df30.apply(calculate_year_months, axis=1)

# 잔여월수열 추가
df30['잔여월수'] = df30.apply(calculate_remaining_months, axis=1)
#df30['연중관리수입']=df30['관리수입']/df30['관리선수금전체월수']*df30['년중월수']
df30['잔여관리선수금']=df30['관리수입']/df30['관리선수금전체월수']*df30['잔여월수']
# '사용료수입' 열을 업데이트
df30['사용료수입'] = df30.apply(check_year, axis=1)
df30['사용료경과분'] = df30.apply(check_year2, axis=1)
# 사용자로부터 연도 입력받기
input_year = input("연도를 입력하세요 (예: 2023): ")

# 계약일에 입력한 연도 포함 여부를 확인하는 함수
def contains_year(date_series, year):
    # 계약일이 NaN이 아닌 경우에만 문자열 변환 후 year 포함 여부 확인
    return date_series.dropna().astype(str).str.contains(year)

# 조건: 입력받은 연도가 2023인 경우
if input_year == input_year:
    # 새로운 열 이름을 생성: '2023-사용선수금'
    new_column_name = f"{input_year}-사용선수금"
    new_column_name2 = f"{input_year}-관리선수금"
    new_column_name3 = f"{input_year}-사용료수입"
    
    # 새로운 열을 데이터프레임에 추가하고 초기값은 None으로 설정
    df30[new_column_name] = None
    df30[new_column_name2] = None
    df30[new_column_name3] = None
    
    # '계약일' 열에 입력받은 연도를 포함하고, 나머지 열이 모두 결측치일 경우 마스크 정의
    mask = (contains_year(df30['계약일'], input_year)) & \
           (df30['봉안일1'].isna()) & \
           (df30['봉안일2'].isna()) & \
           (df30['관리시작'].isna()) & \
           (df30['관리종료'].isna())

    # '봉안일1' 열에 입력받은 연도를 포함하고, 나머지 열이 모두 결측치일 경우 마스크 정의
    mask2 = (contains_year(df30['관리시작'], input_year)) & \
            (df30['사용수입'] > 0)


    # '사용선수' 값이 있는 행에서 '2023-사용선수금' 열에 값을 넣음
    df30.loc[mask, new_column_name] = df30.loc[mask, '사용선수']
    df30.loc[mask, new_column_name2] = df30.loc[mask, '관리선수']
    df30.loc[mask2, new_column_name3] = df30.loc[mask2, '사용수입']

    # 결과 확인
    print(df30)
    df30=df30.sort_values(["봉안함"]) # 일자와 환율을 일자와 환율로 정렬
    # 새로운 파일로 저장
    output_file_path = f"C:\\dev\\myproj01\\금상동\\결과폴더\\금상동1205_{input_year}.xlsx"
    
    df30.to_excel(output_file_path, index=False)
    print(f"새로운 데이터가 {output_file_path} 파일로 저장되었습니다.")









##########################################################################
#        월별결산부통합  ##################################################


# 여러 개의 데이터프레임을 엑셀 파일로 내보내기 함수
def export_multiple_dfs_to_excel(dfs, sheet_names, file_name):
    """
    여러 개의 데이터프레임을 각각 다른 시트에 엑셀 파일로 내보내는 함수.
    : param dfs: 데이터프레임 리스트 
    : param sheet_names: 시트 이름 리스트 
    : param file_name: 저장할 엑셀 파일 이름
   """ 
    
    if len(dfs) != len(sheet_names) :

      raise ValueError("데이터프레임 리스트와 시트 이름 리스트의 길이가 같아야 합니다. ")


    with pd. ExcelWriter (file_name) as writer:
     for df, sheet_name in zip(dfs, sheet_names) :
        df. to_excel(writer, sheet_name=sheet_name,index=False)
    #Print(f"{file_name} 파일이 생성되었습니다.")




#"C:\Users\erato\OneDrive\문서\2.금상동천호성지\2024기초데이터\금상동월별결산부"
local_ip = get_local_ip()

if local_ip == "192.168.0.29":  # 사무실에서 사용하는 IP
    #file_path30 = r'C:\Users\erato2022\OneDrive\문서\2.금상동천호성지\2024기초데이터\통합조회_20241203금상동.xlsx'  # 엑셀 파일 경로를 지정하세요
    #input_folder=r"C:\Users\erato2022\OneDrive\문서\2.금상동천호성지\2024기초데이터\금상동월별결산부"
    #input_folder=r"C:\Users\erato2022\OneDrive\삼례토지\사진\바탕 화면\금상동2024"
    input_folder=r"C:\Users\erato2022\OneDrive\문서\2.금상동천호성지\2024기초데이터\금상동월별결산부\9월통합"  
    #input_folder=r"C:\dev\myproj01\금상동2024"
    #input_folder=r"C:\Users\erato2022\OneDrive\삼례토지\사진\바탕 화면\금상동2024\xlsx"
    #input_folder=r"C:\dev\myproj01\금상동2024\금상동2024-2"
elif local_ip != "192.168.0.29":  # 사무실이 아닌경우 IP
    #input_folder=r"C:\dev\myproj01\금상동2024"
    #input_folder=r"C:\dev\myproj01\금상동2024\금상동2024-2"
    #input_folder=r"C:\Users\erato2022\OneDrive\삼례토지\사진\바탕 화면\금상동2024\xlsx"
    input_folder=r"C:\Users\erato\OneDrive\문서\2.금상동천호성지\2024기초데이터\금상동월별결산부\9월통합"   
        #input_folder=r"C:\Users\erato\OneDrive\문서\2.금상동천호성지\2024기초데이터\금상동월별결산부"
    #file_path30 =r'C:\Users\erato\OneDrive\문서\2.금상동천호성지\2024기초데이터\통합조회_20241203금상동.xlsx'

else:  # 기본 경로 (기타 환경)
    
    file_path30 = "/mnt/default_path/"

#print(f"Current Path: {file_path30}")



raw_data_dir =Path(input_folder)
excel_files=raw_data_dir.glob('24년*')




total_df=pd.DataFrame() #빈 데이터프래임 생성
total_df2=pd.DataFrame() #빈 데이터프래임 생성
total_df3=pd.DataFrame() #빈 데이터프래임 생성
total_df4=pd.DataFrame() #빈 데이터프래임 생성
total_df5=pd.DataFrame() #빈 데이터프래임 생성











i=0
for excel_file in excel_files:
    i=i+1
   
    
    
    
    df =pd.read_excel(excel_file,sheet_name="사용선수금",skiprows=5,usecols="a:g")
    df.columns = ['날짜', '성명', '봉안함', '적요1', '차변', '대변', '잔액']
    #df['날짜'] = pd.to_datetime(df['날짜']).dt.strftime('%Y-%m-%d')
    
    df2 =pd.read_excel(excel_file,sheet_name="일반관리선수-최준철세무사",skiprows=5,usecols="a:g")
    df2.columns = ['날짜', '성명', '봉안함', '적요1', '차변', '대변', '잔액']
    #f2['날짜'] = pd.to_datetime(df2['날짜']).dt.strftime('%Y-%m-%d')
    
    df3 =pd.read_excel(excel_file,sheet_name="사용료수입",skiprows=5,usecols="a:g")
    df3.columns = ['날짜', '성명', '봉안함', '적요1', '차변', '대변', '잔액']
    #df3['날짜'] = pd.to_datetime(df3['날짜']).dt.strftime('%Y-%m-%d')
    
    try:
     # 엑셀 파일의 시트 목록을 가져옴
        xls = pd.ExcelFile(excel_file)
        print(f"'{excel_file}' 파일의 시트 목록: {xls.sheet_names}")
    
        # '회전관선' 시트가 있는지 확인
        if "회전관선" in xls.sheet_names:
          print(f"'{excel_file}' 파일에서 '회전관선' 시트를 읽습니다.")
          # 시트를 읽어 데이터프레임으로 저장
          df4 = pd.read_excel(excel_file, sheet_name="회전관선", skiprows=5, usecols="A:G")
          df4.columns = ['날짜', '성명', '봉안함', '적요1', '차변', '대변', '잔액']
          total_df4=total_df4._append(df4,ignore_index=True)
          print(df4.head())  # 데이터프레임 출력
        else:
          print(f"'{excel_file}' 파일에 '회전관선' 시트가 없습니다. 다른 작업을 진행합니다.")
    
    except ValueError as e:
      print(f"오류 발생: {e}")
    except Exception as e:
      print(f"파일 '{excel_file}' 처리 중 오류 발생: {e}")

    print("프로그램이 종료되었습니다.")
    
       
    
    
    
    
    
    df4 = pd.read_excel(excel_file, sheet_name="회전관선-최준철세무사", skiprows=5, usecols="A:G")
    df4.columns = ['날짜', '성명', '봉안함', '적요1', '차변', '대변', '잔액']

    


    sheet_name = f"24-{i}".strip()  # 공백 제거
    xls = pd.ExcelFile(excel_file)
    print(xls.sheet_names)  # 시트 목록 출력 
    




    df5= pd.read_excel(excel_file,sheet_name=sheet_name,skiprows=2,usecols="a:i")
    # 엑셀 시트 이름 목록 확인
   

   # 시트 존재 여부 확인 후 로드
    if sheet_name in xls.sheet_names:
     df5 = pd.read_excel(excel_file, sheet_name=sheet_name, skiprows=2, usecols="A:I")
    else:
     print(f"시트 '{sheet_name}'가 존재하지 않습니다.")
    
    
    
    
    
    total_df=total_df._append(df,ignore_index=True)	

    total_df2=total_df2._append(df2,ignore_index=True)

    total_df3=total_df3._append(df3,ignore_index=True)

    total_df4=total_df4._append(df4,ignore_index=True)

    total_df5=total_df5._append(df5,ignore_index=True)
    
   
    




total_df2['날짜'] = pd.to_datetime(total_df2['날짜'], errors='coerce')
total_df2['날짜'] = total_df2['날짜'].dt.strftime('%Y-%m-%d')

total_df3['날짜'] = pd.to_datetime(total_df3['날짜'], errors='coerce')
total_df3['날짜'] = total_df3['날짜'].dt.strftime('%Y-%m-%d')
total_df3['날짜']=total_df3['날짜'].ffill() # 최근일 월일없는행은 전일자로 처리

total_df4['날짜'] = pd.to_datetime(total_df4['날짜'], errors='coerce')
total_df4['날짜'] = total_df4['날짜'].dt.strftime('%Y-%m-%d')


total_df5= total_df5.dropna(subset=["적       요"]) #아래 누적액,월계액을 없애기위해 회계계정없는부분을 삭제함
# '월일' 열을 날짜(datetime) 형식으로 변환
total_df5['월일'] = pd.to_datetime(total_df5['월일'], errors='coerce')

# 변환 후 날짜 형식을 'YYYY-MM-DD' 형태의 문자열로 변경
#total_df5['월일'] = total_df5['월일'].dt.strftime('%Y-%m-%d')

total_df5['월일']=total_df5['월일'].dt.strftime('%Y-%m-%d') #구매자결제일이 날짜함수이므로 문자함수로 변환

total_df5['월일']=total_df5['월일'].str.replace('-','')

total_df5.info()
#total_df5['월일']=int(total_df5['월일'])




total_df5.columns = total_df5.columns.str.strip()

#total_df5= total_df5.dropna(subset=["적       요"]) #아래 봉안함번로 없는 행을 없애기위해 봉안함없는 부분을 삭제함
total_df5['월일']=total_df5['월일'].ffill() # 최근일 월일없는행은 전일자로 처리

#total_df5=total_df5['적요'].ffill()) # 최근일 월일없는행은 전일자로 처리
#total_df5['적       요']=total_df5['적       요'].ffill() # 최근일 월일없는행은 전일자로 처리


total_df=total_df.sort_values(["봉안함"]) # 봉안함번호로정렬
total_df2=total_df2.sort_values(["봉안함"]) # 봉안함번호로정렬
total_df3=total_df3.sort_values(["봉안함"]) # 봉안함번호로정렬
total_df4=total_df4.sort_values(["봉안함"]) # 봉안함번호로정렬


total_df= total_df.dropna(subset=["봉안함"]) #아래 봉안함번로 없는 행을 없애기위해 봉안함없는 부분을 삭제함
total_df2= total_df2.dropna(subset=["봉안함"]) #아래 봉안함번로 없는 행을 없애기위해 봉안함없는 부분을 삭제함
total_df3= total_df3.dropna(subset=["봉안함"]) #아래 봉안함번로 없는 행을 없애기위해 봉안함없는 부분을 삭제함
total_df4= total_df4.dropna(subset=["봉안함"]) #아래 봉안함번로 없는 행을 없애기위해 봉안함없는 부분을 삭제함




total_df.columns = ['날짜','성명','봉안함','적요1','차변','대변','잔액']
total_df2.columns = ['날짜','성명','봉안함','적요1','차변','대변','잔액']
total_df3.columns = ['날짜','성명','봉안함','적요1','차변','대변','잔액']
total_df4.columns = ['날짜', '성명','봉안함','적요1','차변','대변','잔액']
    


# 날짜 열을 먼저 datetime 형식으로 변환
total_df['날짜'] = pd.to_datetime(total_df['날짜'], errors='coerce')
total_df['날짜'] = total_df['날짜'].dt.strftime('%Y-%m-%d')

total_df2['날짜'] = pd.to_datetime(total_df2['날짜'], errors='coerce')
total_df2['날짜'] = total_df2['날짜'].dt.strftime('%Y-%m-%d')

total_df3['날짜'] = pd.to_datetime(total_df3['날짜'], errors='coerce')
total_df3['날짜'] = total_df3['날짜'].dt.strftime('%Y-%m-%d')

total_df4['날짜'] = pd.to_datetime(total_df4['날짜'], errors='coerce')
total_df4['날짜'] = total_df4['날짜'].dt.strftime('%Y-%m-%d')


total_df5['월일'] = pd.to_datetime(total_df5['월일'], errors='coerce')
total_df5['월일'] = total_df5['월일'].dt.strftime('%Y-%m-%d')

total_df['날짜']=total_df['날짜'].ffill() # 최근일 월일없는행은 전일자로 처리
total_df=total_df.sort_values(["봉안함"]) # 일자와 환율을 일자와 환율로 정렬
total_df6=total_df.sort_values(["날짜"]) # 일자와 환율을 일자와 환율로 정렬

    
total_df2['날짜']=total_df2['날짜'].ffill() # 최근일 월일없는행은 전일자로 처리
total_df2=total_df2.sort_values(["봉안함"]) # 일자와 환율을 일자와 환율로 정렬
total_df7=total_df2.sort_values(["날짜"]) # 일자와 환율을 일자와 환율로 정렬


#total_df3['날짜']=total_df3['날짜'].ffill()) # 최근일 월일없는행은 전일자로 처리
total_df3['날짜'] = total_df3['날짜'].ffill()
total_df3=total_df3.sort_values(["봉안함"]) # 일자와 환율을 일자와 환율로 정렬
total_df8=total_df3.sort_values(["날짜"]) # 일자와 환율을 일자와 환율로 정렬



#total_df4['날짜']=total_df4['날짜'].ffill()) # 최근일 월일없는행은 전일자로 처리
total_df4['날짜'] = total_df4['날짜'].ffill()
total_df4=total_df4.sort_values(["봉안함"]) # 일자와 환율을 일자와 환율로 정렬
total_df9=total_df4.sort_values(["날짜"]) # 일자와 환율을 일자와 환율로 정렬


#total_df = pd.concat([total_df, df1])
total_df.to_excel("C:\dev\myproj01\금상동\결과폴더\사용선수금2024.xlsx")
total_df2.to_excel("C:\dev\myproj01\금상동\결과폴더\일반관리선수-최준철세무사2024.xlsx")
total_df3.to_excel("C:\dev\myproj01\금상동\결과폴더\사용료수입2024.xlsx")
total_df4.to_excel("C:\dev\myproj01\금상동\결과폴더\회전일반관리선수-최준철세무사2024.xlsx")
total_df5.to_excel("C:\dev\myproj01\금상동\결과폴더\월별결산부2024.xlsx") 




#######################################################################################
# 엑셀 파일 읽어오기
#mon_df = pd.read_excel('C:\dev\myproj01\금상동\결과폴더\월별결산부2024.xlsx')  # 메인 엑셀 파일
mon_df = total_df5
mon_df['수입'] = mon_df['수입'].fillna(0)
mon_df['지출'] = mon_df['지출'].fillna(0)

# '구분1', '수입', '지출' 열에 NaN 값이 있는 행 제거
mon_df = mon_df.dropna(subset=['월일'])
mon_df['금액2']=mon_df['금액2']=0


# 금액 계산을 위한 함수 정의
mon_df['지출']=mon_df['지출'].astype(int)
mon_df['수입']=mon_df['수입'].astype(int)



mon_df['코드1'] = None
mon_df['구분1']=None
mon_df['구분2']=None
mon_df.info()



#mon_df.info()

# 두 문자열 칼럼을 합쳐서 새로운 칼럼에 저장
#mon_df['적요명'] = mon_df['적       요'].astype(str) + mon_df['봉안실'].astype(str)

import pandas as pd

# 1. 엑셀 파일에서 회계 데이터를 불러오기
#mon_df = pd.read_excel('accounting_data.xlsx')

if local_ip == "192.168.0.29":  # 사무실에서 사용하는 IP
    file_path31 =r'C:\Users\erato2022\OneDrive\문서\2.금상동천호성지\2024기초데이터\금상동계정코드(2024-11-21).xlsx' 
elif local_ip != "192.168.0.29":  # 사무실이 아닌경우 IP
    file_path31 =r'C:\Users\erato\OneDrive\문서\2.금상동천호성지\2024기초데이터\금상동계정코드(2024-11-21).xlsx' 
else:  # 기본 경로 (기타 환경)
    file_path30 = "/mnt/default_path/"




account_code_df = pd.read_excel(file_path31)  # 계정코드 엑셀 파일
 
account_code_df.info()

# '코드1'이라는 새로운 열 추가

# 조건에 따른 값을 계산하는 함수 정의
def calculate_gubun(row):
    E = str(row['납부방법'])  # E3608
    D = str(row['계정과목']) # D3608
    G =str(row['회계계정'])  # G3608
    H = row['수입']  # H3608
    I = row['지출']  # I3608

    if  ("이체" in E and "관리수입" in G and H < 10000) : #관리수입중 1만이상인 경우만 인식하기
        return 0 
    elif "현금" in E and H > 0:
        return 2
    elif ("대체" in E or "이체" in E) and "해약" in D and I > 0:
        return 4
    elif (
        ("이체" in E and H > 0) or
        ("이체" in E and "일반관리선수" in G and I > 0) or
        ("이체" in E and "회전관리" in G and I > 0) or
        ("이체" in E and "사용수입" in G and I > 0) or
       
        ("대체" in E and "사용수입" in G and H > 0) or
        ("대체" in E and "관리선수" in G and H > 0) or
        ("대체" in E and "관리선수" in G and I > 0) or
        ("대체" in E and "사용수입" in G and H > 0) or
        ("신용" in E and "사용선수" in G and H > 0) or
        ("신용" in E and "봉안함" in G and H > 0) 
    ):
        return 4
    else:
        return 0

# 각 행에 대해 함수 적용
#mon_df['구분1'] = mon_df.apply(calculate_gubun, axis=1)

#mon_df['구분1']=mon_df['구분1'].astype(float)
#mon_df = pd.read_excel('C:\dev\myproj01\금상동\결과폴더\월별결산부2024.xlsx')  # 메인 엑셀 파일
#mon_df['구분1'] = mon_df['구분1'].fillna(0)



mon_df = mon_df.dropna(subset=['월일'])
mon_df.loc[:, '구분1'] = mon_df.apply(calculate_gubun, axis=1)
mon_df.loc[:, '구분1'] = mon_df['구분1'].astype(float)
mon_df.loc[:, '구분1'] = mon_df['구분1'].fillna(0)






# 인덱스를 재설정
mon_df = mon_df.reset_index(drop=True)

mon_df = mon_df.dropna(subset=['월일'])



# 조건에 따른 값 설정
for index, row in mon_df.iterrows():
    # 마지막 행 처리 방지
    if index + 0 >= len(mon_df):
        continue
    
    # 첫번째 조건
    if "해약" in str(row['계정과목']) and row['납부방법'] == "이체" and (row['회계계정'] in ["관리선수", "사용선수"]) and row['지출'] > 0:
        mon_df.at[index, '코드1'] = 103
    # 두번째 조건
    elif "해약" in str(row['계정과목']) and row['회계계정'] == "사용수입" and row['지출'] > 0:
        mon_df.at[index, '코드1'] = 420
    # 세번째 조건
    elif "관리비" in str(row['계정과목']) and "관리수입" in str(row['회계계정']) and "일반관리선수" in str(mon_df.at[index+1, '회계계정']) and row['수입'] > 0:
        mon_df.at[index, '코드1'] = 274
    # 네번째 조건
    elif "관리비" in str(row['계정과목']) and "회전관리선수" in str(mon_df.iloc[index+1]['회계계정']) and row['수입'] > 0:
        mon_df.at[index, '코드1'] = 275
    elif "관리수입" in str(row['계정과목']) and "회전관리선수" in str(mon_df.iloc[index+1]['회계계정']) and row['수입'] > 0:
        mon_df.at[index, '코드1'] = 275   
    # 다섯번째 조건
    elif "대체" in str(row['납부방법']) and "관리선수" in str(row['회계계정']) and row['지출'] > 0:
        mon_df.at[index, '코드1'] = 421
    
    elif "신용" in str(row['납부방법']) and "사용선수" in str(row['회계계정']) and row['수입'] > 0:
        mon_df.at[index, '코드1'] = 273
       # 여섯번째 조건
    elif "이체" in str(row['납부방법']) and row['회계계정'] == "관리수입":
        mon_df.at[index, '코드1'] = 274
    # 일곱번째 조건
    elif row['수입'] == 0:
        mon_df.at[index, '코드1'] = 0
    # VLOOKUP 유사 작업
    else:
        lookup_value = str(row['회계계정'])[:6]  # G열의 첫 6글자 추출
        code_row = account_code_df[account_code_df['계정코드'].fillna('').str.startswith(lookup_value)]
        
        
        
        #code_row = account_code_df[account_code_df['계정'].str.startswith(lookup_value)]
        if not code_row.empty:
            mon_df.at[index, '코드1'] = code_row.iloc[0]['코드1']
        else:
            mon_df.at[index,'코드1'] = 0  # 일치하는 값이 없으면 0 처리
    
    
    
#mon_df = mon_df.dropna(subset=['월일'])    
# 결과를 엑셀 파일로 저장
# 3. 병합된 결과를 엑셀 파일로 저장
mon_df.to_excel('C:\dev\myproj01\금상동\결과폴더\월별병합엑셀파일9-20(10).xlsx', index=False)

mon_df=mon_df.dropna(subset=['월일'])  
# 월일 열에서 값이 비어 있는 행 삭제


# 2. VLOOKUP 유사 기능: '코드1' 열을 기준으로 lookup_df의 4번째 열과 병합
# lookup_df의 4번째 열의 이름을 임시로 설정
lookup_df_column_name = account_code_df.columns[5]  # 7번째 열의 이름 가져오기

# 병합 수행

print(type(mon_df))  # DataFrame인지 확인
#print(type(account_code_df))  # DataFrame인지 확인


mon_df=mon_df.dropna(subset=['월일']) 

mon_df.to_excel('C:\dev\myproj01\금상동\결과폴더\코드병합전mon_df.xlsx', index=False)
# '월일' 값이 있는 경우만 필터링
#filtered_mon_df = mon_df[mon_df['월일'].notna()] #1225-10
filtered_mon_df = mon_df[mon_df['월일'].notna() & mon_df['회계계정'].notna()] # 필터병합결과
filtered_mon_df.to_excel('C:\dev\myproj01\금상동\결과폴더\필터후구분1.xlsx', index=False)
# 병합 수행
 
merged_df = pd.merge(filtered_mon_df,account_code_df[['코드2', '계정코드2']],how='left',left_on='코드1',right_on='코드2')
merged_df.to_excel('C:\dev\myproj01\금상동\결과폴더\코드병합후.xlsx', index=False)



#######################################################################################################################################





# 3. 불필요한 '코드2' 열 제거 (선택 사항)
merged_df.drop(columns=['코드2'], inplace=True)
merged_df=merged_df.dropna(subset=['월일'])  
#mon_df=mon_df['적       요'].ffill()) # 최근일 월일없는행은 전일자로 처리

merged_df.to_excel('C:\dev\myproj01\금상동\결과폴더\월별병합엑셀파일9-21.xlsx', index=False)

#merged_df.iloc[:, 3] = merged_df.iloc[:, 3].ffill()


# 적요를 적요명으로 봉안실을 봉안함으로 변경
merged_df['적요명'] = merged_df.iloc[:, 2].astype(str) 
merged_df['봉안함'] = merged_df.iloc[:, 5].astype(str)
# 3. 병합된 결과를 엑셀 파일로 저장
merged_df.to_excel('C:\dev\myproj01\금상동\결과폴더\월별병합코드엑셀파일9-20.xlsx', index=False)




merged_df.info()

def calculate_amount(row):
    수입 = row['수입']
    지출 = row['지출']
    구분1 = float(row['구분1'])
    계정과목 = str(row['계정과목'])
    납부방법 = str(row['납부방법'])
    회계계정 = str(row['회계계정'])

    if 구분1 == 0:
        return (수입 + 지출) * 구분1

    # '해약' 관련 조건
    if '해약' in 계정과목 and 지출 > 0:
        if '이체' in 납부방법 and ('관리선수' in 회계계정 or '사용선수' in 회계계정):
            return 수입 + 지출
        return 수입 + 지출 * -1

    # '사용료' 관련 조건
    if '사용료' in 계정과목 and '현금' in 납부방법 and ('사용수입' in 회계계정 or '관리수입' in 회계계정) and 수입 > 0:
        return 수입 + 지출

    # '관리비' 관련 조건
    if '관리비' in 계정과목 and '이체' in 납부방법 and ('관리수입' in 회계계정 or '사용선수' in 회계계정) and 수입 > 0:
        return 수입 + 지출

    # 기본 계산
    return 수입 + 지출

# '금액' 열을 계산하여 추가
merged_df['금액'] = merged_df.apply(calculate_amount, axis=1)


def calculate_amount2(row):
    수입 = row['수입']
    지출 = row['지출']
    구분1 = float(row['구분1'])
    계정과목 = str(row['계정과목'])
    납부방법 = str(row['납부방법'])
    회계계정 = str(row['회계계정'])

    if 구분1 == 0:
        return (수입 + 지출) * 구분1

    # '해약' 관련 조건
    if '해약' in 계정과목 and 지출 > 0:
        if '이체' in 납부방법 and ('관리선수' in 회계계정 or '사용선수' in 회계계정):
            return 수입 + 지출
        return 수입 + 지출 * 1

    # '사용료' 관련 조건
    if '사용료' in 계정과목 and '현금' in 납부방법 and ('사용수입' in 회계계정 or '관리수입' in 회계계정) and 수입 > 0:
        return 수입 + 지출

    # '관리비' 관련 조건
    if '관리비' in 계정과목 and '이체' in 납부방법 and ('관리수입' in 회계계정 or '사용선수' in 회계계정) and 수입 > 0:
        return 수입 + 지출

    # 기본 계산
    return 수입 + 지출

# '금액' 열을 계산하여 추가
merged_df['금액2'] = merged_df.apply(calculate_amount2, axis=1)




# 데이터프레임에 '금액' 열 추가
#mon_df['금액'] = mon_df.apply(calculate_amount, axis=1)

merged_df['지출'] =merged_df['지출'].fillna(0)

merged_df['수입'] =merged_df['수입'].fillna(0)

merged_df['금액'] = merged_df['금액'].fillna(0)

merged_df['금액']=merged_df['금액'].astype(int)






# 코드2 값을 계산하는 함수 정의

def calculate_code2(row):
    if row['구분1'] == 0:
        return 0

    # '납부방법' 및 '회계계정' 조건을 변수로 처리
    납부방법 = str(row['납부방법'])
    회계계정 = str(row['회계계정'])
    계정과목 = str(row.get('계정과목', ''))  # '계정과목'이 없을 경우 빈 문자열

    # 지출이 발생한 경우
    if row['지출'] > 0:
        if '이체' in 납부방법:
            if '일반관리' in 회계계정:
                return 274
            elif '회전관리선수' in 회계계정 or '관리선수' in 회계계정:
                return 274
            elif '회전' in 회계계정:
                return 275
            elif '사용수입' in 회계계정 or '사용선수' in 회계계정 :
                return 273
        elif '대체' in 납부방법:
            if '관리선수' in 회계계정 or '관리수입' in 회계계정:
                return 274
            elif '사용수입' in 회계계정:
                return 273

    # 수입이 발생한 경우
    if row['수입'] > 0:
        if '이체' in 납부방법 and '관리수입' in 회계계정 and row['수입'] < 10000:
            return 0  # C50 외상매출금 참조
        elif '이체' in 납부방법:
            return 108  # C20 미수금 참
        
        elif '신용' in 납부방법:
            return 120  # C20 미수금 참조
        elif '현금' in 납부방법:
            return 103  # C73 현금 참조

    # '대체' 납부방법과 '해약' 또는 '사용선수' 관련 조건
    if '대체' in 납부방법:
        if '해약' in 계정과목 or '사용선수' in 회계계정:
            return row['코드1']
    elif '이체' in 납부방법 and '해약' in 계정과목:
        return row['코드1']

    # 기본값 반환
    return 0

# '코드2' 열을 계산하여 추가
merged_df['코드2'] = merged_df.apply(calculate_code2, axis=1)





# 엑셀 파일 불러오기


if local_ip == "192.168.0.29":  # 사무실에서 사용하는 IP
    file_path32 =r'C:\Users\erato2022\OneDrive\문서\2.금상동천호성지\2024기초데이터\금상동계정코드(2024-11-21).xlsx' 
elif local_ip != "192.168.0.29":  # 사무실이 아닌경우 IP
    file_path32 =r'C:\Users\erato\OneDrive\문서\2.금상동천호성지\2024기초데이터\금상동계정코드(2024-11-21).xlsx' 
else:  # 기본 경로 (기타 환경)
    file_path30 = "/mnt/default_path/"


code_df = pd.read_excel(file_path32)
# 금상동 계정코드 파일을 딕셔너리로 변환
code_dict = code_df.set_index(code_df.columns[0]).to_dict()[code_df.columns[1]]



def calculate_gubun2(row):
    # 납부방법과 회계계정, 계정과목 변수 정의
    납부방법 = str(row['납부방법'])
    회계계정 = str(row['회계계정'])
    계정과목 = str(row.get('계정과목', ''))  # '계정과목'이 없을 경우 빈 문자열

    # '대체' 관련 조건
    if '대체' in 납부방법:
        if '해약' in 계정과목 or '사용선수' in 회계계정:
            return 3
        if '사용수입' in 회계계정 and row['지출'] > 0:
            return 3
        if '관리선수' in 회계계정 and row['지출'] > 0:
            return 3

    # '이체' 관련 조건
    if '이체' in 납부방법 and '해약' in 계정과목:
        return 3

    # 코드1이 0일 때
    if row['코드1'] == 0:
        return 0

    # 구분1이 2일 때
    if row['구분1'] == 2:
        return 0

    # 그 외의 경우
    return row['구분1'] - 1

# '구분2' 열을 계산하여 추가
merged_df['구분2'] = merged_df.apply(calculate_gubun2, axis=1)




code_df1=code_df[['코드2','계정코드2']]

# VLOOKUP 함수 구현
def vlookup(lookup_value, code_df1, col_index, range_lookup=False):
    # 'range_lookup'에 따라 정확히 일치하는 값 또는 근사치를 찾을 수 있지만,
    # 현재는 정확한 일치를 기준으로 구현
    result = code_df1[code_df1.iloc[:, 0] == lookup_value]
    if not result.empty:
        return result.iloc[0, col_index - 1]  # col_index는 1부터 시작
    return 0 if not range_lookup else "Approximate Not Found"



# VLOOKUP을 적용할 데이터
lookup_col = '코드2'  # VLOOKUP에 사용할 열 이름
value_col = '계정코드2'      # 반환할 열 이름
result_col = '계정코드3'  # 결과를 저장할 열 이름

# VLOOKUP 함수 적용
merged_df[result_col] = merged_df[lookup_col].apply(lambda x: vlookup(x, code_df1, 2))

#merged_df['금액2']=None
#merged_df['금액2']=merged_df['금액'].astype(int)


# 금액2 열을 맨 끝으로 이동시킵니다
# 현재 열 순서를 얻습니다
columns = list(merged_df.columns)

# 금액2가 맨 끝으로 오도록 열 순서를 재배열합니다
new_columns_order = [col for col in columns if col != '금액2'] + ['금액2']

# DataFrame을 새로운 열 순서로 재정렬합니다
merged_df=merged_df[new_columns_order]





merged_df['번호']=merged_df['Unnamed: 1']


print(merged_df)
#merged_df.info()

merged_df.to_excel('C:\dev\myproj01\금상동\결과폴더\금상동월별결산부2024업로드20.xlsx', index=False)
upload_df = pd.DataFrame()
dae_df=merged_df[['번호','봉안함','월일','구분1','코드1','계정코드2','적요명','금액']]
cha_df=merged_df[['번호','봉안함','월일','구분2','코드2','계정코드3','적요명','금액2']]
cha_df = cha_df.copy()  # cha_df를 명시적으로 복사
cha_df['구분1']=cha_df['구분2']
cha_df['코드1']=cha_df['코드2']
cha_df['계정코드2']=cha_df['계정코드3']
cha_df['금액']=cha_df['금액2']


#merged_df = pd.merge(mon_df, account_code_df[['코드2','계정코드2']], how='left',left_on='코드1',right_on='코드2')

#upload_df = pd.merge(cha_df, dae_df, how='left',left_on='구분1',right_on='구분1')



cha_df2=cha_df[['번호','봉안함','월일','구분1','코드1','계정코드2','적요명','금액']]







#dae_df.info()
cha_df2.to_excel('C:\dev\myproj01\금상동\결과폴더\차변업로드.xlsx', index=False)
dae_df.to_excel('C:\dev\myproj01\금상동\결과폴더\대변업로드.xlsx', index=False)
#upload_df = pd.merge(cha_df, dae_df[['월일','구분2','코드2','계정코드3']], how='left',left_on='월일',right_on='코드2')

#merged_df = pd.merge(mon_df, account_code_df[['코드2','계정코드2']], how='left',left_on='코드1',right_on='코드2')




# 데이터프레임 1과 2의 행 수가 같은지 확인합니다
#assert len(cha_df) == len(dae_df), "두 데이터프레임의 행 수가 다릅니다."

# 빈 데이터프레임을 만듭니다
upload_df = pd.DataFrame()

# 교차로 행을 추가합니다
for i in range(len(cha_df)):
    upload_df= pd.concat([upload_df, cha_df.iloc[[i]], dae_df.iloc[[i]]], ignore_index=True)

#upload_df['차대']=cha_df['구분1']+dae_df['구분2']
#upload_df=upload_df.drop(columns=['구분1'], inplace=True)
#upload_df.info()



#mon_df['구분1'] > 1) & (mon_df['코드1'] > 0)


#mon_df = mon_df[mon_df['구분1'] >= 1]
#upload_df['구분1']=upload_df['구분1'].astype(int)

upload_df=upload_df[['월일','구분1','코드1','계정코드2','적요명','금액','봉안함']]


#upload_df2.info()

upload_df2= upload_df[(upload_df['구분1'] >= 1) & (upload_df['코드1'] > 0)] #구분1 값이 1보다 큰값은 남기고 코드1값이 0 이면 삭제하기위함


#upload_df2= upload_df[upload_df['구분1'] >= 1]


upload_df2.info()

upload_df2.insert(4, '코드4', '')      # 6th column, index 5
upload_df2.insert(5, '거래처명', '')   # 7th column, index 6
upload_df2.insert(6, '코드5', '')      # 8th column, index 7








#upload_df2.to_excel('C:\dev\myproj01\금상동\결과폴더\금상동세무사랑2024업로드920.xlsx', index=False)


# 1. 엑셀 파일 불러오기
#input_file = 'C:\dev\myproj01\금상동\결과폴더\금상동세무사랑2024업로드920.xlsx'  # 여기에 불러올 엑셀 파일의 경로를 입력하세요.
#df7 = pd.read_excel(input_file)
df7 = upload_df2

df7.rename(columns={'월일' : '년도월일'}, inplace=False)
df7.rename(columns={'계정코드2' : '계정과목'}, inplace=False)
df7.rename(columns={'코드4' : '코드'}, inplace=False)







# 2. 데이터프레임을 엑셀 파일로 저장 (임시 저장)
output_file = 'C:\dev\myproj01\금상동\결과폴더\금상동세무사랑기타업로드수정.xlsx'  # 여기에 저장할 엑셀 파일의 경로를 입력하세요.
df7.to_excel(output_file, index=False)

# 3. openpyxl을 사용하여 엑셀 파일 열기
wb = load_workbook(output_file)
ws = wb.active

# 4. 열 너비 설정
ws.column_dimensions['D'].width = 20  # '구분1' 열 (가정: A열에 위치)
ws.column_dimensions['H'].width = 25  # '적요명' 열 (가정: B열에 위치)

# 5. 엑셀 파일 저장
wb.save(output_file)

#

















#####################################################################################
#total_df4=total_df4.drop(columns=["잔액"])

#total_df4=total_df4.drop(columns=["차변"], inplace=True)


 # 엑셀 파일 불러오기
file_path3 = r"C:\dev\myproj01\금상동세무사랑회전관리선수금2022-2023.xls"   # 엑셀 파일 경로를 지정하세요
df17= pd.read_excel(file_path3,sheet_name="회전관리선수금2022-2023봉안함순",skiprows=0,usecols="a:e") 

df17=df17.rename(columns={"일자":"날짜"})

df17=df17.rename(columns={"계약자":"성명"})
df17=df17.rename(columns={"대    변":"대변"})
df17["잔액"]=0
df17.info()



df17.to_excel(r'C:\dev\myproj01\금상동\결과폴더\회전관리선수금2022-2023봉안함순2.xlsx') 



df20=pd.concat([total_df4,df17])
df20['날짜'] =pd.to_datetime(df20['날짜'])
df20['날짜']=df20['날짜'].dt.strftime('%y-%m-%d') #날짜를 날짜함수이므로 문자함수로 변환
df20=df20.sort_values("봉안함")
df20.to_excel(r'C:\dev\myproj01\금상동\결과폴더\회전관리선수금2022-2023봉안함순.xlsx')







# 데이터프레임과 시트 이름 리스트
dfs = [total_df, total_df2,total_df3,total_df4,df20,merged_df,total_df6,total_df7,total_df8,total_df9]











sheet_names = ['사용선수금', '일반관리선수-최준철세무사','사용료수입','회전관리선수금','회전관리선수금통합-봉안함순','월별결산부','사용선수금-일자순','일반관리선수-최준철세무사-일자순','사용료수입-일자순','회전관리선수금-일자순']
# 함수 호출
#export_multiple_dfs_to_excel(dfs, sheet_names,'C:\dev\myproj01\금상동\결과폴더\월별결산부사용선수금-관리선수금-사용료수입-회전관리선수금2024.xlsx')

export_multiple_dfs_to_excel(dfs, sheet_names,'C:\dev\myproj01\금상동\결과폴더\월별결산부+회전관리선수금통합.xlsx')

# 엑셀 파일 열기
#file_path = r'C:\dev\myproj01\금상동\결과폴더\월별결산부사용선수금-관리선수금-사용료수입-회전관리선수금2024.xlsx'  # 엑셀 파일 경로를 입력하세요
file_path = r'C:\dev\myproj01\금상동\결과폴더\월별결산부+회전관리선수금통합.xlsx'  # 엑셀 파일 경로를 입력하세요
wb = openpyxl.load_workbook(file_path)

# 모든 시트에 대해 A열부터 D열까지의 너비를 25로 설정
columns_to_resize = ['A', 'B', 'C', 'D','E','F','G']  # A열부터 D열까지 지정

for sheet in wb.worksheets:
    for column_letter in columns_to_resize:
        # 지정된 열(column_letter)의 너비를 25로 설정
        sheet.column_dimensions[column_letter].width = 15



# 통화 기호 없는 회계 스타일 정의
accounting_style = NamedStyle(name="accounting_style", number_format='#,##0;[Red](#,##0)')

# 모든 시트에 대해 E열부터 G열까지 회계 스타일 적용
for sheet in wb.worksheets:
    for row in range(1, sheet.max_row + 1):
        for col in ['E', 'F', 'G']:  # E열부터 G열까지
            cell = sheet[f'{col}{row}']
            cell.style = accounting_style








# 수정된 엑셀 파일 저장

# 수정된 엑셀 파일 저장 (경로와 파일명 수정)
#save_path = r'C:\dev\myproj01\금상동\결과폴더\월별결산부사용선수금-관리선수금-사용료수입-회전관리선수금(열조정2024).xlsx'
save_path = r'C:\dev\myproj01\금상동\결과폴더\월별결산부+회전관리선수금통합(열조정2024).xlsx'
wb.save(save_path)












total_df.info()
			
import pandas as pd
from openpyxl import load_workbook

# 파일 경로 설정
#file_path = r'C:\dev\myproj01\금상동\결과폴더\월별결산부사용선수금-관리선수금-사용료수입-회전관리선수금(열조정2024).xlsx'  # 여기에 엑셀 파일 경로 입력
file_path = r'C:\dev\myproj01\금상동\결과폴더\월별결산부+회전관리선수금통합(열조정2024).xlsx'  # 여기에 엑셀 파일 경로 입력



# 엑셀 파일 불러오기 (openpyxl 사용)
wb = load_workbook(file_path)
#sheet_names = ['사용선수금', '일반관리선수-최준철세무사', '사용료수입', '회전관리선수금']  # 작업할 시트 이름들
sheet_names = ['사용선수금', '일반관리선수-최준철세무사','사용료수입','회전관리선수금','회전관리선수금통합-봉안함순','월별결산부','사용선수금-일자순','일반관리선수-최준철세무사-일자순','사용료수입-일자순','회전관리선수금-일자순']  # 작업할 시트 이름들

#'사용선수금', '일반관리선수-최준철세무사','사용료수입','회전관리선수금','월별결산부','사용선수금-일자순','일반관리선수-최준철세무사-일자순','사용료수입-일자순','회전관리선수금-일자순','회전관리선수금누적(봉안함순)'
# 각 시트에 대해 작업 수행
for sheet_name in sheet_names:
    # pandas로 시트를 읽어오기
    df21 = pd.read_excel(file_path, sheet_name=sheet_name)
    
    if '계정과목' not in df21.columns:
        df21['계정과목'] = pd.NA

       # '계정과목'의 결측치 처리: 결측치(NaN)인 경우 시트 이름으로 채움
        df21['계정과목'] = df21['계정과목'].fillna(sheet_name)
    



    # "계정과목" 열을 추가하고 시트 이름으로 채우기
    #df21['계정과목'] = sheet_name
    
    # pandas로 수정된 데이터를 다시 엑셀 파일로 저장
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df21.to_excel(writer, sheet_name=sheet_name, index=False)

print("모든 시트에 '계정과목' 열 추가 완료.")

import pandas as pd

# 파일 경로 설정
#a_file = r'C:\dev\myproj01\금상동\결과폴더\금상동1205_2024.xlsx'  # a파일 경로
a_file = f"C:\\dev\\myproj01\\금상동\\결과폴더\\금상동1205_{input_year}.xlsx"  # a파일 경로

   #f"C:\\dev\\myproj01\\금상동\\결과폴더\\금상동1205_{input_year}.xlsx"
#b_file = r'C:\dev\myproj01\금상동\결과폴더\월별결산부사용선수금-관리선수금-사용료수입-회전관리선수금(열조정2024).xlsx'  # b파일 경로
b_file = r'C:\dev\myproj01\금상동\결과폴더\월별결산부+회전관리선수금통합(열조정2024).xlsx'   # b파일 경로

# 1. a 파일 불러와서 데이터프레임 a 생성
df_a = pd.read_excel(a_file)
#df_a = df30
#df_b = pd.read_excel(b_file)
# 2. b 파일의 첫 번째 시트부터 네 번째 시트까지 데이터프레임 생성
b_sheets = [pd.read_excel(b_file, sheet_name=i) for i in range(5)]

# 3. 첫 번째 시트를 a 데이터프레임에 병합하여 a-1 데이터프레임 생성
df_a1 =df_a


# 4. 두 번째, 세 번째, 네 번째 시트를 차례로 병합
for i in range(1, 5):
    b_sheets[i] = b_sheets[i].drop_duplicates(subset=['봉안함'])  # 봉안함 중복 제거 
    df_a1 = pd.merge(df_a1, 
                     b_sheets[i][['봉안함', '차변', '대변', '계정과목','날짜']], 
                     on='봉안함', 
                     how='left',
                     suffixes=('', f'_b{i}'))


# 4. 병합된 결과를 엑셀 파일로 저장
# 날짜를 문자열로 포맷 (예: '2024-10-03')
today = datetime.today().strftime('%Y-%m-%d')

output_file = f'C:\dev\myproj01\금상동\결과폴더\금상동봉안+월별결산부병합결과-{today}.xlsx'


df_a1=df_a1.rename(columns={"대변_b4":"누적-회전관리선수금"})
df_a1=df_a1.rename(columns={"대변_b3":"2024-회전관리선수금"})
df_a1=df_a1.rename(columns={"차변_b2":"사용료수입"})

df_a1['봉안일1'] = df_a1['봉안일1'].dt.strftime('%Y-%m-%d')
df_a1['봉안일2'] = df_a1['봉안일2'].dt.strftime('%Y-%m-%d')
df_a1['만기일1'] = df_a1['만기일1'].dt.strftime('%Y-%m-%d')
df_a1['만기일2'] = df_a1['만기일2'].dt.strftime('%Y-%m-%d')
df_a1['관리시작'] = df_a1['관리시작'].dt.strftime('%Y-%m-%d')
df_a1['관리종료'] = df_a1['관리종료'].dt.strftime('%Y-%m-%d')


#df_a1['관리수입-누적회전관리선수금']=df_a1['관리수입'] - df_a1['누적-회전관리선수금']

 # 벡터화된 조건 사용
df_a1['연중관리수입2'] = 0  # 기본값 0으로 초기화
df_a1['연중관리수입2'] = df_a1['연중관리수입2'].astype(float)
mask3 = df_a1['관리수입'].notna() & df_a1['누적-회전관리선수금'].isna()
df_a1.loc[mask3, '연중관리수입2'] = (
    df_a1['관리수입'] / df_a1['관리선수금전체월수'] * df_a1['년중월수'])
    
#mask4 = df_a1['관리수입'] > df_a1['누적-회전관리선수금'] or df_a1['누적-회전관리선수금'],isna()
mask4 = (df_a1['관리수입'].notna() & df_a1['누적-회전관리선수금'].notna() & (df_a1['관리수입'] > df_a1['누적-회전관리선수금']))
df_a1.loc[mask4, '연중관리수입2'] = (
    (df_a1['관리수입']-df_a1['누적-회전관리선수금'])/df_a1['관리선수금전체월수']*df_a1['년중월수'])     
#df_a1['관리수입'] <= df_a1['누적-회전관리선수금']:
    #df_a1['연중관리비수입']=(df_a1['관리수입']/df_a1['관리선수금전체월수'])*df_a1['년중월수'] 
    #df_a1['연중관리비수입']=df_a1['관리수입'] 
    #df_a1['연중관리비수입2']=(df_a1['관리수입']/df_a1['관리선수금전체월수'])*df_a1['년중월수'] 
#df_a1['연중관리수입']=(df_a1['관리수입']-df_a1['누적-회전관리선수금'])/df_a1['관리선수금전체월수']*df_a1['년중월수']
df_a1['연중회전관리수입']=(df_a1['누적-회전관리선수금'])/df_a1['회전관리선수금전체월수']*df_a1['년중회전관리선수금월수']

df_a1['연중회전관리수입'] = pd.to_numeric(df_a1['연중회전관리수입'], errors='coerce').fillna(0).astype(int)
df_a1['연중관리수입2'] = pd.to_numeric(df_a1['연중관리수입2'], errors='coerce').fillna(0).astype(int)





df_a1.to_excel(output_file, index=False)

print(f"병합된 데이터가 {output_file}로 저장되었습니다.")


# 엑셀 파일 열기
file_path = f'C:\dev\myproj01\금상동\결과폴더\금상동봉안+월별결산부병합결과-{today}.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb.active  # 활성화된 시트를 선택
# 모든 시트에 대해 A열부터 D열까지의 너비를 25로 설정
columns_to_resize = ['A', 'B', 'C', 'D','E','F','G','Y', 'Z','AA', 'AB','AC','AD','AH', 'AI','AJ', 'AK','AL','AM','AN']  # A열부터 D열까지 지정

for sheet in wb.worksheets:
    for column_letter in columns_to_resize:
        # 지정된 열(column_letter)의 너비를 25로 설정
        sheet.column_dimensions[column_letter].width = 15



       # 통화 기호 없는 회계 스타일 정의
        accounting_style = NamedStyle(name="accounting_style", number_format='#,##0;[Red](#,##0)')

# 모든 시트에 대해 E열부터 G열까지 회계 스타일 적용
    for sheet in wb.worksheets:
      for row in range(1, sheet.max_row + 1):
        for col in ['Y', 'Z','AA', 'AB','AC','AD','AH', 'AI','AJ', 'AK','AL','AM','AN']:  # E열부터 G열까지
            cell = sheet[f'{col}{row}']
            cell.style = accounting_style

    
    # '전체월수' 열의 값이 180을 초과하는 경우 색상을 빨간색으로 변경
    fill_red = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

    # 전체월수 열의 값을 차례대로 불러와서 처리
    for idx, value in enumerate(df_a1['관리선수금전체월수'], start=2):  # start=2는 헤더를 건너뛰기 위해
        if value > 180:
            ws[f'AE{idx}'].fill = fill_red  # C열에 '전체월수'가 있다고 가정
    
    
    
    
    # 수정된 엑셀 파일 저장
    # A2 셀 기준으로 틀 고정 (freeze panes)
    ws.freeze_panes = 'A2'

    # 첫 번째 행에 필터 추가
    ws.auto_filter.ref = ws.dimensions  # ws.dimensions는 사용 중인 셀의 범위를 자동으로 가져옴

    # 수정된 엑셀 파일 저장 (경로와 파일명 수정)
    
    today = datetime.today().strftime('%Y-%m-%d')

    
    
    save_path = f'C:\dev\myproj01\금상동\결과폴더\금상동봉안+월별결산부병합결과-{today}.xlsx' 
    wb.save(save_path)
    print(f"새로운 데이터가 {save_path} 파일로 저장되었습니다.")        

    # 첫 번째 파일 (기존 파일)
file1 = r'C:\dev\myproj01\금상동\결과폴더\월별결산부+회전관리선수금통합(열조정2024).xlsx'

# 두 번째 파일 (추가할 파일)
file2 = r'C:\dev\myproj01\금상동\결과폴더\금상동세무사랑기타업로드수정.xlsx'

# 파일을 읽어들임
dfs_file2 = pd.read_excel(file2, sheet_name=None)  # 두 번째 파일의 모든 시트를 딕셔너리로 읽기

# 엑셀 파일의 기존 워크북 불러오기
with pd.ExcelWriter(file1, engine="openpyxl", mode="a") as writer:
    for sheet_name, df in dfs_file2.items():  # 두 번째 파일의 시트들 순회
        df.to_excel(writer, sheet_name="엑셀기타업로드", index=False)  # 새로운 시트로 추가

print(f"'{file2}' 파일의 모든 시트가 '{file1}' 파일에 추가되었습니다.")


#df_a = pd.read_excel(r"C:\\dev\\myproj01\\금상동\\결과폴더\\금상동1205_2024.xlsx")  # a파일 경로
df_a = pd.read_excel(f'C:\dev\myproj01\금상동\결과폴더\금상동봉안+월별결산부병합결과-{today}.xlsx')  # a파일 경로




#####################################################################################################
##### 월별결산부에는 없고 월별계정시트에 있는  사용료수입,관리선수금 등을 기타업로드에 추가하는 작업 ###

# Step 1: 엑셀 파일 불러오기
#file_path = 'path_to_your_excel_file.xlsx'  # 엑셀 파일 경로
#df = pd.read_excel(file_path)
df37 = pd.read_excel('C:\dev\myproj01\금상동\결과폴더\금상동세무사랑기타업로드수정.xlsx')
# Step 2: 조건 필터링 (계정코드2 == '사용료수입')
#df27 = df37[df37['계정코드2'] == '사용료수입']
df27 = df37[(df37['계정코드2'] == '사용료수입') & (df37['금액'] > 0)]
df28 = df37[df37['계정코드2'] == '관리선수금']
df29 = df37[df37['계정코드2'] == '사용선수금']
df30 = df37[df37['계정코드2'] == '회전관리선수금']


# Step 3:기타업로드에 있는  같은 봉안함 번호별로 금액 열의 합 계산
# '봉안함' 열을 기준으로 그룹화하여 '금액' 열 합계 계산
df_grouped = df27.groupby('봉안함', as_index=False).agg({'금액': 'sum'}) #사용료수입
df_grouped2 = df28.groupby('봉안함', as_index=False).agg({'금액': 'sum'}) #관리선수금
df_grouped3 = df29.groupby('봉안함', as_index=False).agg({'금액': 'sum'}) #사용선수금
df_grouped4 = df30.groupby('봉안함', as_index=False).agg({'금액': 'sum'}) #회전관리선수금

# Step 4: 새로운 데이터프레임으로 저장
df_result = df_grouped.copy() #사용료수입 세무사랑기타업로드수정 df37->df27->df_grouped
df_result2 = df_grouped2.copy() #관리선수금
df_result3 = df_grouped3.copy() #사용선수금
df_result4 = df_grouped4.copy() #회전관리선수금

# Step 5: 결과 출력
#print(df_result)

# Step 6: 결과를 엑셀로 저장 (필요한 경우)
output_path = r"C:\\dev\\myproj01\\금상동\\결과폴더\\금상동기타엑셀사용료수입합계.xlsx"  # 저장할 파일 경로
df_result.to_excel(output_path, index=False)

# Step 6: 결과를 엑셀로 저장 (필요한 경우)
output_path2 = r"C:\\dev\\myproj01\\금상동\\결과폴더\\금상동기타엑셀관리선수금합계.xlsx"  # 저장할 파일 경로
df_result2.to_excel(output_path2, index=False)

output_path3 = r"C:\\dev\\myproj01\\금상동\\결과폴더\\금상동기타엑셀사용선수금합계.xlsx"  # 저장할 파일 경로
df_result3.to_excel(output_path3, index=False)

output_path4 = r"C:\\dev\\myproj01\\금상동\\결과폴더\\금상동기타엑셀회전관리선수금합계.xlsx"  # 저장할 파일 경로
df_result4.to_excel(output_path4, index=False)





#d엑셀 기타업로드에 있고  계정별 각 시트에 있는 금액을 추가하기 위해  기타엑셀업로드의 봉안함리스트를 만듦

# 봉안함 리스트 추출
df_result_list = df_result['봉안함'].tolist() #사용료수입
df_result_list2 = df_result2['봉안함'].tolist() #관리선수금금
df_result_list3 = df_result3['봉안함'].tolist() #사용선수금금
df_result_list4 = df_result4['봉안함'].tolist() #회전관리선수금


# 사용료수입 기타엑셀업로드에 추가 조건에 따라 행 추가
for index, row in df_result.iterrows():
    if row['봉안함'] not in df_result_list and row['차변'] > 0:
        # 첫 번째 행 추가
        df37 = pd.concat([
            df37,
            pd.DataFrame({
                '월일': [row['날짜']],
                '구분1': [3],
                '코드1': ['273'],
                '계정코드2': ['사용선수금'],
                '코드4': [None],  # 필요하면 수정
                '거래처명': [None],
                '코드5': [None],  # 필요하면 수정
                '적요명': [row['성명']],  # 필요하면 수정
                '금액': [row['차변']],
                '봉안함': [row['봉안함']]
            })
        ], ignore_index=True)

        # 두 번째 행 추가
        df37 = pd.concat([
            df37,
            pd.DataFrame({
                '월일': [row['날짜']],
                '구분1': [4],
                '코드1': ['420'],
                '계정코드2': ['사용료수입'],
                '코드4': [None],  # 필요하면 수정
                '거래처명': [None],
                '코드5': [None],  # 필요하면 수정
                '적요명': [row['성명']],  # 필요하면 수정
                '금액': [row['차변']],
                '봉안함': [row['봉안함']]
            })
        ], ignore_index=True)





#  관리선수금 기타엑셀업로드에 추가 조건에 따라 행 추가
for index, row in total_df2.iterrows():
    if row['봉안함'] not in df_result_list2 and row['대변'] > 0:
        # 첫 번째 행 추가
        df37 = pd.concat([
            df37,
            pd.DataFrame({
                '월일': [row['날짜']],
                '구분1': [3],
                '코드1': ['108'],
                '계정코드2': ['외상매출금'],
                '코드4': [None],  # 필요하면 수정
                '거래처명': [None],
                '코드5': [None],  # 필요하면 수정
                '적요명': [row['성명']],  # 필요하면 수정
                '금액': [row['대변']],
                '봉안함': [row['봉안함']]
            })
        ], ignore_index=True)

        # 두 번째 행 추가
        df37 = pd.concat([
            df37,
            pd.DataFrame({
                '월일': [row['날짜']],
                '구분1': [4],
                '코드1': ['274'],
                '계정코드2': ['관리선수금'],
                '코드4': [None],  # 필요하면 수정
                '거래처명': [None],
                '코드5': [None],  # 필요하면 수정
                '적요명': [row['성명']],  # 필요하면 수정
                '금액': [row['대변']],
                '봉안함': [row['봉안함']]
            })
        ], ignore_index=True)


#  사용선수금 기타엑셀업로드에 추가 조건에 따라 행 추가
for index, row in total_df.iterrows():
    if row['봉안함'] not in df_result_list3 and row['대변'] > 0:
        # 첫 번째 행 추가
        df37 = pd.concat([
            df37,
            pd.DataFrame({
                '월일': [row['날짜']],
                '구분1': [3],
                '코드1': ['108'],
                '계정코드2': ['외상매출금'],
                '코드4': [None],  # 필요하면 수정
                '거래처명': [None],
                '코드5': [None],  # 필요하면 수정
                '적요명': [row['성명']],  # 필요하면 수정
                '금액': [row['대변']],
                '봉안함': [row['봉안함']]
            })
        ], ignore_index=True)

        # 두 번째 행 추가
        df37 = pd.concat([
            df37,
            pd.DataFrame({
                '월일': [row['날짜']],
                '구분1': [4],
                '코드1': ['273'],
                '계정코드2': ['사용선수금'],
                '코드4': [None],  # 필요하면 수정
                '거래처명': [None],
                '코드5': [None],  # 필요하면 수정
                '적요명': [row['성명']],  # 필요하면 수정
                '금액': [row['대변']],
                '봉안함': [row['봉안함']]
            })
        ], ignore_index=True)

#  회전관리선수금 기타엑셀업로드에 추가 조건에 따라 행 추가
for index, row in total_df4.iterrows():
    if row['봉안함'] not in df_result_list4 and row['대변'] > 0:
        # 첫 번째 행 추가
        df37 = pd.concat([
            df37,
            pd.DataFrame({
                '월일': [row['날짜']],
                '구분1': [3],
                '코드1': ['108'],
                '계정코드2': ['외상매출금'],
                '코드4': [None],  # 필요하면 수정
                '거래처명': [None],
                '코드5': [None],  # 필요하면 수정
                '적요명': [row['성명']],  # 필요하면 수정
                '금액': [row['대변']],
                '봉안함': [row['봉안함']]
            })
        ], ignore_index=True)

        # 두 번째 행 추가
        df37 = pd.concat([
            df37,
            pd.DataFrame({
                '월일': [row['날짜']],
                '구분1': [4],
                '코드1': ['275'],
                '계정코드2': ['회전관리선수금'],
                '코드4': [None],  # 필요하면 수정
                '거래처명': [None],
                '코드5': [None],  # 필요하면 수정
                '적요명': [row['성명']],  # 필요하면 수정
                '금액': [row['대변']],
                '봉안함': [row['봉안함']]
            })
        ], ignore_index=True)

















# 결과 출력

# Step 5: 결과 출력

# Step 6: 엑셀로 저장 (필요시)
# 1. 비정상적인 값 확인
#print(df37['월일'].unique())

# 2. 결측값(NaN) 또는 비정상적인 값 제외
df37['월일'] = df37['월일'].replace("", None)  # 빈 문자열 제거
df37['월일'] = pd.to_datetime(df37['월일'], errors='coerce')  # 변환 불가능한 값은 NaT로 대체

# 3. NaT(변환 불가능한 값) 제거
df37 = df37.dropna(subset=['월일'])
# df37['월일'] 열이 문자열 형식일 경우 날짜 형식으로 변환
df37['월일'] = pd.to_datetime(df37['월일'])

# 날짜 형식을 'YYYY-MM-DD'로 변환하여 저장
df37['월일'] = df37['월일'].dt.strftime('%Y-%m-%d')



df37.rename(columns={'적요명' : '거래처명'}, inplace=False)

# 관리비수입 기타엑셀업로드에 추가 조건에 따라 행 추가
for _, row in df_a.iterrows():
    if  row['연중관리수입2'] > 0:
        # 첫 번째 행 추가
        df37 = pd.concat([
            df37,
            pd.DataFrame({
                '월일': end_of_input_year,
                '구분1': [3],
                '코드1': ['274'],
                '계정코드2': ['관리선수금'],
                '코드4': [None],  # 필요하면 수정
                '거래처명': [None],
                '코드5': [None],  # 필요하면 수정
                '적요명': [row['계약자명']],  # 필요하면 수정
                '금액': [row['연중관리수입2']],
                '봉안함': [row['봉안함']]
            })
        ], ignore_index=True)

        # 두 번째 행 추가
        df37 = pd.concat([
            df37,
            pd.DataFrame({
                '월일': end_of_input_year,
                '구분1': [4],
                '코드1': ['421'],
                '계정코드2': ['관리비수입'],
                '코드4': [None],  # 필요하면 수정
                '거래처명': [None],
                '코드5': [None],  # 필요하면 수정
                '적요명': [row['계약자명']],  # 필요하면 수정
                '금액': [row['연중관리수입2']],
                '봉안함': [row['봉안함']]
            })
        ], ignore_index=True)




# 연중회전관리비수입 기타엑셀업로드에 추가 조건에 따라 행 추가
for _, row in df_a.iterrows():
    if  row['연중회전관리수입'] > 0:
        # 첫 번째 행 추가
        df37 = pd.concat([
            df37,
            pd.DataFrame({
                '월일': end_of_input_year,
                '구분1': [3],
                '코드1': ['275'],
                '계정코드2': ['회전관리선수금'],
                '코드4': [None],  # 필요하면 수정
                '거래처명': [None],
                '코드5': [None],  # 필요하면 수정
                '적요명': [row['계약자명']],  # 필요하면 수정
                '금액': [row['연중회전관리수입']],
                '봉안함': [row['봉안함']]
            })
        ], ignore_index=True)

        # 두 번째 행 추가
        df37 = pd.concat([
            df37,
            pd.DataFrame({
                '월일': end_of_input_year,
                '구분1': [4],
                '코드1': ['425'],
                '계정코드2': ['회전관리비수입'],
                '코드4': [None],  # 필요하면 수정
                '거래처명': [None],
                '코드5': [None],  # 필요하면 수정
                '적요명': [row['계약자명']],  # 필요하면 수정
                '금액': [row['연중회전관리수입']],
                '봉안함': [row['봉안함']]
            })
        ], ignore_index=True)




df37['적요명'] = df37['적요명'].astype(str) + "**" + df37['봉안함'].astype(str)
# 2. 결측값(NaN) 또는 비정상적인 값 제외
df37['월일'] = df37['월일'].replace("", None)  # 빈 문자열 제거
df37['월일'] = pd.to_datetime(df37['월일'], errors='coerce')  # 변환 불가능한 값은 NaT로 대체

df37['월일'] = df37['월일'].dt.strftime('%Y-%m-%d')
df37.to_excel('C:\dev\myproj01\금상동\결과폴더\금상동세무사랑기타업로드+봉안사용료수입.xlsx', index=False)

# 엑셀 파일 열기
file_path37 = f'C:\dev\myproj01\금상동\결과폴더\금상동세무사랑기타업로드+봉안사용료수입.xlsx'
wb37= openpyxl.load_workbook(file_path37)
ws37 = wb37.active  # 활성화된 시트를 선택

 # A2 셀 기준으로 틀 고정 (freeze panes)
ws37.freeze_panes = 'A2'
 # 첫 번째 행에 필터 추가
ws37.auto_filter.ref = ws37.dimensions  # ws.dimensions는 사용 중인 셀의 범위를 자동으로 가져옴
# 모든 시트에 대해 A열부터 D열까지의 너비를 25로 설정
columns_to_resize = ['h','j']  # A열부터 D열까지 지정
columns_to_resize2 = ['a','d','i']  # A열부터 D열까지 지정

for sheet in wb37.worksheets:
    for column_letter in columns_to_resize:
        # 지정된 열(column_letter)의 너비를 25로 설정
        sheet.column_dimensions[column_letter].width = 25
    
    for column_letter in columns_to_resize2:
        # 지정된 열(column_letter)의 너비를 25로 설정
        sheet.column_dimensions[column_letter].width = 15


       # 통화 기호 없는 회계 스타일 정의
        accounting_style = NamedStyle(name="accounting_style", number_format='#,##0;[Red](#,##0)')

# 모든 시트에 대해 E열부터 G열까지 회계 스타일 적용
    for sheet in wb37.worksheets:
      for row in range(1, sheet.max_row + 1):
        for col in ['I']:  # E열부터 G열까지
            cell = sheet[f'{col}{row}']
            cell.style = accounting_style
save_path2 = f'C:\dev\myproj01\금상동\결과폴더\금상동세무사랑기타업로드+봉안사용료수입-{today}.xlsx' 
 # A2 셀 기준으로 틀 고정 (freeze panes)
ws37.freeze_panes = 'A2'

wb37.save(save_path2)
print(f"새로운 데이터가 {save_path2} 파일로 저장되었습니다.")    





#####################################################################################
####기타엑셀업로드에는 없는 사용료수입을 df_a의 사용료 수입을 인식하기 위한단계계


import pandas as pd
import numpy as np

# 데이터프레임 로드
df_a = pd.read_excel(f'C:\\dev\\myproj01\\금상동\\결과폴더\\금상동봉안+월별결산부병합결과-{today}.xlsx')
df47 = pd.read_excel(f'C:\\dev\\myproj01\\금상동\\결과폴더\\금상동세무사랑기타업로드+봉안사용료수입-{today}.xlsx')

# 2024-사용료수입이 0보다 큰 행 필터링
df_a_filtered = df_a[df_a['2024-사용료수입'] > 0].copy()

# 봉안일 선택 (봉안일1, 봉안일2 중 input_year 포함)
df_a_filtered['봉안일1'] = df_a_filtered['봉안일1'].fillna('')
df_a_filtered['봉안일2'] = df_a_filtered['봉안일2'].fillna('')

df_a_filtered['selected_date'] = np.where(
    df_a_filtered['봉안일1'].str.contains(str(input_year)),
    df_a_filtered['봉안일1'],
    np.where(
        df_a_filtered['봉안일2'].str.contains(str(input_year)),
        df_a_filtered['봉안일2'],
        '2024-12-31'  # 기본값
    )
)

# 결과 저장
df_a_filtered.to_excel('C:\\dev\\myproj01\\금상동\\결과폴더\\df_a_filtered.xlsx', index=False)

# 추가 데이터 처리
#df57 = df47[df47['계정코드2'] == '사용료수입'] & ([df47['금액'] > 0)]
df57 = df47[(df47['계정코드2'] == '사용료수입') & (df37['금액'] > 0)]

df_a57 = df_a[df_a['2024-사용료수입'].notna()]

# '봉안함' 기준으로 그룹화
df_grouped57 = df57.groupby('봉안함', as_index=False).agg({'금액': 'sum'})
df_a_grouped57 = df_a57.groupby('봉안함', as_index=False).agg({'2024-사용료수입': 'sum'})

# 데이터프레임 복사
df_result57 = df_grouped57.copy()
df_a_result57 = df_a_grouped57.copy()

# 저장
df_grouped57.to_excel('C:\\dev\\myproj01\\금상동\\결과폴더\\df_result57.xlsx', index=False)
df_a_grouped57.to_excel('C:\\dev\\myproj01\\금상동\\결과폴더\\df_a_result57.xlsx', index=False)

# 병합
merged_df = pd.merge(
    df_a_result57, df_result57,
    on='봉안함', how='outer', suffixes=('_a', '_r')
)
merged_df.to_excel('C:\\dev\\myproj01\\금상동\\결과폴더\\사용료수입&봉안사용료수입머지.xlsx', index=False)
# 금액 차이 계산
#merged_df['금액_차이'] = max(0,(merged_df['2024-사용료수입'].fillna(0) - merged_df['금액'].fillna(0)))
# 금액 차이를 계산할 때 max 함수를 벡터화된 연산으로 변경
merged_df['금액_차이'] = np.maximum(0, merged_df['2024-사용료수입'].fillna(0) - merged_df['금액'].fillna(0))

# 금액 차이가 있는 행 필터링
diff_rows = merged_df[merged_df['금액_차이'] != 0]

# 수정된 행 추가
for _, row in diff_rows.iterrows():
    # 봉안일 데이터 가져오기
    matching_rows = df_a_filtered[df_a_filtered['봉안함'] == row['봉안함']]
    if not matching_rows.empty:
        selected_date = matching_rows['selected_date'].iloc[0]
    else:
        selected_date = '2024-12-31'  # 기본값

    # 사용선수금 행 추가
    new_row1 = pd.DataFrame({
        '월일': [selected_date],  # 동적으로 설정된 날짜
        '구분1': [3],
        '코드1': ['273'],
        '계정코드2': ['사용선수금'],
        '코드4': [None],
        '거래처명':[None],
        '코드5': [None],
        '적요명':[None],
        '금액': [row['금액_차이']],
        '봉안함': [row['봉안함']]
    })

    # 사용료수입 행 추가
    new_row2 = pd.DataFrame({
        '월일': [selected_date],  # 동적으로 설정된 날짜
        '구분1': [4],
        '코드1': ['420'],
        '계정코드2': ['사용료수입'],
        '코드4': [None],
        '거래처명': [None],
        '코드5': [None],
        '적요명': [None],
        '금액': [row['금액_차이']],
        '봉안함': [row['봉안함']]
    })

    # 새로운 행 추가
    df47 = pd.concat([df47, new_row1, new_row2], ignore_index=True)
# '봉안함'을 기준으로 두 데이터프레임 병합
merged_df = pd.merge(
    df47, 
    df_a[['봉안함', '계약자명']], 
    on='봉안함', 
    how='left', 
    suffixes=('', '_from_a')
)

# '적요명'이 NaN인 경우 '계약자명'과 '봉안함'을 결합하여 채움
merged_df['적요명'] = merged_df['적요명'].fillna(
    merged_df['계약자명'] + ' & ' + merged_df['봉안함'].astype(str)
)

# 결과를 df47에 다시 저장
df47 = merged_df.drop(columns=['계약자명'])


# 최종 결과 저장
df47.to_excel('C:\\dev\\myproj01\\금상동\\결과폴더\\금상동세무사랑기타업로드+봉안사용료수입추가1225-(9월에합산).xlsx', index=False)

# 차변,대변 출력후 한칸씩 띄우는 작업  대차차액 발생을 방지하기위함함
# 엑셀 워크북 생성
# 새로운 데이터프레임을 생성할 리스트
new_rows = []

# 이전 행의 값 초기화
previous_value = None

# 데이터프레임을 순회하며 조건에 맞는 행 추가
for index, row in df47.iterrows():
    # 현재 행 추가
    new_rows.append(row)
    
    # 현재 행의 '구분1' 값
    current_value = row['구분1']
    
    # 조건에 따라 빈 행 추가
    if (previous_value == 3 and current_value == 4) or (current_value == 2):
        new_rows.append(pd.Series([None] * len(row), index=row.index))  # 빈 행 추가
    
    # 이전 값 업데이트
    previous_value = current_value

# 새로운 데이터프레임 생성
updated_df47 = pd.DataFrame(new_rows)

# 엑셀 파일로 저장
updated_df47.to_excel('C:\\dev\\myproj01\\금상동\\결과폴더\\updated_df47.xlsx', index=False)
# 결과 저장
