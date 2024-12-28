import pandas as pd			
from pathlib import Path
import openpyxl	
import pandas as pd

from openpyxl import load_workbook

import pandas as pd			
from pathlib import Path
import openpyxl	
from datetime import datetime
from dateutil.relativedelta import relativedelta
import pandas as pd			
from pathlib import Path
import openpyxl	
from openpyxl.styles import NamedStyle
from openpyxl import Workbook
from openpyxl.styles import PatternFill #엑셀셀을 빨간색으로 표시
import socket
from openpyxl.utils.dataframe import dataframe_to_rows
from pathlib import Path
import openpyxl	
from openpyxl.styles import NamedStyle

from openpyxl.styles import PatternFill #엑셀셀을 빨간색으로 표시
import socket
import re


def get_local_ip():
    """현재 컴퓨터의 IP 주소를 반환합니다."""
    hostname = socket.gethostname()  # 컴퓨터의 호스트 이름 가져오기
    local_ip = socket.gethostbyname(hostname)  # 호스트 이름 기반 IP 가져오기
    return local_ip

# IP 주소에 따른 경로 설정
local_ip = get_local_ip()

if local_ip == "192.168.0.29":  # 사무실에서 사용하는 IP
    file_path30 = r'C:\Users\erato2022\OneDrive\문서\2.금상동천호성지\2024기초데이터\통합조회_20241203금상동.xlsx'  # 엑셀 파일 경로를 지정하세요
elif local_ip != "192.168.0.29":  # 사무실이 아닌경우 IP
    file_path30 =r'C:\Users\erato\OneDrive\문서\2.금상동천호성지\2024기초데이터\통합조회_20241203금상동.xlsx'
else:  # 기본 경로 (기타 환경)
    file_path30 = "/mnt/default_path/"

#print(f"Current Path: {file_path30}")






#########################################################################################
#####    봉안프로그램 

# 엑셀 파일 불러오기
#file_path30 = r'C:\dev\myproj01\통합조회_20241123금상동.xlsx'  # 엑셀 파일 경로를 지정하세요
#file_path30 = r'C:\Users\erato2022\OneDrive\문서\2.금상동천호성지\2024기초데이터\통합조회_20241128금상동.xlsx'  # 엑셀 파일 경로를 지정하세요
#file_path30 =r'C:\Users\erato\OneDrive\문서\2.금상동천호성지\2024기초데이터\통합조회_20241128금상동.xlsx'
             

if local_ip == "192.168.0.29":  # 사무실에서 사용하는 IP
    #file_path30 = r'C:\Users\erato2022\OneDrive\문서\2.금상동천호성지\2024기초데이터\통합조회_20241128금상동.xlsx'  # 엑셀 파일 경로를 지정하세요
    output_file = r'C:\Users\erato2022\OneDrive\문서\2.금상동천호성지\2024기초데이터\통합조회_20241128금상동(관리선수금월수).xlsx'
elif local_ip != "192.168.0.29":  # 사무실이 아닌경우 IP
    #file_path30 =r'C:\Users\erato\OneDrive\문서\2.금상동천호성지\2024기초데이터\통합조회_20241128금상동.xlsx'
     output_file = r'C:\Users\erato\OneDrive\문서\2.금상동천호성지\2024기초데이터\통합조회_20241128금상동(관리선수금월수).xlsx'
else:  # 기본 경로 (기타 환경)
    file_path30 = "/mnt/default_path/"

df30 = pd.read_excel(file_path30)
df30.info()

#output_file = r'C:\Users\erato2022\OneDrive\문서\2.금상동천호성지\2024기초데이터\통합조회_20241128금상동(관리선수금월수).xlsx'

# 날짜가 유효하지 않다면 2월 28일로 수정하는 함수
def fix_invalid_date(date_str):
    try:
        # 날짜가 유효한지 확인
        date_obj = pd.to_datetime(date_str)
    except ValueError:
        # 유효하지 않은 날짜인 경우 처리 (윤년이 아닌 해의 2월 29일 등)
        if "02-29" in date_str:
            # 날짜가 2월 29일인 경우 2월 28일로 수정
            date_str = date_str.replace("02-29", "02-28")
            date_obj = pd.to_datetime(date_str)
        else:
            date_obj = pd.NaT
    return date_obj

# '관리시작열'과 '관리종료열'의 날짜를 변환하며 오류를 수정
df30['관리시작'] = df30['관리시작'].apply(fix_invalid_date)
df30['관리종료'] = df30['관리종료'].apply(fix_invalid_date)
df30['봉안일1'] = df30['봉안일1'].apply(fix_invalid_date)
df30['봉안일2'] = df30['봉안일2'].apply(fix_invalid_date)
df30['만기일1'] = df30['만기일1'].apply(fix_invalid_date)
df30['만기일2'] = df30['만기일2'].apply(fix_invalid_date)



# 연도를 입력받음
input_year = int(input("연도를 입력하세요 (예: 2023): "))
start_of_input_year = datetime(input_year, 1, 1)
end_of_input_year = datetime(input_year, 12, 31)



# 날짜 형식 변환
df30['봉안일1'] = pd.to_datetime(df30['봉안일1'],errors='coerce')
df30['봉안일2'] = pd.to_datetime(df30['봉안일2'],errors='coerce')
df30['만기일1'] = pd.to_datetime(df30['만기일1'],errors='coerce')
df30['만기일2'] = pd.to_datetime(df30['만기일2'],errors='coerce')

df30['관리시작'] = pd.to_datetime(df30['관리시작'],errors='coerce')
df30['관리종료'] = pd.to_datetime(df30['관리종료'],errors='coerce')

# 관리선수금총월수를 계산하는 함수
def calculate_total_months_with_extra_month(row):
    #def calculate_total_months(row):
    try:
        # 날짜 값이 모두 유효한 경우에만 계산
        if pd.notna(row['관리종료']) and pd.notna(row['관리시작']):
    
          if pd.notna(row['봉안일1']) or pd.notna(row['봉안일2']):
             #조건 A
             #if row['봉안일1'] == row['봉안일2'] == row['관리시작'] or row['봉안일1'] == row['봉안일2']:
             # 관리종료 - 관리시작의 월수 계산
              delta = relativedelta(row['관리종료'], row['관리시작'])
              total_months =delta.years * 12 + delta.months+1
              return total_months               
             
              
         
    except Exception as e:
        print(f"Error processing row: {row}, Error: {e}")
        return 0       
    
      # 조건을 만족하지 않는 경우 0 반환



# 회전관리선수금총월수를 계산하는 함수
def calculate_total_months_with_extra_month2(row):
    #def calculate_total_months(row):
    try:
        # 날짜 값이 모두 유효한 경우에만 계산
        if pd.notna(row['만기일1']) or  pd.notna(row['만기일2']):
    
          if pd.notna(row['봉안일1']) :
             #조건 A
             #if row['봉안일1'] == row['봉안일2'] == row['관리시작'] or row['봉안일1'] == row['봉안일2']:
             # 관리종료 - 관리시작의 월수 계산
              delta = relativedelta(row['만기일1'],row['관리종료'], )
              total_months =delta.years * 12 + delta.months
              return total_months               
             
          if pd.notna(row['봉안일2']) :
             #조건 A
             #if row['봉안일1'] == row['봉안일2'] == row['관리시작'] or row['봉안일1'] == row['봉안일2']:
             # 관리종료 - 관리시작의 월수 계산
              delta = relativedelta(row['만기일2'],row['관리종료'], )
              total_months =delta.years * 12 + delta.months
              return total_months     
         
    except Exception as e:
        print(f"Error processing row: {row}, Error: {e}")
        return 0       
    
      # 조건을 만족하지 않는 경우 0 반환


# 관리선수금 년중월수 계산 함수
def calculate_year_months(row):
    start_date = pd.to_datetime(row['관리시작'])
    end_date = pd.to_datetime(row['관리종료'])

    # 관리 시작 날짜가 입력된 연도보다 이전인 경우 처리
    if start_date < start_of_input_year:
        # 입력된 연도의 시작일부터 계산
        actual_start = start_of_input_year
    else:
        actual_start = start_date

    # 관리 종료 날짜와 입력된 연도의 종료일 중 빠른 날짜 선택
    actual_end = min(end_date, end_of_input_year)

    # 기간이 겹치지 않는 경우 0 반환
    if actual_end < actual_start:
        return 0

    # 월 수 계산
    months_in_year = (actual_end.year - actual_start.year) * 12 + actual_end.month - actual_start.month
    return months_in_year+1  # 시작과 종료 포함


# 연중회전관리선수금월수를 계산하는 함수
def calculate_total_months_with_extra_month3(row):
    #def calculate_total_months(row):
    try:
        # 날짜 값이 모두 유효한 경우에만 계산
        if pd.notna(row['만기일1']) or  pd.notna(row['만기일2']):
    
          if pd.notna(row['봉안일1']) :
            start_date = pd.to_datetime(row['관리종료'])
            end_date = pd.to_datetime(row['만기일1'])
          
            # 관리 시작 날짜가 입력된 연도보다 이전인 경우 처리
            if start_date < start_of_input_year:
             # 입력된 연도의 시작일부터 계산
               actual_start = start_of_input_year
            else:
               actual_start = start_date

               # 관리 종료 날짜와 입력된 연도의 종료일 중 빠른 날짜 선택
            actual_end = min(end_date, end_of_input_year)

            # 기간이 겹치지 않는 경우 0 반환
            if actual_end < actual_start:
              return 0
            months_in_year = (actual_end.year - actual_start.year) * 12 + actual_end.month - actual_start.month
            return months_in_year+1
            # 시작과 종료 포함





          if pd.notna(row['봉안일2']) :
            start_date = pd.to_datetime(row['관리종료'])
            end_date = pd.to_datetime(row['만기일2'])
          
            # 관리 시작 날짜가 입력된 연도보다 이전인 경우 처리
            if start_date < start_of_input_year:
             # 입력된 연도의 시작일부터 계산
               actual_start = start_of_input_year
            else:
               actual_start = start_date

               # 관리 종료 날짜와 입력된 연도의 종료일 중 빠른 날짜 선택
            actual_end = min(end_date, end_of_input_year)

            # 기간이 겹치지 않는 경우 0 반환
            if actual_end < actual_start:
              return 0# 월 수 계산
            months_in_year = (actual_end.year - actual_start.year) * 12 + actual_end.month - actual_start.month
            return months_in_year+1  # 시작과 종료 포함 #조건 A
            
    except Exception as e:
        print(f"Error processing row: {row}, Error: {e}")
        return 0       















# 연말잔여월수 계산 함수
# 입력된 연도의 종료일을 계산
end_of_input_year = datetime(input_year, 12, 31)

# 잔여월수 계산 함수
def calculate_remaining_months(row):
    end_date = pd.to_datetime(row['관리종료'])
    
    # 관리종료열이 입력된 연도의 종료일보다 빠른 경우 0 반환
    if end_date <= end_of_input_year:
        return 0
    
    # 관리종료열이 입력된 연도의 종료일보다 늦은 경우 잔여 월수 계산
    else:
        remaining_months = (end_date.year - end_of_input_year.year) * 12 + (end_date.month - end_of_input_year.month)
        return remaining_months + 1  # 종료 월을 포함한 월수 반환




# 3. 관리시작열에서 입력된 연도를 포함하는지 확인 후, 사용료수입 열에 값을 추가하거나 0으로 설정
def check_year(row):
    # 관리시작열에서 입력된 연도를 포함하면 사용료수입 값을 유지, 포함하지 않으면 0을 반환
    if str(input_year) in str(row['관리시작']):
        return row['사용수입']
    else:
        return 0

def check_year2(row):
    start_date = pd.to_datetime(row['관리시작'])
    end_date = pd.to_datetime(row['관리종료'])
    
    if start_date < start_of_input_year:
        # 입력된 연도의 시작일부터 계산
        return row['사용수입']
    else:
        return 0



# 각 행에 대해 함수 적용
df30['관리선수금전체월수'] = df30.apply(calculate_total_months_with_extra_month, axis=1)

# 년중월수열 추가
df30['년중월수'] = df30.apply(calculate_year_months, axis=1)

# 잔여월수열 추가
df30['잔여월수'] = df30.apply(calculate_remaining_months, axis=1)
#df30['연중관리수입']=df30['관리수입']/df30['관리선수금전체월수']*df30['년중월수']
df30['잔여관리선수금']=df30['관리수입']/df30['관리선수금전체월수']*df30['잔여월수']
# '사용료수입' 열을 업데이트
df30['사용료수입'] = df30.apply(check_year, axis=1)
df30['사용료경과분'] = df30.apply(check_year2, axis=1)
# 사용자로부터 연도 입력받기
input_year = input("연도를 입력하세요 (예: 2023): ")

# 계약일에 입력한 연도 포함 여부를 확인하는 함수
def contains_year(date_series, year):
    # 계약일이 NaN이 아닌 경우에만 문자열 변환 후 year 포함 여부 확인
    return date_series.dropna().astype(str).str.contains(year)

# 조건: 입력받은 연도가 2023인 경우
if input_year == input_year:
    # 새로운 열 이름을 생성: '2023-사용선수금'
    new_column_name = f"{input_year}-사용선수금"
    new_column_name2 = f"{input_year}-관리선수금"
    new_column_name3 = f"{input_year}-사용료수입"
    
    # 새로운 열을 데이터프레임에 추가하고 초기값은 None으로 설정
    df30[new_column_name] = None
    df30[new_column_name2] = None
    df30[new_column_name3] = None
    
    # '계약일' 열에 입력받은 연도를 포함하고, 나머지 열이 모두 결측치일 경우 마스크 정의
    mask = (contains_year(df30['계약일'], input_year)) & \
           (df30['봉안일1'].isna()) & \
           (df30['봉안일2'].isna()) & \
           (df30['관리시작'].isna()) & \
           (df30['관리종료'].isna())

    # '봉안일1' 열에 입력받은 연도를 포함하고, 나머지 열이 모두 결측치일 경우 마스크 정의
    mask2 = (contains_year(df30['관리시작'], input_year)) & \
            (df30['사용수입'] > 0)


    # '사용선수' 값이 있는 행에서 '2023-사용선수금' 열에 값을 넣음
    df30.loc[mask, new_column_name] = df30.loc[mask, '사용선수']
    df30.loc[mask, new_column_name2] = df30.loc[mask, '관리선수']
    df30.loc[mask2, new_column_name3] = df30.loc[mask2, '사용수입']

    # 결과 확인
    print(df30)
    df30=df30.sort_values(["봉안함"]) # 일자와 환율을 일자와 환율로 정렬
    # 새로운 파일로 저장
    output_file_path = f"C:\\dev\\myproj01\\금상동\\결과폴더\\천호성지1205_{input_year}.xlsx"
    
    df30.to_excel(output_file_path, index=False)
    print(f"새로운 데이터가 {output_file_path} 파일로 저장되었습니다.")









##########################################################################
#        월별결산부통합  ##################################################


# 여러 개의 데이터프레임을 엑셀 파일로 내보내기 함수
def export_multiple_dfs_to_excel(dfs, sheet_names, file_name):
    """
    여러 개의 데이터프레임을 각각 다른 시트에 엑셀 파일로 내보내는 함수.
    : param dfs: 데이터프레임 리스트 
    : param sheet_names: 시트 이름 리스트 
    : param file_name: 저장할 엑셀 파일 이름
   """ 
    
    if len(dfs) != len(sheet_names) :

      raise ValueError("데이터프레임 리스트와 시트 이름 리스트의 길이가 같아야 합니다. ")


    with pd. ExcelWriter (file_name) as writer:
     for df, sheet_name in zip(dfs, sheet_names) :
        df. to_excel(writer, sheet_name=sheet_name,index=False)
    #Print(f"{file_name} 파일이 생성되었습니다.")




#"C:\Users\erato\OneDrive\문서\2.금상동천호성지\2024기초데이터\금상동월별결산부"
local_ip = get_local_ip()

if local_ip == "192.168.0.30":  # 사무실에서 사용하는 IP
    #file_path30 = r'C:\Users\erato2022\OneDrive\문서\2.금상동천호성지\2024기초데이터\통합조회_20241203천호성지.xlsx'  # 엑셀 파일 경로를 지정하세요
    #input_folder=r"C:\Users\erato2022\OneDrive\문서\2.금상동천호성지\2024기초데이터\사용료현금출납부-2024"
    file_path31 = r"C:\Users\erato2022\OneDrive\문서\2.금상동천호성지\2024기초데이터\사용료 현금출납부-2024.xlsx" 
    file_path32 = r"C:\Users\erato2022\OneDrive\문서\2.금상동천호성지\2024기초데이터\봉안경당매일시재(계약선수금)-2024.xlsx" 
    file_path33 = r"C:\Users\erato2022\OneDrive\문서\2.금상동천호성지\2024기초데이터\봉안경당 매일시재(봉안)-2024.xlsx" 
    
    
elif local_ip != "192.168.0.30":  # 사무실이 아닌경우 IP
    #input_folder=r"C:\Users\erato\OneDrive\문서\2.금상동천호성지\2024기초데이터\사용료현금출납부-2024"
    #file_path30 =r'C:\Users\erato\OneDrive\문서\2.금상동천호성지\2024기초데이터\통합조회_20241203금상동.xlsx'
     file_path31 = r"C:\Users\erato\OneDrive\문서\2.금상동천호성지\2024기초데이터\사용료 현금출납부-2024.xlsx" 
     file_path32 = r"C:\Users\erato\OneDrive\문서\2.금상동천호성지\2024기초데이터\봉안경당매일시재(계약선수금)-2024.xlsx" 
     file_path33 = r"C:\Users\erato\OneDrive\문서\2.금상동천호성지\2024기초데이터\봉안경당 매일시재(봉안)-2024.xlsx" 
    #"C:\Users\erato2022\OneDrive\문서\2.금상동천호성지\2024기초데이터\사용료 현금출납부-2024.xlsx"
else:  # 기본 경로 (기타 환경)
    
    file_path31 = "/mnt/default_path/"
    file_path32 = "/mnt/default_path/"
    file_path33 = "/mnt/default_path/"

#print(f"Current Path: {file_path30}")

# 엑셀 파일 불러오기
wb31 = openpyxl.load_workbook(file_path31)
wb32 = openpyxl.load_workbook(file_path32)
wb33 = openpyxl.load_workbook(file_path33)

ws31 = wb31.active
ws32 = wb32.active
ws33 = wb33.active


# 병합된 셀 정보 확인
print("병합된 셀 범위:", ws31.merged_cells.ranges)

# 병합된 셀의 범위를 리스트로 복사
merged_ranges = list(ws31.merged_cells.ranges)
merged_ranges2 = list(ws32.merged_cells.ranges)
merged_ranges3 = list(ws33.merged_cells.ranges)

# 병합 해제
for merge_range in merged_ranges:
    ws31.unmerge_cells(str(merge_range))

for merge_range in merged_ranges2:
    ws32.unmerge_cells(str(merge_range))
    
for merge_range in merged_ranges3:
    ws33.unmerge_cells(str(merge_range))   



if local_ip == "192.168.0.30":  # 사무실에서 사용하는 IP


   # 수정된 파일 저장
   wb31.save(r"c:\Users\erato2022\OneDrive\문서\2.금상동천호성지\2024기초데이터\사용료 현금출납부병합-2024.xlsx")
   wb32.save(r"c:\Users\erato2022\OneDrive\문서\2.금상동천호성지\2024기초데이터\봉안경당매일시재(계약선수금)병합-2024.xlsx")
   wb33.save(r"c:\Users\erato2022\OneDrive\문서\2.금상동천호성지\2024기초데이터\봉안경당매일시재(봉안)병합-2024.xlsx")

elif local_ip != "192.168.0.30":  # 사무실이 아닌경우 IP

     # 수정된 파일 저장
   wb31.save(r"c:\Users\erato\OneDrive\문서\2.금상동천호성지\2024기초데이터\사용료 현금출납부병합-2024.xlsx")
   wb32.save(r"c:\Users\erato\OneDrive\문서\2.금상동천호성지\2024기초데이터\봉안경당매일시재(계약선수금)병합-2024.xlsx")
   wb33.save(r"c:\Users\erato\OneDrive\문서\2.금상동천호성지\2024기초데이터\봉안경당매일시재(봉안)병합-2024.xlsx")



print("병합된 셀이 모두 해제되었습니다.")


if local_ip == "192.168.0.30":  # 사무실에서 사용하는 IP
    
    file_path31 = r"C:\Users\erato2022\OneDrive\문서\2.금상동천호성지\2024기초데이터\사용료 현금출납부-2024.xlsx" 
    file_path32 = r"C:\Users\erato2022\OneDrive\문서\2.금상동천호성지\2024기초데이터\봉안경당매일시재(계약선수금)-2024.xlsx" 
    file_path33 = r"C:\Users\erato2022\OneDrive\문서\2.금상동천호성지\2024기초데이터\봉안경당 매일시재(봉안)-2024.xlsx" 
    file_path34 = r'C:\Users\erato2022\OneDrive\문서\2.금상동천호성지\2024기초데이터\사용료현금출납부셀병합해제-2024.xlsx' 
    
    
elif local_ip != "192.168.0.30":  # 사무실이 아닌경우 IP
   
     file_path31 = r"C:\Users\erato\OneDrive\문서\2.금상동천호성지\2024기초데이터\사용료 현금출납부-2024.xlsx" 
     file_path32 = r"C:\Users\erato\OneDrive\문서\2.금상동천호성지\2024기초데이터\봉안경당매일시재(계약선수금)-2024.xlsx" 
     file_path33 = r"C:\Users\erato\OneDrive\문서\2.금상동천호성지\2024기초데이터\봉안경당 매일시재(봉안)-2024.xlsx" 
     file_path34 = r'C:\Users\erato\OneDrive\문서\2.금상동천호성지\2024기초데이터\사용료현금출납부셀병합해제-2024.xlsx' 
else:  # 기본 경로 (기타 환경)
    
    file_path31 = "/mnt/default_path/"
    file_path32 = "/mnt/default_path/"
    file_path33 = "/mnt/default_path/"


if local_ip == "192.168.0.30":  # 사무실에서 사용하는 IP

 df31 = pd.read_excel(r'C:\Users\erato2022\OneDrive\문서\2.금상동천호성지\2024기초데이터\사용료현금출납부셀병합해제-2024.xlsx',sheet_name="2024년도",skiprows=3,usecols="a:g")
 df32 = pd.read_excel(r'C:\Users\erato2022\OneDrive\문서\2.금상동천호성지\2024기초데이터\봉안경당매일시재(계약선수금)셀병합해제-2024.xlsx',sheet_name="2024년도",skiprows=4,usecols="a:g")
 df33 = pd.read_excel(r'C:\Users\erato2022\OneDrive\문서\2.금상동천호성지\2024기초데이터\봉안경당매일시재(봉안)셀병합해제-2024.xlsx',sheet_name="2024년도",skiprows=4,usecols="a:g")

elif local_ip != "192.168.0.30":  # 사무실이 아닌경우 IP
  
  df31 = pd.read_excel(r'C:\Users\erato\OneDrive\문서\2.금상동천호성지\2024기초데이터\사용료현금출납부셀병합해제-2024.xlsx',sheet_name="2024년도",skiprows=3,usecols="a:g")
  df32 = pd.read_excel(r'C:\Users\erato\OneDrive\문서\2.금상동천호성지\2024기초데이터\봉안경당매일시재(계약선수금)셀병합해제-2024.xlsx',sheet_name="2024년도",skiprows=4,usecols="a:g")
  df33 = pd.read_excel(r'C:\Users\erato\OneDrive\문서\2.금상동천호성지\2024기초데이터\봉안경당매일시재(봉안)셀병합해제-2024.xlsx',sheet_name="2024년도",skiprows=4,usecols="a:g")




df31.columns = ['월일','계정','적요','수입','지출','잔액','비고']
df32.columns = ['월일','봉안함','사용료계약금','사용료','소계','누계','비고']
df33.columns = ['월일','봉안함','사용료','관리비','소계','누계','비고']
#df31.info()

df31= df31.dropna(subset=["계정"]) #아래 누적액,월계액을 없애기위해 회계계정없는부분을 삭제함

df31['적요'] = df31['적요'].fillna("").astype(str)  # NaN 값을 빈 문자열로 대체하고 문자열로
    

# 괄호 안의 값을 추출하는 함수
def extract_bracket_value(text):
    if isinstance(text, str):
        match = re.search(r'\((.*?)\)', text)
        if match:
            return match.group(1)
    return None

# '적요' 열 처리 함수
def process_column(df):
    df['적요'] = df['적요'].fillna('값(0)')  # 결측값을 '값(0)'으로 채우기
    prev_value = None  # 이전 행의 괄호 안 값 저장
    
    for idx, row in df.iterrows():
        current_value = extract_bracket_value(row['적요'])  # 현재 행의 괄호 안 값 추출
        # 조건: 괄호 안 값에 ","와 ":"가 모두 포함된 경우
        if current_value and '"' in current_value and ':' in current_value:
            # 이전 값으로 대체
            if prev_value:
                df.loc[idx, '적요'] = re.sub(r'\(.*?\)', f"({prev_value})", row['적요'])
        else:
            # 유효한 값인 경우 이전 값 갱신
            if current_value:
                prev_value = current_value

# 함수 실행
process_column(df31)

# 함수: 적요열에서 끝에서 첫번째 글자를 제외하고 7번째 글자부터 7개 문자를 추출하고 숫자인지 확인
def extract_and_check_numeric(text):
    if isinstance(text, str) and len(text) >= 14:  # 7번째 글자부터 7개 문자를 추출하려면 길이가 최소 14이어야 함
        # 끝에서 첫번째 문자를 제외한 7번째 글자부터 7개 문자를 추출
        substring = text[-8:-1]
        if substring.isdigit():  # 숫자 여부 확인
            return substring
    return None  # 숫자가 아니면 None 반환

# '봉안함' 열을 새로 생성하여 값 저장
df31['봉안함'] = df31['적요'].apply(extract_and_check_numeric)







df31.info()

#df31.info()
#df31.to_excel(r'C:\dev\myproj01\금상동\결과폴더\사용료현금출납부엑셀정리-2024.xlsx', index=False)
output_file31=r'C:\dev\myproj01\금상동\결과폴더\사용료현금출납부엑셀정리-2024.xlsx'
df31.to_excel(output_file31, index=False)


df32= df32.dropna(subset=["봉안함"]) #아래 누적액,월계액을 없애기위해 회계계정없는부분을 삭제함
#df31.info()
#df31.to_excel(r'C:\dev\myproj01\금상동\결과폴더\사용료현금출납부엑셀정리-2024.xlsx', index=False)
output_file32=r'C:\dev\myproj01\금상동\결과폴더\봉안경당매일시재(계약선수금)엑셀정리-2024.xlsx'
df32.to_excel(output_file32, index=False)

df33= df33.dropna(subset=["봉안함"]) #아래 누적액,월계액을 없애기위해 회계계정없는부분을 삭제함
#df31.info()
#df31.to_excel(r'C:\dev\myproj01\금상동\결과폴더\사용료현금출납부엑셀정리-2024.xlsx', index=False)
output_file33=r'C:\dev\myproj01\금상동\결과폴더\봉안경당매일시재(봉안)엑셀정리-2024.xlsx'
df33.to_excel(output_file33, index=False)





print(f"병합된 데이터가 {output_file31}로 저장되었습니다.")
print(f"병합된 데이터가 {output_file32}로 저장되었습니다.")
print(f"병합된 데이터가 {output_file33}로 저장되었습니다.")

# 데이터프레임과 시트 이름 리스트
dfs = [df31,df32,df33]

sheet_names = ['사용료현금출납부', '봉안경당매일시재(계약선수금)','봉안경당매일시재(봉안)']
# 함수 호출
#export_multiple_dfs_to_excel(dfs, sheet_names,'C:\dev\myproj01\금상동\결과폴더\월별결산부사용선수금-관리선수금-사용료수입-회전관리선수금2024.xlsx')
output_file34=r'C:\dev\myproj01\금상동\결과폴더\천호성지엑셀통합.xlsx'
export_multiple_dfs_to_excel(dfs, sheet_names,output_file34)
print(f"병합된 데이터가 {output_file34}로 저장되었습니다.")

# 엑셀 파일 경로
file_path35 = r'C:\dev\myproj01\금상동\결과폴더\천호성지엑셀통합.xlsx'

# 모든 시트를 데이터프레임으로 불러오기
sheets = pd.read_excel(file_path35, sheet_name=None)  # 모든 시트를 딕셔너리로 불러옴

# 각 시트 처리
processed_sheets = {}
for sheet_name, df in sheets.items():
    if '월일' in df.columns:  # 월일 열이 존재하는 경우에만 처리
        # 빈 문자열("")을 NaN으로 변환
        df['월일'] = df['월일'].replace('"', pd.NA)

        
        # '월일' 열에 대해 결측값을 앞의 값으로 채우는 방법
        df.loc[:, '월일'] = df['월일'].ffill()
        
        
        df['월일'] = df['월일'].apply(lambda x: f"2024-{x.replace(',', '-')}")
        # '월일' 열에서 '월' 또는 '월일'이 포함된 행 삭제
        df = df[~df['월일'].str.contains('월|월일', na=False)]
    if '계정' in df.columns:  # 계정 열이 존재하는 경우에만 처리
        # DataFrame 복사본 생성 (필요한 경우)
        df = df.copy()

        # '계정' 열에서 '"'를 pd.NA로 대체
        df.loc[:, '계정'] = df['계정'].replace('"', pd.NA)

        # 결측값을 앞의 값으로 채우기
        df.loc[:, '계정'] = df['계정'].ffill()
        
        
        
        
        
        # 빈 문자열("")을 NaN으로 변환
        df.loc[:, '계정'] = df['계정'].replace('"', pd.NA)
        df.loc[:, '계정'] = df['계정'].ffill()

    # 2. 월일 열의 값을 차례대로 변환
    # 각 값에 대해 "YYYY-MM-DD" 형식으로 변경
         
    

    



    # 처리된 시트를 저장
    processed_sheets[sheet_name] = df

# 처리된 데이터프레임을 엑셀로 저장
output_path = r'C:\dev\myproj01\금상동\결과폴더\천호성지엑셀통합일자채움.xlsx'
with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
    for sheet_name, df in processed_sheets.items():
        df.to_excel(writer, index=False, sheet_name=sheet_name)

print("모든 시트의 '월일' 열을 처리하여 저장했습니다.")



# 빈 DataFrame을 올바르게 초기화
#df37 = pd.DataFrame(columns=['월일', '구분1', '코드1', '계정코드2', '코드4', '거래처명', '코드5', '적요명', '금액', '봉안함'])

# 데이터프레임 초기화 (데이터 타입 명시)
df37 = pd.DataFrame({
    '월일': pd.Series(dtype='str'),
    '구분1': pd.Series(dtype='int'),
    '코드1': pd.Series(dtype='str'),
    '계정코드2': pd.Series(dtype='str'),
    '코드4': pd.Series(dtype='str'),
    '거래처명': pd.Series(dtype='str'),
    '코드5': pd.Series(dtype='str'),
    '적요명': pd.Series(dtype='str'),
    '금액': pd.Series(dtype='float'),
    '봉안함': pd.Series(dtype='str'),
})




# df32의 절대값이 0보다 큰 경우 필터링

df34 = pd.read_excel(r'C:\dev\myproj01\금상동\결과폴더\천호성지엑셀통합일자채움.xlsx',sheet_name="봉안경당매일시재(계약선수금)",skiprows=0,usecols="a:g")
df35 = pd.read_excel(r'C:\dev\myproj01\금상동\결과폴더\천호성지엑셀통합일자채움.xlsx',sheet_name="봉안경당매일시재(봉안)",skiprows=0,usecols="a:g")

# '사용료' 열을 숫자형으로 변환 (문자열이나 비정상 데이터는 NaN 처리)
df34['사용료'] = pd.to_numeric(df34['사용료'], errors='coerce')
df35['사용료'] = pd.to_numeric(df35['사용료'], errors='coerce')
df35['관리비'] = pd.to_numeric(df35['관리비'], errors='coerce')

 # 빈 문자열("")을 NaN으로 변환
df34.loc[:, '봉안함'] = df34['봉안함'].replace('"', pd.NA)
df34.loc[:, '봉안함'] = df34['봉안함'].ffill()

df35.loc[:, '봉안함'] = df35['봉안함'].replace('"', pd.NA)
df35.loc[:, '봉안함'] = df35['봉안함'].ffill()





mask = df34['사용료'].abs() > 0
#  사용선수금 기타엑셀업로드에 추가 조건에 따라 행 추가
# 필터링된 데이터 반복 처리
for _, row in df34[mask].iterrows():
    new_row1 = {
        '월일': row['월일'],          # 날짜 열에서 값 가져오기
        '구분1': 3,
        '코드1': '108',
        '계정코드2': '외상매출금',
        '코드4': None,               # 필요시 수정
        '거래처명': None,            # 필요시 수정
        '코드5': None,               # 필요시 수정
        '적요명': row['비고'],       # 성명 열에서 값 가져오기
        '금액': abs(row['사용료']),         # 대변 열에서 값 가져오기
        '봉안함': row['봉안함']      # 봉안함 열에서 값 가져오기
    }
   
    df37 = pd.concat([df37, pd.DataFrame([new_row1])], ignore_index=True)   
   
    new_row2 = {
        '월일': row['월일'],          # 날짜 열에서 값 가져오기
        '구분1': 4,
        '코드1': '273',
        '계정코드2': '사용선수금',
        '코드4': None,               # 필요시 수정
        '거래처명': None,            # 필요시 수정
        '코드5': None,               # 필요시 수정
        '적요명': row['비고'],       # 성명 열에서 값 가져오기
        '금액': abs(row['사용료']),         # 대변 열에서 값 가져오기
        '봉안함': row['봉안함']      # 봉안함 열에서 값 가져오기
    }


    df37 = pd.concat([df37, pd.DataFrame([new_row2])], ignore_index=True)


mask2 = df35['사용료'] > 0
#  사용선수금 기타엑셀업로드에 추가 조건에 따라 행 추가
# 필터링된 데이터 반복 처리
for _, row in df35[mask2].iterrows():
    new_row1 = {
        '월일': row['월일'],          # 날짜 열에서 값 가져오기
        '구분1': 3,
        '코드1': '273',
        '계정코드2': '사용선수금',
        '코드4': None,               # 필요시 수정
        '거래처명': None,            # 필요시 수정
        '코드5': None,               # 필요시 수정
        '적요명': row['비고'],       # 성명 열에서 값 가져오기
        '금액': abs(row['사용료']),         # 대변 열에서 값 가져오기
        '봉안함': row['봉안함']      # 봉안함 열에서 값 가져오기
    }
   
    df37 = pd.concat([df37, pd.DataFrame([new_row1])], ignore_index=True)   
   
    new_row2 = {
        '월일': row['월일'],          # 날짜 열에서 값 가져오기
        '구분1': 4,
        '코드1': '420',
        '계정코드2': '사용료수입',
        '코드4': None,               # 필요시 수정
        '거래처명': None,            # 필요시 수정
        '코드5': None,               # 필요시 수정
        '적요명': row['비고'],       # 성명 열에서 값 가져오기
        '금액': abs(row['사용료']),         # 대변 열에서 값 가져오기
        '봉안함': row['봉안함']      # 봉안함 열에서 값 가져오기
    }


    df37 = pd.concat([df37, pd.DataFrame([new_row2])], ignore_index=True) 
    
    
    
mask3 = df35['관리비'] > 0
#  사용선수금 기타엑셀업로드에 추가 조건에 따라 행 추가
# 필터링된 데이터 반복 처리
for _, row in df35[mask3].iterrows():
    new_row1 = {
        '월일': row['월일'],          # 날짜 열에서 값 가져오기
        '구분1': 3,
        '코드1': '108',
        '계정코드2': '외상매출금',
        '코드4': None,               # 필요시 수정
        '거래처명': None,            # 필요시 수정
        '코드5': None,               # 필요시 수정
        '적요명': row['비고'],       # 성명 열에서 값 가져오기
        '금액': abs(row['관리비']),         # 대변 열에서 값 가져오기
        '봉안함': row['봉안함']      # 봉안함 열에서 값 가져오기
    }
   
    df37 = pd.concat([df37, pd.DataFrame([new_row1])], ignore_index=True)   
   
    new_row2 = {
        '월일': row['월일'],          # 날짜 열에서 값 가져오기
        '구분1': 4,
        '코드1': '274',
        '계정코드2': '관리선수금',
        '코드4': None,               # 필요시 수정
        '거래처명': None,            # 필요시 수정
        '코드5': None,               # 필요시 수정
        '적요명': row['비고'],       # 성명 열에서 값 가져오기
        '금액': abs(row['관리비']),         # 대변 열에서 값 가져오기
        '봉안함': row['봉안함']      # 봉안함 열에서 값 가져오기
    }


    df37 = pd.concat([df37, pd.DataFrame([new_row2])], ignore_index=True) 
    











# 데이터가 올바르게 추가되었는지 확인
if df37.empty:
    print("df37이 비어 있습니다. 데이터 추가를 확인하세요.")
else:
    print("df37에 데이터가 성공적으로 추가되었습니다.")    

output_file34=r'C:\dev\myproj01\금상동\결과폴더\천호엑셀기타업로드-2024.xlsx'
df37['적요명'] = df37['적요명'].astype(str) + "**" + df37['봉안함'].astype(str)   
    

#df37.to_excel(output_file34, index=False)








#__________________________________________________________________________________________________________






# 2. 데이터프레임을 엑셀 파일로 저장 (임시 저장)
output_file = 'C:\dev\myproj01\금상동\결과폴더\천호엑셀기타업로드-2024.xlsx'  # 여기에 저장할 엑셀 파일의 경로를 입력하세요.
df37.to_excel(output_file, index=False)

# 3. openpyxl을 사용하여 엑셀 파일 열기
wb = load_workbook(output_file)
ws = wb.active

# 4. 열 너비 설정
ws.column_dimensions['D'].width = 20  # '구분1' 열 (가정: A열에 위치)
ws.column_dimensions['H'].width = 25  # '적요명' 열 (가정: B열에 위치)

# 5. 엑셀 파일 저장
wb.save(output_file)

#

# 엑셀 파일 열기
#file_path = r'C:\dev\myproj01\금상동\결과폴더\월별결산부사용선수금-관리선수금-사용료수입-회전관리선수금2024.xlsx'  # 엑셀 파일 경로를 입력하세요
file_path = r'C:\dev\myproj01\금상동\결과폴더\천호성지엑셀통합일자채움.xlsx'  # 엑셀 파일 경로를 입력하세요
wb = openpyxl.load_workbook(file_path)

# 모든 시트에 대해 A열부터 D열까지의 너비를 25로 설정
columns_to_resize = ['A', 'B', 'C', 'D','E','F','G']  # A열부터 D열까지 지정

for sheet in wb.worksheets:
    for column_letter in columns_to_resize:
        # 지정된 열(column_letter)의 너비를 25로 설정
        sheet.column_dimensions[column_letter].width = 15



# 통화 기호 없는 회계 스타일 정의
accounting_style = NamedStyle(name="accounting_style", number_format='#,##0;[Red](#,##0)')

# 모든 시트에 대해 E열부터 G열까지 회계 스타일 적용
for sheet in wb.worksheets:
    for row in range(1, sheet.max_row + 1):
        for col in ['E', 'F', 'G']:  # E열부터 G열까지
            cell = sheet[f'{col}{row}']
            cell.style = accounting_style








# 수정된 엑셀 파일 저장

# 수정된 엑셀 파일 저장 (경로와 파일명 수정)
#save_path = r'C:\dev\myproj01\금상동\결과폴더\월별결산부사용선수금-관리선수금-사용료수입-회전관리선수금(열조정2024).xlsx'
save_path = r'C:\dev\myproj01\금상동\결과폴더\천호성지엑셀통합일자채움.xlsx'
wb.save(save_path)



import pandas as pd
from openpyxl import load_workbook

# 파일 경로 설정
#file_path = r'C:\dev\myproj01\금상동\결과폴더\월별결산부사용선수금-관리선수금-사용료수입-회전관리선수금(열조정2024).xlsx'  # 여기에 엑셀 파일 경로 입력
file_path = r'C:\dev\myproj01\금상동\결과폴더\천호성지엑셀통합일자채움.xlsx'

# 엑셀 파일 불러오기 (openpyxl 사용)
wb = load_workbook(file_path)
#sheet_names = ['사용선수금', '일반관리선수-최준철세무사', '사용료수입', '회전관리선수금']  # 작업할 시트 이름들
sheet_names = ['사용료현금출납부', '봉안경당매일시재(계약선수금)','봉안경당매일시재(봉안)']  # 작업할 시트 이름들

#'사용선수금', '일반관리선수-최준철세무사','사용료수입','회전관리선수금','월별결산부','사용선수금-일자순','일반관리선수-최준철세무사-일자순','사용료수입-일자순','회전관리선수금-일자순','회전관리선수금누적(봉안함순)'
# 각 시트에 대해 작업 수행
for sheet_name in sheet_names:
    # pandas로 시트를 읽어오기
    df21 = pd.read_excel(file_path, sheet_name=sheet_name)
    
    if '시트명' not in df21.columns:
        df21['시트명'] = pd.NA

       # '계정과목'의 결측치 처리: 결측치(NaN)인 경우 시트 이름으로 채움
        df21['시트명'] = df21['시트명'].fillna(sheet_name)
    



    # "계정과목" 열을 추가하고 시트 이름으로 채우기
    #df21['계정과목'] = sheet_name
    
    # pandas로 수정된 데이터를 다시 엑셀 파일로 저장
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df21.to_excel(writer, sheet_name=sheet_name, index=False)

print("모든 시트에 '시트명' 열 추가 완료.")

import pandas as pd

# 파일 경로 설정
#a_file = r'C:\dev\myproj01\금상동\결과폴더\금상동1205_2024.xlsx'  # a파일 경로
a_file = f"C:\\dev\\myproj01\\금상동\\결과폴더\\천호성지1205_{input_year}.xlsx"  # a파일 경로

   #f"C:\\dev\\myproj01\\금상동\\결과폴더\\금상동1205_{input_year}.xlsx"
#b_file = r'C:\dev\myproj01\금상동\결과폴더\월별결산부사용선수금-관리선수금-사용료수입-회전관리선수금(열조정2024).xlsx'  # b파일 경로
b_file = r'C:\dev\myproj01\금상동\결과폴더\천호성지엑셀통합일자채움.xlsx'   # b파일 경로

# 1. a 파일 불러와서 데이터프레임 a 생성
df_a = pd.read_excel(a_file)
#df_a = df30
#df_b = pd.read_excel(b_file)
# 2. b 파일의 첫 번째 시트부터 네 번째 시트까지 데이터프레임 생성
b_sheets = [pd.read_excel(b_file, sheet_name=i) for i in range(3)]

# 3. 첫 번째 시트를 a 데이터프레임에 병합하여 a-1 데이터프레임 생성
df_a1 =df_a


# 4. 두 번째, 세 번째, 네 번째 시트를 차례로 병합
for i in range(2, 3):
    b_sheets[i] = b_sheets[i].drop_duplicates(subset=['봉안함'])  # 봉안함 중복 제거 
    df_a1 = pd.merge(df_a1, 
                     b_sheets[i][['봉안함', '사용료', '관리비', '시트명','월일']], 
                     on='봉안함', 
                     how='left',
                     suffixes=('', f'_b{i}'))


# 4. 병합된 결과를 엑셀 파일로 저장
# 날짜를 문자열로 포맷 (예: '2024-10-03')
today = datetime.today().strftime('%Y-%m-%d')

output_file = f'C:\dev\myproj01\금상동\결과폴더\천호성지봉안+기타엑셀업로드병합결과-{today}.xlsx'


df_a1=df_a1.rename(columns={"대변_b4":"누적-회전관리선수금"})
df_a1=df_a1.rename(columns={"대변_b3":"2024-회전관리선수금"})
df_a1=df_a1.rename(columns={"차변_b2":"사용료수입"})

df_a1['봉안일1'] = df_a1['봉안일1'].dt.strftime('%Y-%m-%d')
df_a1['봉안일2'] = df_a1['봉안일2'].dt.strftime('%Y-%m-%d')
df_a1['만기일1'] = df_a1['만기일1'].dt.strftime('%Y-%m-%d')
df_a1['만기일2'] = df_a1['만기일2'].dt.strftime('%Y-%m-%d')
df_a1['관리시작'] = df_a1['관리시작'].dt.strftime('%Y-%m-%d')
df_a1['관리종료'] = df_a1['관리종료'].dt.strftime('%Y-%m-%d')


#df_a1['관리수입-누적회전관리선수금']=df_a1['관리수입'] - df_a1['누적-회전관리선수금']

 # 벡터화된 조건 사용
df_a1['연중관리수입2'] = 0  # 기본값 0으로 초기화
df_a1['연중관리수입2'] = df_a1['연중관리수입2'].astype(float)
mask3 = df_a1['관리수입'].notna() 
df_a1.loc[mask3, '연중관리수입2'] = (
    df_a1['관리수입'] / df_a1['관리선수금전체월수'] * df_a1['년중월수'])
    
#mask4 = df_a1['관리수입'] > df_a1['누적-회전관리선수금'] 
#mask4 = (df_a1['관리수입'].notna() & df_a1['누적-회전관리선수금'].notna() & (df_a1['관리수입'] > df_a1['누적-회전관리선수금']))
#df_a1.loc[mask4, '연중관리수입2'] = (
#    (df_a1['관리수입'])/df_a1['관리선수금전체월수']*df_a1['년중월수'])     
#df_a1['관리수입'] <= df_a1['누적-회전관리선수금']:
    #df_a1['연중관리비수입']=(df_a1['관리수입']/df_a1['관리선수금전체월수'])*df_a1['년중월수'] 
    #df_a1['연중관리비수입']=df_a1['관리수입'] 
    #df_a1['연중관리비수입2']=(df_a1['관리수입']/df_a1['관리선수금전체월수'])*df_a1['년중월수'] 
#df_a1['연중관리수입']=(df_a1['관리수입']-df_a1['누적-회전관리선수금'])/df_a1['관리선수금전체월수']*df_a1['년중월수']
#df_a1['연중회전관리수입']=(df_a1['누적-회전관리선수금'])/df_a1['회전관리선수금전체월수']*df_a1['년중회전관리선수금월수']

#df_a1['연중회전관리수입'] = pd.to_numeric(df_a1['연중회전관리수입'], errors='coerce').fillna(0).astype(int)
df_a1['연중관리수입2'] = pd.to_numeric(df_a1['연중관리수입2'], errors='coerce').fillna(0).astype(int)





df_a1.to_excel(output_file, index=False)

print(f"병합된 데이터가 {output_file}로 저장되었습니다.")


# 엑셀 파일 열기
file_path = f'C:\dev\myproj01\금상동\결과폴더\천호성지봉안+기타엑셀업로드병합결과-{today}.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb.active  # 활성화된 시트를 선택
# 모든 시트에 대해 A열부터 D열까지의 너비를 25로 설정
columns_to_resize = ['A', 'B', 'C', 'D','E','F','G','Y', 'Z','AA', 'AB','AC','AD','AH', 'AI','AJ', 'AK','AL','AM','AN']  # A열부터 D열까지 지정

for sheet in wb.worksheets:
    for column_letter in columns_to_resize:
        # 지정된 열(column_letter)의 너비를 25로 설정
        sheet.column_dimensions[column_letter].width = 15



       # 통화 기호 없는 회계 스타일 정의
        accounting_style = NamedStyle(name="accounting_style", number_format='#,##0;[Red](#,##0)')

# 모든 시트에 대해 E열부터 G열까지 회계 스타일 적용
    for sheet in wb.worksheets:
      for row in range(1, sheet.max_row + 1):
        for col in ['Y', 'Z','AA', 'AB','AC','AD','AH', 'AI','AJ', 'AK','AL','AM','AN']:  # E열부터 G열까지
            cell = sheet[f'{col}{row}']
            cell.style = accounting_style

    
    # '전체월수' 열의 값이 180을 초과하는 경우 색상을 빨간색으로 변경
    fill_red = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

    # 전체월수 열의 값을 차례대로 불러와서 처리
    for idx, value in enumerate(df_a1['관리선수금전체월수'], start=2):  # start=2는 헤더를 건너뛰기 위해
        if value > 180:
            ws[f'AE{idx}'].fill = fill_red  # C열에 '전체월수'가 있다고 가정
    
    
    
    
    # 수정된 엑셀 파일 저장
    # A2 셀 기준으로 틀 고정 (freeze panes)
    ws.freeze_panes = 'A2'

    # 첫 번째 행에 필터 추가
    ws.auto_filter.ref = ws.dimensions  # ws.dimensions는 사용 중인 셀의 범위를 자동으로 가져옴

    # 수정된 엑셀 파일 저장 (경로와 파일명 수정)
    
    today = datetime.today().strftime('%Y-%m-%d')

    
    
    save_path = f'C:\dev\myproj01\금상동\결과폴더\천호성지봉안+기타엑셀업로드병합결과-{today}.xlsx' 
    wb.save(save_path)
    print(f"새로운 데이터가 {save_path} 파일로 저장되었습니다.")        

    # 첫 번째 파일 (기존 파일)
file1 = r'C:\dev\myproj01\금상동\결과폴더\천호엑셀기타업로드-2024(열조정2024).xlsx'

# 두 번째 파일 (추가할 파일)
file2 = f'C:\dev\myproj01\금상동\결과폴더\천호성지봉안+기타엑셀업로드병합결과-{today}.xlsx'

# 파일을 읽어들임
dfs_file2 = pd.read_excel(file2, sheet_name=None)  # 두 번째 파일의 모든 시트를 딕셔너리로 읽기

# 엑셀 파일의 기존 워크북 불러오기
with pd.ExcelWriter(file1, engine="openpyxl", mode="a") as writer:
    for sheet_name, df in dfs_file2.items():  # 두 번째 파일의 시트들 순회
        df.to_excel(writer, sheet_name="엑셀통합시트", index=False)  # 새로운 시트로 추가

print(f"'{file2}' 파일의 모든 시트가 '{file1}' 파일에 추가되었습니다.")

#####################################################################################
####기타엑셀업로드에는 없는 사용료수입을 df_a의 사용료 수입을 인식하기 위한단계계
import numpy as np

# 데이터프레임 로드
df_a = pd.read_excel(f'C:\\dev\\myproj01\\금상동\\결과폴더\\천호성지봉안+기타엑셀업로드병합결과-{today}.xlsx'
)
df47 = pd.read_excel(r'C:\dev\myproj01\금상동\결과폴더\천호엑셀기타업로드-2024(열조정2024).xlsx')
# '봉안함' 열의 데이터 유형을 문자열(object)로 변환
df_a['봉안함'] = df_a['봉안함'].astype(str)  # 첫 번째 데이터프레임
df47['봉안함'] = df47['봉안함'].astype(str)  # 두 번째 데이터프레임


# 2024-사용료수입이 0보다 큰 행 필터링
df_a_filtered = df_a[df_a['2024-사용료수입'] > 0].copy()

# 봉안일 선택 (봉안일1, 봉안일2 중 input_year 포함)
df_a_filtered['봉안일1'] = df_a_filtered['봉안일1'].fillna('')
df_a_filtered['봉안일2'] = df_a_filtered['봉안일2'].fillna('')

df_a_filtered['selected_date'] = np.where(
    df_a_filtered['봉안일1'].str.contains(str(input_year)),
    df_a_filtered['봉안일1'],
    np.where(
        df_a_filtered['봉안일2'].str.contains(str(input_year)),
        df_a_filtered['봉안일2'],
        '2024-12-31'  # 기본값
    )
)

# 결과 저장
df_a_filtered.to_excel('C:\\dev\\myproj01\\금상동\\결과폴더\\df_a_filtered.xlsx', index=False)

# 추가 데이터 처리
#df57 = df47[df47['계정코드2'] == '사용료수입'] & ([df47['금액'] > 0)]
df57 = df47[(df47['계정코드2'] == '사용료수입') & (df37['금액'] > 0)]

df_a57 = df_a[df_a['2024-사용료수입'].notna()]

# '봉안함' 기준으로 그룹화
df_grouped57 = df57.groupby('봉안함', as_index=False).agg({'금액': 'sum'})
df_a_grouped57 = df_a57.groupby('봉안함', as_index=False).agg({'2024-사용료수입': 'sum'})

# 데이터프레임 복사
df_result57 = df_grouped57.copy()
df_a_result57 = df_a_grouped57.copy()

# 저장
df_grouped57.to_excel('C:\\dev\\myproj01\\금상동\\결과폴더\\df_result57.xlsx', index=False)
df_a_grouped57.to_excel('C:\\dev\\myproj01\\금상동\\결과폴더\\df_a_result57.xlsx', index=False)

# 병합
merged_df = pd.merge(
    df_a_result57, df_result57,
    on='봉안함', how='outer', suffixes=('_a', '_r')
)
merged_df.to_excel('C:\\dev\\myproj01\\금상동\\결과폴더\\사용료수입&봉안사용료수입머지.xlsx', index=False)
# 금액 차이 계산
#merged_df['금액_차이'] = max(0,(merged_df['2024-사용료수입'].fillna(0) - merged_df['금액'].fillna(0)))
# 금액 차이를 계산할 때 max 함수를 벡터화된 연산으로 변경
merged_df['금액_차이'] = np.maximum(0, merged_df['2024-사용료수입'].fillna(0) - merged_df['금액'].fillna(0))

# 금액 차이가 있는 행 필터링
diff_rows = merged_df[merged_df['금액_차이'] != 0]

# 수정된 행 추가
for _, row in diff_rows.iterrows():
    # 봉안일 데이터 가져오기
    matching_rows = df_a_filtered[df_a_filtered['봉안함'] == row['봉안함']]
    if not matching_rows.empty:
        selected_date = matching_rows['selected_date'].iloc[0]
    else:
        selected_date = '2024-12-31'  # 기본값

    # 사용선수금 행 추가
    new_row1 = pd.DataFrame({
        '월일': [selected_date],  # 동적으로 설정된 날짜
        '구분1': [3],
        '코드1': ['273'],
        '계정코드2': ['사용선수금'],
        '코드4': [None],
        '거래처명':[None],
        '코드5': [None],
        '적요명':[None],
        '금액': [row['금액_차이']],
        '봉안함': [row['봉안함']]
    })

    # 사용료수입 행 추가
    new_row2 = pd.DataFrame({
        '월일': [selected_date],  # 동적으로 설정된 날짜
        '구분1': [4],
        '코드1': ['420'],
        '계정코드2': ['사용료수입'],
        '코드4': [None],
        '거래처명': [None],
        '코드5': [None],
        '적요명': [None],
        '금액': [row['금액_차이']],
        '봉안함': [row['봉안함']]
    })

    # 새로운 행 추가
    df47 = pd.concat([df47, new_row1, new_row2], ignore_index=True)
# '봉안함'을 기준으로 두 데이터프레임 병합
merged_df = pd.merge(
    df47, 
    df_a[['봉안함', '계약자명']], 
    on='봉안함', 
    how='left', 
    suffixes=('', '_from_a')
)

# '적요명'이 NaN인 경우 '계약자명'과 '봉안함'을 결합하여 채움
merged_df['적요명'] = merged_df['적요명'].fillna(
    merged_df['계약자명'] + ' & ' + merged_df['봉안함'].astype(str)
)

# 결과를 df47에 다시 저장
df47 = merged_df.drop(columns=['계약자명'])


# 최종 결과 저장
df47.to_excel('C:\\dev\\myproj01\\금상동\\결과폴더\\천호성지세무사랑기타업로드+봉안사용료수입추가1225-8.xlsx', index=False)

# 차변,대변 출력후 한칸씩 띄우는 작업  대차차액 발생을 방지하기위함함
# 엑셀 워크북 생성
# 새로운 데이터프레임을 생성할 리스트
new_rows = []

# 이전 행의 값 초기화
previous_value = None

# 데이터프레임을 순회하며 조건에 맞는 행 추가
for index, row in df47.iterrows():
    # 현재 행 추가
    new_rows.append(row)
    
    # 현재 행의 '구분1' 값
    current_value = row['구분1']
    
    # 조건에 따라 빈 행 추가
    if (previous_value == 3 and current_value == 4) or (current_value == 2):
        new_rows.append(pd.Series([None] * len(row), index=row.index))  # 빈 행 추가
    
    # 이전 값 업데이트
    previous_value = current_value

# 새로운 데이터프레임 생성
updated_df47 = pd.DataFrame(new_rows)

# 엑셀 파일로 저장
updated_df47.to_excel('C:\\dev\\myproj01\\금상동\\결과폴더\\updated_df47.xlsx', index=False)
# 결과 저장
