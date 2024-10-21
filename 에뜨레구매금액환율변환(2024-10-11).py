import pandas as pd			
import openpyxl		

import datetime as dt
			
#엑셀파일의 모든 시트를 리스트로 가져오기			
sheet = openpyxl.load_workbook("C:\dev\myproj01\Q10 매출관련 자료 (2024-7-9월).xlsx").sheetnames			
print('시트갯수:', len(sheet))		
# 엑셀파일에서 데이터프레임을 읽은뒤 합치기			
df = pd.DataFrame()			
#for i in sheet:	
#"C:\dev\myproj01\Q10 매출관련 자료 (6).xlsx"

for i in range(1,4):
      #df1 = pd.read_excel("E16EXAMPLE.xlsx", sheet_name=i)
    df1 = pd.read_excel("C:\dev\myproj01\Q10 매출관련 자료 (2024-7-9월).xlsx", sheet_name=-i)			
    df1= df1.dropna(subset=["구매자결제일"])
   	
    df1["월"] = i  #시트이름의 날짜정보를 새로운 열로 생성			
    df = pd.concat([df, df1])				

			
#파일로 저장			
df['구매자결제일']=df['구매자결제일'].dt.strftime('%y-%m-%d') #구매자결제일이 날짜함수이므로 문자함수로 변환

df3 =pd.read_excel("C:\dev\myproj01\StdExRate(2024-7-9월).xlsx",sheet_name="StdExRate(7-9월)",skiprows=8,usecols="a:e")

df3=df3.rename(columns={"날짜":"구매자결제일"})
df3['구매자결제일']=df3['구매자결제일'].str.replace('.','-')

df3.info()
df2 = df[["구매자결제일","상품결제금"]] #상품결제일과 결제금액 열을 추출
df2['구매자결제일']=df2['구매자결제일'].str.replace('24-','2024-')
df4 = df3[["구매자결제일","환율"]]   #환율고시일과 환율을 추출

df6 = pd.concat([df2, df4])   #상품결제일과 환율고시일자를  일자를 기준으로 합침

df6.merge(df4,how="outer",on=["구매자결제일"])

df7 =df6.sort_values("구매자결제일")
print(df7)
df7.merge(df4,how="outer",on=["구매자결제일"])
df9=df7.sort_values(["구매자결제일","환율"]) # 일자와 환율을 일자와 환율로 정렬
#df9= df9.dropna(subset=["환율"])

df9.iloc[:,[0,2]].to_excel("c:/dev/myproj01/적용환율7-9월.xlsx")

# wget , openpyxl, time을  import 해줍니다.
#import wget # 다운로드
import openpyxl as xl # 엑셀읽고 쓰기
import time 
import pandas as pd		
import numpy as np 


xl_file_path = 'c:/dev/myproj01/적용환율7-9월.xlsx' # 엑셀파일이 있는 곳을 적어줍니다. 
#df10 =pd.read_excel("c:/dev/myproj01/적용환율for9.xlsx",sheet_name="Sheet1",skiprows=0,usecols="a:e")
df10=df9.iloc[:,[0,2]]
wb = xl.load_workbook(xl_file_path, data_only=True) # 엑셀파일을 열고 읽습니다.
ws = wb.worksheets[0] # 몇번째 시트에 해당 내용이 있는지에 따라 번호를 달리합니다. 저는 첫번째 시트, 0번째에 있습니다

df11= df10.drop_duplicates(subset=['구매자결제일','환율']) #결제일과 환율이 중복되는  행삭제

df12=df11.sort_values(["구매자결제일","환율"]) # 일자와 환율을 일자와 환율로 정렬
df12.drop_duplicates(subset=['구매자결제일','환율'],keep='first')  #일자로 먼저 정렬하고 환율값이 있는 열을 우선 추출후 나머지 삭제

df13=df12.fillna(method='ffill') # 최근일 이전환율을 채움
#
df14=df13.drop_duplicates(subset=['구매자결제일','환율'])


df16=df14.merge(df2,how="outer",on=["구매자결제일"])

df17= df16.dropna(subset=["상품결제금","환율"])

df17['매출액*환율']=df17['상품결제금']*df17['환율']

df17['통화명']='jpy'
df17['원화환산금액']=(round(df17['매출액*환율']/100))
df18=df17.sort_values(["구매자결제일"],ascending = True) # 일자로 정렬
df18.iloc[:,[0,1,2,3,4,5]].to_excel("c:/dev/myproj01/환율적용매출금액2024-07-9월.xlsx")


