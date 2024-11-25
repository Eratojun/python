import pandas as pd
from datetime import datetime
import pandas as pd			
from pathlib import Path
import openpyxl	
import pandas as pd

from openpyxl import load_workbook

import pandas as pd			
from pathlib import Path
import openpyxl	
from datetime import datetime
import pandas as pd			
from pathlib import Path
import openpyxl	
from openpyxl.styles import NamedStyle
from openpyxl.styles import PatternFill #엑셀셀을 빨간색으로 표시










# 엑셀 파일 불러오기
file_path = r'C:\dev\myproj01\통합조회_20241112천호성지.xlsx'  # 엑셀 파일 경로를 지정하세요
df = pd.read_excel(file_path)



# 날짜가 유효하지 않다면 2월 28일로 수정하는 함수
def fix_invalid_date(date_str):
    try:
        # 날짜가 유효한지 확인
        date_obj = pd.to_datetime(date_str)
    except ValueError:
        # 유효하지 않은 날짜인 경우 처리 (윤년이 아닌 해의 2월 29일 등)
        if "02-29" in date_str:
            # 날짜가 2월 29일인 경우 2월 28일로 수정
            date_str = date_str.replace("02-29", "02-28")
            date_obj = pd.to_datetime(date_str)
        else:
            date_obj = pd.NaT
    return date_obj

# '관리시작열'과 '관리종료열'의 날짜를 변환하며 오류를 수정
df['관리시작'] = df['관리시작'].apply(fix_invalid_date)
df['관리종료'] = df['관리종료'].apply(fix_invalid_date)




# 연도를 입력받음
input_year = int(input("연도를 입력하세요 (예: 2023): "))
start_of_input_year = datetime(input_year, 1, 1)
end_of_input_year = datetime(input_year, 12, 31)

#df['관리시작']=df['관리시작'].dt.strftime('%y-%m-%d') #구매자결제일이 날짜함수이므로 문자함수로 변환
#df['관리종료']=df['관리종료'].dt.strftime('%y-%m-%d') #구매자결제일이 날짜함수이므로 문자함수로 변환





# 전체월수 계산 함수2024
def calculate_total_months(row):
    start_date = pd.to_datetime(row['관리시작'])
    end_date = pd.to_datetime(row['관리종료'])
    
    # 월 수 계산 (종료일과 시작일의 차이를 기준으로)
    total_months = (end_date.year - start_date.year) * 12 + end_date.month - start_date.month
    return total_months   # 시작과 종료 포함

# 년중월수 계산 함수
def calculate_year_months(row):
    start_date = pd.to_datetime(row['관리시작'])
    end_date = pd.to_datetime(row['관리종료'])

    # 관리 시작 날짜가 입력된 연도보다 이전인 경우 처리
    if start_date < start_of_input_year:
        # 입력된 연도의 시작일부터 계산
        actual_start = start_of_input_year
    else:
        actual_start = start_date

    # 관리 종료 날짜와 입력된 연도의 종료일 중 빠른 날짜 선택
    actual_end = min(end_date, end_of_input_year)

    # 기간이 겹치지 않는 경우 0 반환
    if actual_end < actual_start:
        return 0

    # 월 수 계산
    months_in_year = (actual_end.year - actual_start.year) * 12 + actual_end.month - actual_start.month
    return months_in_year+1  # 시작과 종료 포함



# 연말잔여월수 계산 함수
# 입력된 연도의 종료일을 계산
end_of_input_year = datetime(input_year, 12, 31)

# 잔여월수 계산 함수
def calculate_remaining_months(row):
    end_date = pd.to_datetime(row['관리종료'])
    
    # 관리종료열이 입력된 연도의 종료일보다 빠른 경우 0 반환
    if end_date <= end_of_input_year:
        return 0
    
    # 관리종료열이 입력된 연도의 종료일보다 늦은 경우 잔여 월수 계산
    else:
        remaining_months = (end_date.year - end_of_input_year.year) * 12 + (end_date.month - end_of_input_year.month)
        return remaining_months + 1  # 종료 월을 포함한 월수 반환




# 3. 관리시작열에서 입력된 연도를 포함하는지 확인 후, 사용료수입 열에 값을 추가하거나 0으로 설정
def check_year(row):
    # 관리시작열에서 입력된 연도를 포함하면 사용료수입 값을 유지, 포함하지 않으면 0을 반환
    if str(input_year) in str(row['관리시작']):
        return row['사용수입']
    else:
        return 0

def check_year2(row):
    start_date = pd.to_datetime(row['관리시작'])
    end_date = pd.to_datetime(row['관리종료'])
    
    if start_date < start_of_input_year:
        # 입력된 연도의 시작일부터 계산
        return row['사용수입']
    else:
        return 0





# 전체월수열 추가
df['전체월수'] = df.apply(calculate_total_months, axis=1)









# 년중월수열 추가
df['년중월수'] = df.apply(calculate_year_months, axis=1)

# 잔여월수열 추가
df['잔여월수'] = df.apply(calculate_remaining_months, axis=1)
df['연중관리수입']=df['관리수입']/df['전체월수']*df['년중월수']
df['잔여관리선수금']=df['관리수입']/df['전체월수']*df['잔여월수']
# '사용료수입' 열을 업데이트
df['사용료수입'] = df.apply(check_year, axis=1)
df['사용료경과분'] = df.apply(check_year2, axis=1)



# 결과를 엑셀 파일로 저장
output_path = r'C:\dev\myproj01\전체월수연중월수.xlsx'  # 결과 파일 경로를 지정하세요
df.to_excel(output_path,sheet_name=str(input_year), index=False)
df.info()
print("엑셀 파일이 성공적으로 저장되었습니다.")
# 엑셀 파일 경로 지정
file_path = r'C:\dev\myproj01\전체월수연중월수.xlsx'  # 결과 파일 경로를 지정하세요
df = pd.read_excel(file_path)
df['관리시작'] = pd.to_datetime(df['관리시작'], errors='coerce')
df['관리시작'] = df['관리시작'].dt.strftime('%Y-%m-%d')
df['관리종료'] = pd.to_datetime(df['관리종료'], errors='coerce')
df['관리종료'] = df['관리종료'].dt.strftime('%Y-%m-%d')
# 사용자로부터 연도 입력받기
input_year = input("연도를 입력하세요 (예: 2023): ")

# 계약일에 입력한 연도 포함 여부를 확인하는 함수
def contains_year(date_series, year):
    # 계약일이 NaN이 아닌 경우에만 문자열 변환 후 year 포함 여부 확인
    return date_series.dropna().astype(str).str.contains(year)

# 조건: 입력받은 연도가 2023인 경우
if input_year == input_year:
    # 새로운 열 이름을 생성: '2023-사용선수금'
    new_column_name = f"{input_year}-사용선수금"
    new_column_name2 = f"{input_year}-관리선수금"
    new_column_name3 = f"{input_year}-사용료수입"
    
    # 새로운 열을 데이터프레임에 추가하고 초기값은 None으로 설정
    df[new_column_name] = None
    df[new_column_name2] = None
    df[new_column_name3] = None
    
    # '계약일' 열에 입력받은 연도를 포함하고, 나머지 열이 모두 결측치일 경우 마스크 정의
    mask = (contains_year(df['계약일'], input_year)) & \
           (df['봉안일1'].isna()) & \
           (df['봉안일2'].isna()) & \
           (df['관리시작'].isna()) & \
           (df['관리종료'].isna())

    # '봉안일1' 열에 입력받은 연도를 포함하고, 나머지 열이 모두 결측치일 경우 마스크 정의
    mask2 = (contains_year(df['관리시작'], input_year)) & \
            (df['사용수입'] > 0)


    # '사용선수' 값이 있는 행에서 '2023-사용선수금' 열에 값을 넣음
    df.loc[mask, new_column_name] = df.loc[mask, '사용선수']
    df.loc[mask, new_column_name2] = df.loc[mask, '관리선수']
    df.loc[mask2, new_column_name3] = df.loc[mask2, '사용수입']

    # 결과 확인
    print(df)
    df=df.sort_values(["봉안함"]) # 일자와 환율을 일자와 환율로 정렬
    # 새로운 파일로 저장
    output_file_path = f"C:\\dev\\myproj01\\금상동\\결과폴더\\천호성지_{input_year}.xlsx"
    
    df.to_excel(output_file_path, index=False)
    print(f"새로운 데이터가 {output_file_path} 파일로 저장되었습니다.")

    # 엑셀 파일 열기
    file_path = f"C:\\dev\\myproj01\\금상동\\결과폴더\\천호성지_{input_year}.xlsx"  # 엑셀 파일 경로를 입력하세요
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active  # 활성화된 시트를 선택
    # 모든 시트에 대해 A열부터 D열까지의 너비를 25로 설정
    columns_to_resize = ['A', 'B', 'C', 'D','E','F','G','Y', 'Z','AA', 'AB','AC','AD','AH', 'AI','AJ', 'AK','AL','AM','AN']  # A열부터 D열까지 지정

    for sheet in wb.worksheets:
     for column_letter in columns_to_resize:
        # 지정된 열(column_letter)의 너비를 25로 설정
        sheet.column_dimensions[column_letter].width = 15



       # 통화 기호 없는 회계 스타일 정의
        accounting_style = NamedStyle(name="accounting_style", number_format='#,##0;[Red](#,##0)')

    # 모든 시트에 대해 E열부터 G열까지 회계 스타일 적용
    for sheet in wb.worksheets:
      for row in range(1, sheet.max_row + 1):
        for col in ['Y', 'Z','AA', 'AB','AC','AD','AH', 'AI','AJ', 'AK','AL','AM','AN']:  # E열부터 G열까지
            cell = sheet[f'{col}{row}']
            cell.style = accounting_style

    
    # '전체월수' 열의 값이 180을 초과하는 경우 색상을 빨간색으로 변경
    fill_red = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

    # 전체월수 열의 값을 차례대로 불러와서 처리
    for idx, value in enumerate(df['전체월수'], start=2):  # start=2는 헤더를 건너뛰기 위해
        if value > 180:
            ws[f'AE{idx}'].fill = fill_red  # C열에 '전체월수'가 있다고 가정
    
    
    
    
    # 수정된 엑셀 파일 저장
    # A2 셀 기준으로 틀 고정 (freeze panes)
    ws.freeze_panes = 'A2'

    # 첫 번째 행에 필터 추가
    ws.auto_filter.ref = ws.dimensions  # ws.dimensions는 사용 중인 셀의 범위를 자동으로 가져옴

    # 수정된 엑셀 파일 저장 (경로와 파일명 수정)
    save_path = f"C:\\dev\\myproj01\\금상동\\결과폴더\\천호성지_{input_year}.xlsx" 
    wb.save(save_path)

else:
    print("입력한 연도가 2023이 아닙니다. 새로운 열이 추가되지 않았습니다.")